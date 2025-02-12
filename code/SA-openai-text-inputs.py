import os
import json
import openai
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import logging
import time
from datetime import datetime
import tenacity
import traceback
import tiktoken

# Add API stats tracking
class APIStats:
    def __init__(self):
        self.total_requests = 0
        self.total_input_tokens = 0
        self.total_output_tokens = 0
        
    def update(self, input_tokens, output_tokens):
        self.total_requests += 1
        self.total_input_tokens += input_tokens
        self.total_output_tokens += output_tokens
        
    @property
    def total_tokens(self):
        return self.total_input_tokens + self.total_output_tokens

api_stats = APIStats()

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def preprocess_ocr_text(text):
    # Replace common OCR errors
    replacements = {
        "Sv'cink i^idam": "Frank Adam",
        # Add more common OCR errors here
    }
    
    for error, correction in replacements.items():
        text = text.replace(error, correction)
    
    # Remove unusual characters
    text = re.sub(r'[^\w\s.,;:!?()-]', '', text)
    
    return text

def clean_json_string(json_str):
    # Remove any markdown code block indicators
    json_str = re.sub(r'```json\s*|\s*```', '', json_str)
    
    # Remove any leading/trailing whitespace
    json_str = json_str.strip()
    
    # Handle potential issues with quotes
    json_str = json_str.replace('\\"', '"').replace("'", '"')
    
    # Escape any unescaped double quotes within value strings
    json_str = re.sub(r'(?<!\\)\"([^\"]*)\"', lambda m: '"{}"'.format(m.group(1).replace('"', '\\"')), json_str)
    
    # Handle potential issues with newlines in values
    json_str = json_str.replace('\n', '\\n')
    
    # Remove any potential control characters
    json_str = ''.join(ch for ch in json_str if ord(ch) >= 32)
    
    return json_str

def num_tokens_from_string(string: str, model: str) -> int:
    encoding = tiktoken.encoding_for_model(model)
    return len(encoding.encode(string))

@tenacity.retry(
    wait=tenacity.wait_exponential(multiplier=1, min=4, max=10),
    stop=tenacity.stop_after_attempt(5),
    retry=tenacity.retry_if_exception_type(openai.RateLimitError)
)
def make_api_call(client, prompt, content, is_short):
    try:
        model = "gpt-4o-2024-08-06"  
        system_message = "You are an AI archival expert tasked with cleaning OCR text and extracting metadata from it."
        user_message = prompt + "\n\nHere's the content to analyze:\n\n" + content
        
        user_message += f"\n\nNote: This content is {'short (less than 250 characters)' if is_short else 'of normal length'}."

        input_tokens = num_tokens_from_string(system_message, model) + num_tokens_from_string(user_message, model)

        response = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system_message},
                {"role": "user", "content": user_message}
            ],
            max_tokens=3000
        )
        response_text = response.choices[0].message.content
        logging.info(f"Raw API response: {response_text}")

        output_tokens = num_tokens_from_string(response_text, model)
        
        # Update API stats
        api_stats.update(input_tokens, output_tokens)

        return response_text, input_tokens, output_tokens, input_tokens + output_tokens

    except Exception as e:
        logging.error(f"Error in make_api_call: {str(e)}")
        return None, 0, 0, 0
    
def calculate_totals(ws):
    input_total = sum(ws[f'I{i}'].value for i in range(2, ws.max_row + 1) if ws[f'I{i}'].value)
    output_total = sum(ws[f'J{i}'].value for i in range(2, ws.max_row + 1) if ws[f'J{i}'].value)
    total_total = sum(ws[f'K{i}'].value for i in range(2, ws.max_row + 1) if ws[f'K{i}'].value)
    return input_total, output_total, total_total

def postprocess_api_response(response_data):
    if 'namedEntities' in response_data:
        # Remove duplicates while preserving order
        response_data['namedEntities'] = list(dict.fromkeys(response_data['namedEntities']))
        
        # Remove any entities that are just single letters or numbers
        response_data['namedEntities'] = [entity for entity in response_data['namedEntities'] if len(entity) > 1 or not entity.isalnum()]
    
    # Ensure 'contentWarning' field exists and is properly formatted
    if 'contentWarning' not in response_data:
        response_data['contentWarning'] = 'None'
    elif response_data['contentWarning'].lower() == 'none' or response_data['contentWarning'].strip() == '':
        response_data['contentWarning'] = 'None'
    else:
        # Capitalize the first letter and ensure it ends with a period
        response_data['contentWarning'] = response_data['contentWarning'].capitalize().rstrip('.') + '.'
    
    return response_data

def process_file(file_path, client, folder_name, page_number):
    filename = os.path.basename(file_path)
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()

        # Preprocess the OCR text
        content = preprocess_ocr_text(content)
        
        # Check if the content is empty or only whitespace
        if not content.strip():
            return [
                folder_name,
                page_number,
                filename,
                "",
                "No readable text on this page",
                "",
                "N/A",
                "",  # Empty content warning
                0,
                0,
                0
            ]

        is_short = len(content.strip()) < 225  # Adjust this threshold as needed
        is_cover = page_number == 1  # Assume page 1 is always the cover

        ocr_cleaning_instructions = """
Before analyzing the content, please clean the OCR text according to these guidelines:
1. Correct misspellings that are clearly due to OCR errors.
2. Fix obvious punctuation errors.
3. Correct word splits (e.g., "architec ture" should be "architecture").
4. Do not alter proper names, even if they seem unusual, unless it's an obvious OCR error.
5. Do not add or remove content; your task is to clean, not to summarize or expand.
6. If you're unsure about a word or phrase, leave it as is.
7. Maintain the original formatting as much as possible.  Remember that it is a page from an old magazine.  Many pages begin with the name of the periodical, the date, and the page number.  This information may be edited slightly for clarity by, for example, separating the date from the page number.

Provide the cleaned text followed by '---' before proceeding with the metadata extraction.
"""

        content_warning_instructions = """
content_warning_instructions:
After analyzing the content, assess whether it contains any language or themes that might warrant a content warning. Consider the following categories:
1. Racist or ethnically insensitive language
2. Sexist or gender-discriminatory content
3. Violence or graphic descriptions
4. Other potentially offensive or outdated terminology

If you identify any such content, briefly describe it using one of the following phrases in a 'contentWarning' field:
- For outdated and/or racist terms: Some of the language in this resource uses outdated and/or racist terminology.
- For imagery and/or descriptions of violence: Some of the language and/or images in this resource depict graphic and violent content.
- For combinations of depictions of violence and outdated terms: Some of the language and/or images in this resource depict graphic and violent content and outdated and/or racist terminology.

If no warning is necessary, set this field to "None".

"""

        if is_cover:
            prompt_text = f"""{ocr_cleaning_instructions}
You are an archivist at the University of Texas at Austin, cataloging a collection of issues of The Southern Architect (1892-1931).
This page is the cover of an issue. The OCR text for this cover is:

{content.strip()}

After cleaning the OCR text, please create the following metadata fields in JSON format:
'tocEntry': Start with "Cover:" followed by a brief description of the cover content, based on the text.
'namedEntities': A list of key entities mentioned in the text, including any significant names, titles, or locations.
'subjectHeadings': Include "Magazine covers" and any other relevant subject headings based on the content.
{content_warning_instructions}

Provide the cleaned OCR text, followed by '---', then ONLY the JSON response, with no additional text before or after."""
        elif is_short:
            prompt_text = f"""{ocr_cleaning_instructions}
You are an archivist at the University of Texas at Austin, cataloging a collection of issues of The Southern Architect (1892-1931).
This page likely contains an image. Consequently, the OCR text for this image is short, containing only the following text:

{content.strip()}

After cleaning the OCR text, please create the following metadata fields in JSON format:
'tocEntry': If photography is mentioned in the text, start with "Photograph:"; if there is no mention of a photo, start with "Image:" followed by a brief description of what the image likely shows, based on the text.
'namedEntities': A list of key entities mentioned in the text, including architects, locations, and any other significant names.
'subjectHeadings': Include "Architectural photography" or if unsure of the type of image, "Architecture" and/or any other relevant subject headings based on the content.
{content_warning_instructions}

Provide the cleaned OCR text, followed by '---', then ONLY the JSON response, with no additional text before or after."""
        else:
            prompt_text = f"""{ocr_cleaning_instructions}
You are an archivist at the University of Texas at Austin, cataloging a collection of issues of The Southern Architect (1892-1931). 
Important: you are working from OCR text that may contain errors, so use your judgment to decipher meaning.  
Audience: you are creating this metadata for mainly architectural historians and architectural history students. Architectural styles, movements, trends, and other historically important information should be prioritized.  

After cleaning the OCR text, create the following metadata fields in JSON:
'tocEntry': One short phrase that succinctly describes the main content of the page. Always note the type of content that it is - for example, "Advertisement for an architecture firm" or "Article on the latest architectural trends in Dallas, Texas".
'namedEntities': A list of key entities directly relevant to the main content. Include significant people, companies, products, and locations central to the topic. Exclude lists of branch offices or minor mentions.
'subjectHeadings': Two to three Library of Congress Authoritative subject headings that best represent the main themes or topics of the page.
{content_warning_instructions}

Focus on capturing the essence of the page's content rather than exhaustive detail. 

Rules for identifying advertisements:
1. Look for company names prominently displayed
2. Check for product descriptions or benefits
3. Look for contact information or calls to action
4. Be aware of persuasive language or marketing slogans
5. Note any pricing information or special offers

Here's the OCR text to clean and analyze:

{content.strip()}

Provide the cleaned OCR text, followed by '---', then ONLY the JSON response, with no additional text before or after."""

        response_data, input_tokens, output_tokens, total_tokens = make_api_call(client, prompt_text, content, is_short or is_cover)
        
        if response_data:
            # Split the response into cleaned text and JSON data
            parts = response_data.split('---', 1)
            if len(parts) == 2:
                cleaned_content, json_data = parts
                cleaned_content = cleaned_content.strip()
                
                # Remove any markdown formatting from the JSON data
                json_data = re.sub(r'^```json\s*|\s*```$', '', json_data.strip(), flags=re.MULTILINE)
                
                try:
                    # Parse the JSON data
                    metadata = json.loads(json_data)

                    # Postprocess the API response
                    metadata = postprocess_api_response(metadata)
                    
                    if metadata.get('contentWarning', 'None') != 'None':
                        logging.info(f"Content warning identified in {filename}: {metadata['contentWarning']}")

                    return [
                        folder_name,
                        page_number,
                        filename,
                        cleaned_content,
                        metadata.get('tocEntry', 'N/A'),
                        ', '.join(metadata.get('namedEntities', [])),
                        ', '.join(metadata.get('subjectHeadings', [])),
                        metadata.get('contentWarning', 'None'),  # New field for content warning
                        input_tokens,
                        output_tokens,
                        total_tokens
                    ]
                except json.JSONDecodeError as e:
                    logging.error(f"JSON parsing error in {filename}: {str(e)}")
                    logging.error(f"Problematic JSON data: {json_data}")
                    return f"{folder_name}/{filename}: Error: JSON Parsing Error - {str(e)}"
            else:
                logging.error(f"Unexpected response format in {filename}")
                return f"{folder_name}/{filename}: Error: Unexpected response format"
        else:
            return f"{folder_name}/{filename}: Error: Failed to get valid response"

    except Exception as e:
        logging.error(f"Error processing file {filename}: {str(e)}")
        logging.error(f"Problematic content: {content[:500]}...")  # Log the first 500 characters of the content
        logging.error(traceback.format_exc())  # Log the full traceback
        return f"{folder_name}/{filename}: Error: Processing Error - {str(e)}"
    
def process_folder_recursive(folder_path, client, wb):
    ws = wb.active
    headers = ['Folder', 'Page Number', 'Page Title', 'Cleaned OCR Text', 'TOC Entry', 'Named Entities', 'Subject Headings', 'Content Warning', 'Input Tokens', 'Output Tokens', 'Total Tokens']
    ws.append(headers)

    # Set column widths
    column_widths = [15, 10, 30, 50, 30, 30, 30, 30, 15, 15, 15]  # Added width for new column
    for i, width in enumerate(column_widths):
        ws.column_dimensions[ws.cell(row=1, column=i+1).column_letter].width = width

    items_with_issues = []
    total_items = 0
    all_data = []

    def folder_sort_key(x):
        parts = x.split('-')
        return tuple(parts + [''] * (3 - len(parts)))

    # Sort folders
    sorted_folders = sorted(os.listdir(folder_path), key=folder_sort_key, reverse=False)

    for folder_name in sorted_folders:
        folder_full_path = os.path.join(folder_path, folder_name)
        if os.path.isdir(folder_full_path):
            # Sort files within the folder by page number
            files = sorted(
                [f for f in os.listdir(folder_full_path) if f.endswith('.txt')],
                key=lambda x: int(re.search(r'page(\d+)', x).group(1))
            )
            
            for filename in files:
                total_items += 1
                file_path = os.path.join(folder_full_path, filename)
                
                # Extract page number
                page_number = int(re.search(r'page(\d+)', filename).group(1))
                
                result = process_file(file_path, client, folder_name, page_number)
                if isinstance(result, str):  # Error occurred
                    items_with_issues.append(result)
                else:
                    all_data.append(result)

                time.sleep(1)  # Add a 1-second delay between API calls

    # Sort all_data by folder and page number
    all_data.sort(key=lambda x: (folder_sort_key(x[0]), int(x[1])), reverse=False)

    # Add sorted data to worksheet
    for row_data in all_data:
        ws.append(row_data)

    # Apply text wrapping and vertical alignment to all rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Calculate and add totals row
    input_total, output_total, total_total = calculate_totals(ws)
    ws.append(['Totals', '', '', '', '', '', '', '', input_total, output_total, total_total])
    
    # Style the totals row
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    # Create a new sheet for items with issues
    ws_issues = wb.create_sheet(title="Issues")
    ws_issues.append(["Filename"])
    for item in items_with_issues:
        ws_issues.append([item])

    logging.info(f"Processed {total_items} items. {len(items_with_issues)} items had issues.")
    return total_items, len(items_with_issues)

def main():
    start_time = time.time()
    
    logging.info("Starting the script")
    folder_path = "/Users/hannahmoutran/Desktop/ai-projects-ut-austin-fall-2024/southern_architect/southern-architect-page-level-file-folders-small copy"
    client = openai.OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))
    wb = Workbook()

    logging.info("Processing folder")
    total_items, items_with_issues = process_folder_recursive(folder_path, client, wb)

    wb.active.freeze_panes = 'E2' 
    
    current_date = datetime.now().strftime("%Y-%m-%d")
    output_file = f"SA-gpt-4o-text-inputs-{current_date}.xlsx"
    output_dir = "/Users/hannahmoutran/Desktop/ai-projects-ut-austin-fall-2024/southern_architect/Comparing Models - FINAL"
    os.makedirs(output_dir, exist_ok=True)
    full_output_path = os.path.join(output_dir, output_file)

    logging.info(f"Attempting to save file to {full_output_path}")
    try:
        wb.save(full_output_path)
        logging.info(f"Results saved to {full_output_path}")
    except Exception as e:
        logging.error(f"Error saving file: {str(e)}", exc_info=True)

    # Print API usage statistics
    print("\nAPI Usage Statistics:")
    print(f"Total API Requests: {api_stats.total_requests}")
    print(f"Total Input Tokens: {api_stats.total_input_tokens}")
    print(f"Total Output Tokens: {api_stats.total_output_tokens}")
    print(f"Total Tokens: {api_stats.total_tokens}")
    
    # Print execution time statistics
    total_time = time.time() - start_time
    print("\nTotal execution time:")
    print(f"Seconds: {total_time:.2f}")
    print(f"Minutes: {total_time/60:.2f}")
    print(f"Hours: {total_time/3600:.2f}")

    logging.info(f"Summary: Processed {total_items} items, {items_with_issues} with issues.")

if __name__ == "__main__":
    main()
    
# Add Getty and LCSH API calls here