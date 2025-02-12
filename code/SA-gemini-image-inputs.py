import os
import json
import base64
import logging
import google.generativeai as genai
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment
from PIL import Image as PILImage
from io import BytesIO
from openpyxl.drawing.image import Image as XLImage
import time
import re
import httpx
import warnings

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Add API stats tracking
class APIStats:
    def __init__(self):
        self.total_requests = 0
        self.total_input_tokens = 0
        self.total_output_tokens = 0
        self.total_tokens = 0

api_stats = APIStats()

def get_combined_prompt():
    return """This image is from 'Southern Architect and Building News', a periodical published from 1892 to 1931 that covered topics of interest to persons in the architecture, building, and hardware trades in the American South. We are creating metadata for the entire collection for the Architecture and Planning Library Special Collections, University of Texas Libraries, The University of Texas at Austin.  Please analyze this image and return information in EXACTLY this format:

TEXT TRANSCRIPTION:
[Full, exact transcription of all text visible in the image. 
- Maintain all original spelling, punctuation, and terminology exactly as shown
- Preserve line breaks and paragraph structure
- Include ALL text visible in the image including headers, footers, and italicized text
- Do not correct, modernize, or sanitize any language
- Use [illegible] for text that cannot be read
- Use [...] for partially visible or cut-off text]

VISUAL DESCRIPTION:
[Detailed description of any images. Include:
- Layout and design of the page
- Detailed descriptions of illustrations and photographs
- Placement and appearance of elements on the page, but keep it concise]

TABLE OF CONTENTS ENTRY:
[A descriptive entry appropriate for a table of contents. Include a short description of the content and whether the page is a:
- Cover
- Table of contents
- Advertisement
- Editorial
- Article
- Other type of page]

NAMED ENTITIES:
[List key entities including:
- Architects and architectural firms
- Geographic locations
- Building names
- Significant people mentioned
- Organizations]
- [Entity 1]
- [Entity 2]
...

SUBJECT HEADINGS:
[Two to three Library of Congress Subject Headings, paying special attention to:
- Architectural styles and movements
- Building types and purposes
- Historical significance
- Geographic locations]
- [Subject 1]
- [Subject 2]
...

CONTENT WARNING:
[Note potentially sensitive content, or 'None' if none exists. Consider:
- Historical language or terminology
- Culturally sensitive material
- Potentially offensive or harmful imagery or language]"""

def parse_combined_response(text):
    sections = {}
    current_section = None
    current_content = []
    
    for line in text.split('\n'):
        line = line.strip()
        if line in ['TEXT TRANSCRIPTION:', 'VISUAL DESCRIPTION:', 'TABLE OF CONTENTS ENTRY:', 
                   'NAMED ENTITIES:', 'SUBJECT HEADINGS:', 'CONTENT WARNING:']:
            if current_section:
                content = '\n'.join(current_content).strip()
                if current_section in ['named_entities', 'subject_headings']:
                    sections[current_section] = [x.strip('- ') for x in content.split('\n') if x.strip('- ')]
                else:
                    sections[current_section] = content
                current_content = []
            current_section = line.replace(':', '').lower().replace(' ', '_')
        elif current_section and line:
            current_content.append(line)
    
    if current_section:
        content = '\n'.join(current_content).strip()
        if current_section in ['named_entities', 'subject_headings']:
            sections[current_section] = [x.strip('- ') for x in content.split('\n') if x.strip('- ')]
        else:
            sections[current_section] = content
    
    return {
        'ocr_analysis': {
            'textTranscription': sections.get('text_transcription', ''),
            'visualDescription': sections.get('visual_description', '')
        },
        'metadata_analysis': {
            'tocEntry': sections.get('table_of_contents_entry', ''),
            'namedEntities': sections.get('named_entities', []),
            'subjectHeadings': sections.get('subject_headings', []),
            'contentWarning': sections.get('content_warning', 'None')
        }
    }

def process_image(image_path, model):
    try:
        # Load the image using PIL
        image = PILImage.open(image_path)
        prompt = get_combined_prompt()
        
        # Count input tokens before making the request
        input_tokens = model.count_tokens([prompt, image])
        logging.info(f"Input tokens for {image_path}: {input_tokens.total_tokens}")
        
        # Track request
        api_stats.total_requests += 1

        # Generate content and wait for response
        response = model.generate_content([
            image,
            prompt
        ])
        
        # Important: Need to wait for response to complete
        response.resolve()

        # Check for blocked response or copyright issues
        if response.candidates[0].finish_reason == 4:  # Copyright block
            return json.dumps({
                'ocr_analysis': {
                    'textTranscription': 'Content blocked due to copyright detection',
                    'visualDescription': 'Content blocked due to copyright detection'
                },
                'metadata_analysis': {
                    'tocEntry': 'Content blocked',
                    'namedEntities': [],
                    'subjectHeadings': [],
                    'contentWarning': 'None'
                }
            }), "Content blocked due to copyright detection"


        # Get token usage from response
        api_stats.total_input_tokens += response.usage_metadata.prompt_token_count
        api_stats.total_output_tokens += response.usage_metadata.candidates_token_count
        api_stats.total_tokens += response.usage_metadata.total_token_count
        
        logging.info(f"Token usage for {image_path}:")
        logging.info(f"Input tokens: {response.usage_metadata.prompt_token_count}")
        logging.info(f"Output tokens: {response.usage_metadata.candidates_token_count}")
        logging.info(f"Total tokens: {response.usage_metadata.total_token_count}")

        # Get the text response and ensure it's properly formatted
        response_text = response.text
        logging.debug(f"Raw response text: {response_text}")

        # Check if response has the expected sections
        if not all(section in response_text for section in [
            'TEXT TRANSCRIPTION:',
            'VISUAL DESCRIPTION:',
            'TABLE OF CONTENTS ENTRY:',
            'NAMED ENTITIES:',
            'SUBJECT HEADINGS:',
            'CONTENT WARNING:'
        ]):
            # If missing sections, try to reformat the response
            logging.warning(f"Response missing sections for {image_path}. Attempting to standardize format.")
            
            # Extract whatever content we can get
            response_text = f"""
TEXT TRANSCRIPTION:
{response_text.split('TEXT TRANSCRIPTION:', 1)[-1].split('VISUAL DESCRIPTION:', 1)[0] if 'TEXT TRANSCRIPTION:' in response_text else ''}

VISUAL DESCRIPTION:
{response_text.split('VISUAL DESCRIPTION:', 1)[-1].split('TABLE OF CONTENTS ENTRY:', 1)[0] if 'VISUAL DESCRIPTION:' in response_text else ''}

TABLE OF CONTENTS ENTRY:
{response_text.split('TABLE OF CONTENTS ENTRY:', 1)[-1].split('NAMED ENTITIES:', 1)[0] if 'TABLE OF CONTENTS ENTRY:' in response_text else ''}

NAMED ENTITIES:
{response_text.split('NAMED ENTITIES:', 1)[-1].split('SUBJECT HEADINGS:', 1)[0] if 'NAMED ENTITIES:' in response_text else ''}

SUBJECT HEADINGS:
{response_text.split('SUBJECT HEADINGS:', 1)[-1].split('CONTENT WARNING:', 1)[0] if 'SUBJECT HEADINGS:' in response_text else ''}

CONTENT WARNING:
{response_text.split('CONTENT WARNING:', 1)[-1] if 'CONTENT WARNING:' in response_text else 'None'}
"""

        # Parse the response
        try:
            parsed_response = parse_combined_response(response_text)
            logging.debug(f"Parsed response: {json.dumps(parsed_response, indent=2)}")
            
            # Verify we got some content
            if not any([
                parsed_response['ocr_analysis']['textTranscription'],
                parsed_response['ocr_analysis']['visualDescription'],
                parsed_response['metadata_analysis']['tocEntry'],
                parsed_response['metadata_analysis']['namedEntities'],
                parsed_response['metadata_analysis']['subjectHeadings']
            ]):
                # Try one more time with a modified prompt if we got no content
                logging.warning(f"Empty response for {image_path}. Retrying with simpler prompt...")
                
                retry_prompt = "Please analyze this image and describe: 1) Any text you can see 2) What's in the image 3) What topics it covers"
                retry_response = model.generate_content([image, retry_prompt])
                retry_response.resolve()
                
                # Use retry response if we got one
                if retry_response.text:
                    response_text = f"""
TEXT TRANSCRIPTION:
{retry_response.text.split('1)')[1].split('2)')[0] if '1)' in retry_response.text else ''}

VISUAL DESCRIPTION:
{retry_response.text.split('2)')[1].split('3)')[0] if '2)' in retry_response.text else ''}

TABLE OF CONTENTS ENTRY:
Analysis of historical architectural document

NAMED ENTITIES:
{', '.join(name.strip() for name in retry_response.text.split() if name[0].isupper())}

SUBJECT HEADINGS:
- Architecture
- Historical Documents
- Building Design

CONTENT WARNING:
None
"""
                    parsed_response = parse_combined_response(response_text)
            
            return json.dumps(parsed_response), response_text
            
        except Exception as parse_error:
            logging.error(f"Error parsing response for {image_path}: {str(parse_error)}")
            raise parse_error
                
    except Exception as e:
        logging.error(f"Error processing image {image_path}: {str(e)}")
        return str(e), str(e)

def process_folder(input_folder, output_dir):
    # Configure Gemini with safety settings
    genai.configure(api_key=os.environ.get("GEMINI_API_KEY"))
    
    # Create model with relaxed safety settings if needed
    generation_config = {
        "temperature": 0.4,  # Lower temperature for more consistent outputs
        "top_p": 0.8,
        "top_k": 40,
        "max_output_tokens": 2048,
    }
    
    safety_settings = [
        {
            "category": "HARM_CATEGORY_HARASSMENT",
            "threshold": "BLOCK_NONE",
        },
        {
            "category": "HARM_CATEGORY_HATE_SPEECH",
            "threshold": "BLOCK_NONE",
        },
        {
            "category": "HARM_CATEGORY_SEXUALLY_EXPLICIT",
            "threshold": "BLOCK_NONE",
        },
        {
            "category": "HARM_CATEGORY_DANGEROUS_CONTENT",
            "threshold": "BLOCK_NONE",
        },
    ]
    
    model = genai.GenerativeModel(
        model_name='gemini-2.0-flash-exp',
        generation_config=generation_config,
        safety_settings=safety_settings
    )
    
    # Get model information
    model_info = genai.get_model('models/gemini-2.0-flash-exp')
    logging.info(f"Model input token limit: {model_info.input_token_limit}")
    logging.info(f"Model output token limit: {model_info.output_token_limit}")
    
    # Initialize workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis"
    
    # Set up headers and formatting
    headers = [
        'Folder', 'Page Number', 'Image Path', 
        'Text Transcription', 'Visual Description',
        'TOC Entry', 'Named Entities', 'Subject Headings', 'Content Warning'
    ]
    ws.append(headers)
    
    # Create API Stats sheet
    stats_sheet = wb.create_sheet("API Stats")
    stats_sheet.append(["Metric", "Value"])
    
    # Freeze the top row
    ws.freeze_panes = 'A2'
    
    # Set column widths
    column_widths = [15, 11, 30, 70, 50, 30, 30, 30, 30]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[ws.cell(row=1, column=i+1).column_letter].width = width
    
    all_results = []
    raw_responses = []  # New list to store raw responses
    
    for folder_name in sorted(os.listdir(input_folder)):
        folder_path = os.path.join(input_folder, folder_name)
        if os.path.isdir(folder_path):
            image_files = sorted([f for f in os.listdir(folder_path) 
                                if f.lower().endswith(('.jpg', '.jpeg', '.png', '.webp', '.heic', '.heif'))],
                               key=lambda x: int(re.search(r'page(\d+)', x).group(1)))
            
            for img_file in image_files:
                page_number = int(re.search(r'page(\d+)', img_file).group(1))
                img_path = os.path.join(folder_path, img_file)
                
                try:
                    parsed_response, raw_response = process_image(img_path, model)
                    response_data = json.loads(parsed_response)

                    # Store raw response with metadata
                    raw_response_entry = {
                        'folder': str(folder_name),
                        'page_number': int(page_number),
                        'image_path': str(img_path),
                        'raw_response': raw_response
                    }
                    raw_responses.append(raw_response_entry)
                    
                    # Add results to worksheet
                    row = [
                        folder_name,
                        page_number,
                        img_path,
                        response_data['ocr_analysis']['textTranscription'],
                        response_data['ocr_analysis']['visualDescription'],
                        response_data['metadata_analysis']['tocEntry'],
                        ', '.join(response_data['metadata_analysis']['namedEntities']),
                        ', '.join(response_data['metadata_analysis']['subjectHeadings']),
                        response_data['metadata_analysis']['contentWarning']
                    ]
                    ws.append(row)
                    
                    # Add thumbnail
                    img = PILImage.open(img_path)
                    img.thumbnail((200, 200))
                    output = BytesIO()
                    img.save(output, format='JPEG')
                    output.seek(0)
                    img_excel = XLImage(output)
                    img_excel.anchor = ws.cell(row=ws.max_row, column=3).coordinate
                    ws.add_image(img_excel)
                    
                    # Format cells
                    for cell in ws[ws.max_row]:
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                    ws.row_dimensions[ws.max_row].height = 150
                    
                    # Update the results storage
                    result = {
                        'folder': str(folder_name),
                        'page_number': int(page_number),
                        'image_path': str(img_path),
                        'ocr_analysis': response_data['ocr_analysis'],
                        'metadata_analysis': response_data['metadata_analysis']
                    }
                    all_results.append(result)
                    
                except Exception as e:
                    logging.error(f"Error processing {img_path}: {str(e)}")
                    error_row = [
                        folder_name,
                        page_number,
                        img_path,
                        f"Error: {str(e)}",
                        "",
                        "",
                        "",
                        "",
                        ""
                    ]
                    ws.append(error_row)
    
    # Enhanced API stats to Excel
    api_summary = {
        "Total API Requests": api_stats.total_requests,
        "Total Input Tokens": api_stats.total_input_tokens,
        "Total Output Tokens": api_stats.total_output_tokens,
        "Total Tokens": api_stats.total_tokens,
        "Average Tokens per Request": api_stats.total_tokens / api_stats.total_requests if api_stats.total_requests > 0 else 0
    }
    
    for key, value in api_summary.items():
        stats_sheet.append([key, value])
    
    # Add API stats to results
    all_results.append({"api_stats": api_summary})
    
    # Save results
    current_date = datetime.now().strftime("%Y-%m-%d")
    excel_path = os.path.join(output_dir, f"SA-image-analysis-gemini-2-0-flash-{current_date}.xlsx")
    json_path = os.path.join(output_dir, f"SA-image-analysis-gemini-2-0-flash-{current_date}.json")
    
    wb.save(excel_path)
    with open(json_path, 'w') as f:
        json.dump(all_results, f, indent=2)
    
    return all_results, api_summary

def main():
    start_time = time.time()
    
    input_folder = "/Users/hannahmoutran/Desktop/ai-projects-ut-austin-fall-2024/southern_architect/sa-test-10-pages"
    output_dir = "/Users/hannahmoutran/Desktop/ai-projects-ut-austin-fall-2024/southern_architect/sa-test-outputs-image-workflow"
    os.makedirs(output_dir, exist_ok=True)
    
    results, api_summary = process_folder(input_folder, output_dir)
    
    # Enhanced API usage statistics
    print("\nAPI Usage Statistics:")
    print(f"Total API Requests: {api_summary['Total API Requests']}")
    print(f"Total Input Tokens: {api_summary['Total Input Tokens']}")
    print(f"Total Output Tokens: {api_summary['Total Output Tokens']}")
    print(f"Total Tokens: {api_summary['Total Tokens']}")
    print(f"Average Tokens per Request: {api_summary['Average Tokens per Request']:.2f}")
    
    logging.info(f"Processing complete. Image analysis results saved in {output_dir}")
    
    total_time = time.time() - start_time
    print(f"\nTotal execution time:")
    print(f"Seconds: {total_time:.2f}")
    print(f"Minutes: {total_time/60:.2f}")
    print(f"Hours: {total_time/3600:.2f}")

if __name__ == "__main__":
    main()