import os
import json
import base64
import logging
import anthropic
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment
from PIL import Image as PILImage
from io import BytesIO
from openpyxl.drawing.image import Image as XLImage
import time
import re
import tiktoken
import warnings

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Add API stats tracking
class APIStats:
    def __init__(self):
        self.total_requests = 0
        self.total_input_tokens = 0
        self.total_output_tokens = 0

api_stats = APIStats()

def get_combined_prompt():
    return """This image is from 'Southern Architect and Building News', a periodical published from 1892 to 1931 that covered topics of interest to persons in the architecture, building, and hardware trades in the American South. We are creating metadata for the entire collection for the Architecture and Planning Library Special Collections, University of Texas Libraries, The University of Texas at Austin.  Please analyze this image and return information in EXACTLY this format:

TEXT TRANSCRIPTION:
[Full, exact transcription of all text visible in the image. 
- Maintain all original spelling, punctuation, and terminology exactly as shown
- Preserve line breaks and paragraph structure
- Include ALL text visible in the image including headers, footers, and italicized text
- Do not correct, modernize, or sanitize any language
- Use [illegible] for text that cannot be read
- Use [...] for partially visible or cut-off text]

VISUAL DESCRIPTION:
[Detailed description of any images. Include:
- Layout and design of the page
- Detailed descriptions of illustrations and photographs
- Placement and appearance of elements on the page, but keep it concise]

TABLE OF CONTENTS ENTRY:
[A descriptive entry appropriate for a table of contents. Include a short description of the content and whether the page is a:
- Cover
- Table of contents
- Advertisement
- Editorial
- Article
- Other type of page]

NAMED ENTITIES:
[List key entities including:
- Architects and architectural firms
- Geographic locations
- Building names
- Significant people mentioned
- Organizations]
- [Entity 1]
- [Entity 2]
...

SUBJECT HEADINGS:
[Two to three Library of Congress Subject Headings, paying special attention to:
- Architectural styles and movements
- Building types and purposes
- Historical significance
- Geographic locations]
- [Subject 1]
- [Subject 2]
...

CONTENT WARNING:
[Note potentially sensitive content, or 'None' if none exists. Consider:
- Historical language or terminology
- Culturally sensitive material
- Potentially offensive or harmful imagery or language]"""

def parse_combined_response(text):
    sections = {}
    current_section = None
    current_content = []
    
    for line in text.split('\n'):
        line = line.strip()
        if line in ['TEXT TRANSCRIPTION:', 'VISUAL DESCRIPTION:', 'TABLE OF CONTENTS ENTRY:', 
                   'NAMED ENTITIES:', 'SUBJECT HEADINGS:', 'CONTENT WARNING:']:
            if current_section:
                content = '\n'.join(current_content).strip()
                if current_section in ['named_entities', 'subject_headings']:
                    sections[current_section] = [x.strip('- ') for x in content.split('\n') if x.strip('- ')]
                else:
                    sections[current_section] = content
                current_content = []
            current_section = line.replace(':', '').lower().replace(' ', '_')
        elif current_section and line:
            current_content.append(line)
    
    if current_section:
        content = '\n'.join(current_content).strip()
        if current_section in ['named_entities', 'subject_headings']:
            sections[current_section] = [x.strip('- ') for x in content.split('\n') if x.strip('- ')]
        else:
            sections[current_section] = content
    
    return {
        'ocr_analysis': {
            'textTranscription': sections.get('text_transcription', ''),
            'visualDescription': sections.get('visual_description', '')
        },
        'metadata_analysis': {
            'tocEntry': sections.get('table_of_contents_entry', ''),
            'namedEntities': sections.get('named_entities', []),
            'subjectHeadings': sections.get('subject_headings', []),
            'contentWarning': sections.get('content_warning', 'None')
        }
    }

def process_image(image_path, client):
    try:
        with open(image_path, "rb") as image_file:
            base64_image = base64.b64encode(image_file.read()).decode('utf-8')
            
            image_format = image_path.lower().split('.')[-1]
            if image_format == 'jpg':
                image_format = 'jpeg'
            media_type = f"image/{image_format}"

            # Track request
            api_stats.total_requests += 1

            response = client.messages.create(
                model="claude-3-haiku-20240307",
                max_tokens=1500,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image",
                                "source": {
                                    "type": "base64",
                                    "media_type": media_type,
                                    "data": base64_image,
                                }
                            },
                            {
                                "type": "text",
                                "text": get_combined_prompt()
                            }
                        ]
                    }
                ]
            )

            # Track tokens
            api_stats.total_input_tokens += response.usage.input_tokens
            api_stats.total_output_tokens += response.usage.output_tokens

            response_text = response.content[0].text
            parsed_response = parse_combined_response(response_text)
            return json.dumps(parsed_response), response_text
                
    except Exception as e:
        logging.error(f"Error processing image {image_path}: {str(e)}")
        return str(e), str(e)
      
def num_tokens_from_string(string: str) -> int:
    try:
        encoding = tiktoken.encoding_for_model("claude-3-haiku-20240307")
    except KeyError:
        warnings.warn("Using cl100k_base tokenizer as fallback")
        encoding = tiktoken.get_encoding("cl100k_base")
    return len(encoding.encode(string))

def process_folder(input_folder, output_dir):
    client = anthropic.Anthropic(api_key=os.environ.get("CLAUDE_API_KEY"))
    
    # Initialize workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis"
    
    # Set up headers and formatting
    headers = [
        'Folder', 'Page Number', 'Image Path', 
        'Text Transcription', 'Visual Description',
        'TOC Entry', 'Named Entities', 'Subject Headings', 'Content Warning'
    ]
    ws.append(headers)
    
    # Create API Stats sheet
    stats_sheet = wb.create_sheet("API Stats")
    stats_sheet.append(["Metric", "Value"])
    
    # Freeze the top row
    ws.freeze_panes = 'A2'
    
    # Set column widths
    column_widths = [15, 11, 30, 70, 50, 30, 30, 30, 30]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[ws.cell(row=1, column=i+1).column_letter].width = width
    
    all_results = []
    
    for folder_name in sorted(os.listdir(input_folder)):
        folder_path = os.path.join(input_folder, folder_name)
        if os.path.isdir(folder_path):
            image_files = sorted([f for f in os.listdir(folder_path) 
                                if f.lower().endswith(('.jpg', '.jpeg', '.png'))],
                               key=lambda x: int(re.search(r'page(\d+)', x).group(1)))
            
            for img_file in image_files:
                page_number = int(re.search(r'page(\d+)', img_file).group(1))
                img_path = os.path.join(folder_path, img_file)
                
                try:
                    parsed_response, raw_response = process_image(img_path, client)
                    response_data = json.loads(parsed_response)

                    # Add results to worksheet
                    row = [
                        folder_name,
                        page_number,
                        img_path,
                        response_data['ocr_analysis']['textTranscription'],
                        response_data['ocr_analysis']['visualDescription'],
                        response_data['metadata_analysis']['tocEntry'],
                        ', '.join(response_data['metadata_analysis']['namedEntities']),
                        ', '.join(response_data['metadata_analysis']['subjectHeadings']),
                        response_data['metadata_analysis']['contentWarning']
                    ]
                    ws.append(row)
                    
                    # Add thumbnail
                    img = PILImage.open(img_path)
                    img.thumbnail((200, 200))
                    output = BytesIO()
                    img.save(output, format='JPEG')
                    output.seek(0)
                    img_excel = XLImage(output)
                    img_excel.anchor = ws.cell(row=ws.max_row, column=3).coordinate
                    ws.add_image(img_excel)
                    
                    # Format cells
                    for cell in ws[ws.max_row]:
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                    ws.row_dimensions[ws.max_row].height = 150
                    
                    # Update the results storage
                    result = {
                        'folder': str(folder_name),
                        'page_number': int(page_number),
                        'image_path': str(img_path),
                        'ocr_analysis': response_data['ocr_analysis'],
                        'metadata_analysis': response_data['metadata_analysis']
                    }
                    all_results.append(result)
                    
                except Exception as e:
                    logging.error(f"Error processing {img_path}: {str(e)}")
                    error_row = [
                        folder_name,
                        page_number,
                        img_path,
                        f"Error: {str(e)}",
                        "",
                        "",
                        "",
                        "",
                        ""
                    ]
                    ws.append(error_row)
    
# Add API stats to Excel
    api_summary = {
        "Total API Requests": api_stats.total_requests,
        "Total Input Tokens": api_stats.total_input_tokens,
        "Total Output Tokens": api_stats.total_output_tokens,
        "Total Tokens": api_stats.total_input_tokens + api_stats.total_output_tokens
    }
    
    for key, value in api_summary.items():
        stats_sheet.append([key, value])
    
    # Add API stats to results
    all_results.append({"api_stats": api_summary})
    
    # Save results
    current_date = datetime.now().strftime("%Y-%m-%d")
    excel_path = os.path.join(output_dir, f"SA-image-analysis-claude-haiku-{current_date}.xlsx")
    json_path = os.path.join(output_dir, f"SA-image-analysis-claude-haiku-{current_date}.json")
    
    wb.save(excel_path)
    with open(json_path, 'w') as f:
        json.dump(all_results, f, indent=2)
    
    return all_results, api_summary

def main():
    start_time = time.time()
    
    input_folder = "/Users/hannahmoutran/Desktop/ai-projects-ut-austin-fall-2024/southern_architect/southern-architect-testing-pages"
    output_dir = "/Users/hannahmoutran/Desktop/ai-projects-ut-austin-fall-2024/southern_architect/southern-architect-testing-sheets-json-api-image-inputs"
    os.makedirs(output_dir, exist_ok=True)
    
    results, api_summary = process_folder(input_folder, output_dir)
    
    # Print API usage statistics
    print("\nAPI Usage Statistics:")
    print(f"Total API Requests: {api_summary['Total API Requests']}")
    print(f"Total Input Tokens: {api_summary['Total Input Tokens']}")
    print(f"Total Output Tokens: {api_summary['Total Output Tokens']}")
    print(f"Total Tokens: {api_summary['Total Tokens']}")
    
    logging.info(f"Processing complete. Image analysis results saved in {output_dir}")
    
    total_time = time.time() - start_time
    print(f"\nTotal execution time:")
    print(f"Seconds: {total_time:.2f}")
    print(f"Minutes: {total_time/60:.2f}")
    print(f"Hours: {total_time/3600:.2f}")

if __name__ == "__main__":
    main()