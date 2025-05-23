import os
import json
import base64
import logging
from datetime import datetime
from openai import OpenAI
import tenacity
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment
from PIL import Image as PILImage
from io import BytesIO
from openpyxl.drawing.image import Image as XLImage
import time

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))

class APIStats:
    def __init__(self):
        self.total_requests = 0
        self.total_input_tokens = 0
        self.total_output_tokens = 0

api_stats = APIStats()

def get_combined_prompt():
    return """This image is from 'Southern Architect and Building News', a periodical published from 1892 to 1931 that covered topics of interest to persons in the architecture, building, and hardware trades in the American South. We are creating metadata for the entire collection for the Architecture and Planning Library Special Collections, University of Texas Libraries, The University of Texas at Austin. Please analyze this image and return ONLY a JSON response in this exact format:

{
    "textTranscription": "Full, exact transcription of all text visible in the image. 
        - Maintain all original spelling, punctuation, and terminology exactly as shown
        - Preserve line breaks and paragraph structure
        - Include ALL text visible including headers, footers, and italicized text
        - Do not correct, modernize, or sanitize any language
        - Use [illegible] for text that cannot be read
        - Use [...] for partially visible or cut-off text",
    
    "visualDescription": "Detailed description of any images. Include:
        - Layout and design of the page
        - Detailed descriptions of illustrations and photographs
        - Placement and appearance of elements on the page, but keep it concise",
    
    "tocEntry": "A descriptive entry appropriate for a table of contents. Include:
        - Short description of the content
        - Page type (cover, table of contents, advertisement, editorial, article, or other)",
    
    "namedEntities": [
        "List key entities including:
        - Architects and architectural firms
        - Geographic locations 
        - Building names
        - Significant people mentioned
        - Organizations"
    ],
    
    "subjects": [
        "Two to three subjects that will later be used to search for verified headings in FAST, focusing on:
        - Architectural styles and movements
        - Building types and purposes
        - Historical significance
        - Geographic locations"
    ],
    
    "contentWarning": "Note potentially sensitive content, or 'None' if none exists. Consider:
        - Historical language or terminology
        - Culturally sensitive material
        - Potentially offensive or harmful imagery or language"
}"""

@tenacity.retry(
    wait=tenacity.wait_exponential(multiplier=1, min=4, max=10),
    stop=tenacity.stop_after_attempt(5),
    retry=tenacity.retry_if_exception_type(Exception)
)
def process_image(image_path):
    with open(image_path, "rb") as image_file:
        base64_image = base64.b64encode(image_file.read()).decode('utf-8')

    api_stats.total_requests += 1
    response = client.chat.completions.create(
        model="gpt-4o-mini-2024-07-18",
        messages=[{
            "role": "user",
            "content": [
                {"type": "text", "text": get_combined_prompt()},
                {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}}
            ]
        }],
        max_tokens=2000
    )
    
    api_stats.total_input_tokens += response.usage.prompt_tokens
    api_stats.total_output_tokens += response.usage.completion_tokens
    
    raw_response = response.choices[0].message.content.strip()
    
    try:
        cleaned_response = re.sub(r'```json\s*|\s*```', '', raw_response)
        cleaned_response = re.sub(r'^[^{]*({.*})[^}]*$', r'\1', cleaned_response, flags=re.DOTALL)
        parsed_json = json.loads(cleaned_response)
        
        required_fields = ['textTranscription', 'visualDescription', 'tocEntry', 
                          'namedEntities', 'subjects', 'contentWarning']
        if not all(key in parsed_json for key in required_fields):
            raise KeyError("Missing required fields in JSON response")
            
        return json.dumps(parsed_json), raw_response
        
    except Exception as e:
        logging.warning(f"Initial JSON parsing failed for {image_path}: {str(e)}")
        try:
            match = re.search(r'{.*}', raw_response, re.DOTALL)
            if not match:
                raise ValueError("No JSON object found in response")
                
            json_str = match.group(0)
            parsed_json = json.loads(json_str)
            
            if not all(key in parsed_json for key in required_fields):
                raise KeyError("Missing required fields in JSON response")
                
            return json.dumps(parsed_json), raw_response
                
        except Exception as e2:
            logging.error(f"JSON parsing failed after second attempt for {image_path}: {str(e2)}\nRaw response: {raw_response}")
            return raw_response, raw_response

def process_folder(input_folder, output_dir):
    wb = Workbook()
    
    analysis_sheet = wb.active
    analysis_sheet.title = "Analysis"
    analysis_headers = [
        'Folder', 'Page Number', 'Image Path', 
        'Text Transcription', 'Visual Description',
        'TOC Entry', 'Named Entities', 'Subjects', 'Content Warning'
    ]
    analysis_sheet.append(analysis_headers)
    analysis_sheet.freeze_panes = 'A2'
    
    column_widths = [11, 11, 30, 70, 50, 30, 30, 30, 30]
    for i, width in enumerate(column_widths):
        analysis_sheet.column_dimensions[analysis_sheet.cell(row=1, column=i+1).column_letter].width = width
    
    raw_sheet = wb.create_sheet("Raw Responses")
    raw_headers = ['Folder', 'Page Number', 'API Response']
    raw_sheet.append(raw_headers)
    raw_sheet.freeze_panes = 'A2'
    
    for i, width in enumerate([15, 10, 120]):
        raw_sheet.column_dimensions[raw_sheet.cell(row=1, column=i+1).column_letter].width = width
    
    max_row_height = 409
    analysis_sheet.row_dimensions[1].height = 15
    raw_sheet.row_dimensions[1].height = 15

    all_results = []
    
    for folder_name in sorted(os.listdir(input_folder)):
        folder_path = os.path.join(input_folder, folder_name)
        if os.path.isdir(folder_path):
            image_files = sorted([f for f in os.listdir(folder_path) 
                                if f.lower().endswith(('.jpg', '.jpeg', '.png'))],
                               key=lambda x: int(re.search(r'page(\d+)', x).group(1)))
            
            for img_file in image_files:
                page_number = int(re.search(r'page(\d+)', img_file).group(1))
                img_path = os.path.join(folder_path, img_file)
                
                try:
                    cleaned_response, raw_response = process_image(img_path)
                    try:
                        response_data = json.loads(cleaned_response)
                    except json.JSONDecodeError:
                        logging.warning(f"Failed to parse JSON for {img_path}")
                        response_data = {
                            "textTranscription": raw_response,
                            "visualDescription": raw_response,
                            "tocEntry": raw_response,
                            "namedEntities": [],
                            "subjects": [],
                            "contentWarning": raw_response
                        }
                    
                    analysis_row = [
                        folder_name,
                        page_number,
                        img_path,
                        response_data.get('textTranscription', raw_response),
                        response_data.get('visualDescription', raw_response),
                        response_data.get('tocEntry', raw_response),
                        ', '.join(response_data.get('namedEntities', [])),
                        ', '.join(response_data.get('subjects', [])),
                        response_data.get('contentWarning', raw_response)
                    ]
                    analysis_sheet.append(analysis_row)
                    
                    img = PILImage.open(img_path)
                    img.thumbnail((200, 200))
                    output = BytesIO()
                    img.save(output, format='JPEG')
                    output.seek(0)
                    img_excel = XLImage(output)
                    img_excel.anchor = analysis_sheet.cell(row=analysis_sheet.max_row, column=3).coordinate
                    analysis_sheet.add_image(img_excel)
                    
                    analysis_sheet.row_dimensions[analysis_sheet.max_row].height = max_row_height
                    
                    current_row = analysis_sheet.max_row
                    for col_idx, cell in enumerate(analysis_sheet[current_row], 1):
                        if col_idx == 3:
                            cell.alignment = Alignment(vertical='bottom', wrap_text=True)
                        else:
                            cell.alignment = Alignment(vertical='top', wrap_text=True)
                    
                    raw_row = [folder_name, page_number, raw_response]
                    raw_sheet.append(raw_row)
                    raw_sheet.row_dimensions[raw_sheet.max_row].height = max_row_height
                    
                    for cell in raw_sheet[raw_sheet.max_row]:
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                    
                    entry_result = {
                        'folder': folder_name,
                        'page_number': page_number,
                        'image_path': img_path,
                        'analysis': {
                            'text_transcription': response_data.get('textTranscription', raw_response),
                            'visual_description': response_data.get('visualDescription', raw_response),
                            'toc_entry': response_data.get('tocEntry', raw_response),
                            'named_entities': response_data.get('namedEntities', []),
                            'subject_headings': response_data.get('subjects', []),
                            'content_warning': response_data.get('contentWarning', raw_response),
                            'raw_response': raw_response
                        }
                    }
                    all_results.append(entry_result)
                    
                except Exception as e:
                    logging.error(f"Error processing {img_path}: {str(e)}")
                    error_entry = {
                        'folder': folder_name,
                        'page_number': page_number,
                        'image_path': img_path,
                        'error': str(e),
                        'analysis': {
                            'text_transcription': 'Processing error',
                            'visual_description': 'Processing error',
                            'toc_entry': 'Processing error',
                            'named_entities': [],
                            'subject_headings': [],
                            'content_warning': 'Processing error',
                            'raw_response': str(e)
                        }
                    }
                    all_results.append(error_entry)
                    
                    error_row = [
                        folder_name,
                        page_number,
                        img_path,
                        'Processing error',
                        'Processing error',
                        'Processing error',
                        '',
                        '',
                        str(e)
                    ]
                    analysis_sheet.append(error_row)
                    
                    raw_error_row = [folder_name, page_number, str(e)]
                    raw_sheet.append(raw_error_row)
    
    api_summary = {
        "total_requests": api_stats.total_requests,
        "total_input_tokens": api_stats.total_input_tokens,
        "total_output_tokens": api_stats.total_output_tokens,
        "total_tokens": api_stats.total_input_tokens + api_stats.total_output_tokens
    }
    
    stats_sheet = wb.create_sheet("API Stats")
    stats_sheet.append(["Metric", "Value"])
    for key, value in api_summary.items():
        stats_sheet.append([key, value])
    
    current_date = datetime.now().strftime("%Y-%m-%d")
    excel_path = os.path.join(output_dir, f"SA-image-gpt-4o-mini-{current_date}.xlsx")
    json_path = os.path.join(output_dir, f"SA-image-gpt-4o-mini-{current_date}.json")
    
    all_results.append({"api_stats": api_summary})
    
    wb.save(excel_path)
    with open(json_path, 'w') as f:
        json.dump(all_results, f, indent=2)
    
    return all_results, api_summary

def main():
    start_time = time.time()  
    
    input_folder = "/Users/hannahmoutran/Desktop/ai-projects-ut-austin-fall-2024/southern_architect/sa-test-10-pages"
    output_dir = "/Users/hannahmoutran/Desktop/ai-projects-ut-austin-fall-2024/southern_architect/sa-test-outputs-image-workflow"
    os.makedirs(output_dir, exist_ok=True)
    
    results, api_summary = process_folder(input_folder, output_dir)
    
    print("\nAPI Usage Statistics:")
    print(f"Total API Requests: {api_summary['total_requests']}")
    print(f"Total Input Tokens: {api_summary['total_input_tokens']}")
    print(f"Total Output Tokens: {api_summary['total_output_tokens']}")
    print(f"Total Tokens: {api_summary['total_tokens']}")
    
    total_time = time.time() - start_time
    print(f"\nTotal execution time:")
    print(f"Seconds: {total_time:.2f}")
    print(f"Minutes: {total_time/60:.2f}")
    print(f"Hours: {total_time/3600:.2f}")

if __name__ == "__main__":
    main()