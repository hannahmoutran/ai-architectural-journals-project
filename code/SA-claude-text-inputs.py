import os
import json
import anthropic
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import logging
import time
from datetime import datetime
import traceback
import tiktoken
import warnings

class APIStats:
    def __init__(self):
        self.total_requests = 0
        self.total_input_tokens = 0
        self.total_output_tokens = 0
        self.total_tokens = 0

api_stats = APIStats()

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def clean_json_string(json_str):
    json_str = re.sub(r'```json\s*|\s*```', '', json_str)
    json_str = json_str.strip()
    json_str = json_str.replace('\\"', '"').replace("'", '"')
    json_str = re.sub(r'(?<!\\)"([^"]*)"', lambda m: '"{}"'.format(m.group(1).replace('"', '\\"')), json_str)
    json_str = json_str.replace('\n', '\\n')
    json_str = ''.join(ch for ch in json_str if ord(ch) >= 32)
    return json_str

def preprocess_ocr_text(text):
    # Replace common OCR errors
    replacements = {
        "Sv'cink i^idam": "Frank Adam",
        # Add more common OCR errors here
    }
    
    for error, correction in replacements.items():
        text = text.replace(error, correction)
    
    # Remove unusual characters
    text = re.sub(r'[^\w\s.,;:!?()-]', '', text)
    
    return text

def num_tokens_from_string(string: str, model: str) -> int:
    try:
        encoding = tiktoken.encoding_for_model(model)
    except KeyError:
        warnings.warn(f"No specific tokenizer found for {model}. Falling back to cl100k_base.")
        encoding = tiktoken.get_encoding("cl100k_base")
    return len(encoding.encode(string))

def make_api_call(client, prompt, content, is_short, max_retries=3, initial_delay=1):
    model = "claude-3-sonnet-20240229"  
    system_message = "You are an AI assistant tasked with cleaning OCR text and extracting metadata. Always format your response as follows: cleaned OCR text between '---' markers, followed by JSON metadata. It's crucial to include both opening and closing '---' markers."
    user_message = prompt + "\n\nHere's the content to analyze:\n\n" + content
    
    user_message += f"\n\nNote: This content is {'short (less than 250 characters)' if is_short else 'of normal length'}."
    user_message += "\n\nPlease format your response exactly as follows:\n---\n[Cleaned OCR text here]\n---\n{\"JSON metadata here\"}"

    input_tokens = num_tokens_from_string(system_message + user_message, model)

    api_stats.total_requests += 1  # Track request count

    for attempt in range(max_retries):
        try:
            message = client.messages.create(
                model=model,
                max_tokens=3000,
                system=system_message,
                messages=[{"role": "user", "content": user_message}]
            )
            response_text = message.content[0].text
            output_tokens = num_tokens_from_string(response_text, model)
            total_tokens = input_tokens + output_tokens

            # Add the token tracking here, after getting the counts but before processing the response
            api_stats.total_input_tokens += input_tokens
            api_stats.total_output_tokens += output_tokens
            api_stats.total_tokens += total_tokens

            parts = response_text.split('---')
            if len(parts) >= 3:
                cleaned_content = parts[1].strip()
                json_data = parts[2].strip()
                try:
                    parsed_json = json.loads(json_data)
                except json.JSONDecodeError:
                    parsed_json = parse_response(json_data)
            else:
                cleaned_content = response_text
                parsed_json = parse_response(response_text)

            if is_valid_response(parsed_json):
                return cleaned_content, parsed_json, input_tokens, output_tokens, total_tokens

        except anthropic.InternalServerError as e:
            if "overloaded" in str(e).lower():
                time.sleep(initial_delay * (2 ** attempt))
                continue
            raise
        except Exception as e:
            logging.error(f"Error on attempt {attempt + 1}: {str(e)}")
            
        if attempt < max_retries - 1:
            time.sleep(initial_delay)

    return None, None, 0, 0, 0

def parse_response(response_text):
    # First, try to find and parse a JSON object in the response
    json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
    if json_match:
        try:
            return json.loads(json_match.group(0))
        except json.JSONDecodeError:
            pass  # If JSON parsing fails, fall through to regex parsing

    # If JSON parsing fails, use regex to extract fields
    toc_entry = re.search(r'"tocEntry":\s*"(.*?)"', response_text, re.DOTALL)
    named_entities = re.findall(r'"namedEntities":\s*\[(.*?)\]', response_text, re.DOTALL)
    subject_headings = re.findall(r'"subjectHeadings":\s*\[(.*?)\]', response_text, re.DOTALL)
    content_warning = re.search(r'"contentWarning":\s*"(.*?)"', response_text, re.DOTALL)

    return {
        "tocEntry": toc_entry.group(1) if toc_entry else "",
        "namedEntities": [entity.strip().strip('"') for entity in named_entities[0].split(',')] if named_entities else [],
        "subjectHeadings": [heading.strip().strip('"') for heading in subject_headings[0].split(',')] if subject_headings else [],
        "contentWarning": content_warning.group(1) if content_warning else "None"
    }
        
        
def is_valid_response(parsed_json):
    return all(key in parsed_json for key in ['tocEntry', 'namedEntities', 'subjectHeadings']) and any(parsed_json.values())

def calculate_totals(ws):
    def safe_sum(column):
        total = 0
        for i in range(2, ws.max_row + 1):
            cell_value = ws[f'{column}{i}'].value
            if cell_value:
                try:
                    total += int(cell_value)
                except ValueError:
                    logging.warning(f"Non-integer value '{cell_value}' found in column {column}, row {i}. Skipping.")
        return total

    input_total = safe_sum('I')  # Column I for Input Tokens
    output_total = safe_sum('J')  # Column J for Output Tokens
    total_total = safe_sum('K')  # Column K for Total Tokens

    return input_total, output_total, total_total

def postprocess_api_response(response_data):
    if 'namedEntities' in response_data:
        # Remove duplicates while preserving order
        response_data['namedEntities'] = list(dict.fromkeys(response_data['namedEntities']))
        
        # Remove any entities that are just single letters or numbers
        response_data['namedEntities'] = [entity for entity in response_data['namedEntities'] if len(entity) > 1 or not entity.isalnum()]
    
    return response_data

def process_file(file_path, client, folder_name, page_number):
    filename = os.path.basename(file_path)
    logging.info(f"Starting to process file: {filename}")
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()

        # Preprocess the OCR text
        content = preprocess_ocr_text(content)

        # Check if the content is empty or only whitespace
        if not content.strip():
            logging.info(f"No readable text in {filename}")
            return [
                folder_name,
                page_number,
                filename,
                "",
                "No readable text on this page",
                "",
                "Blank pages",
                "None",  # Content warning
                0,
                0,
                0
            ]

        is_short = len(content.strip()) < 225  # Adjust this threshold as needed
        is_cover = page_number == 1  # Assume page 1 is always the cover

        ocr_cleaning_instructions = """
Before analyzing the content, please clean the OCR text according to these guidelines:
1. Correct misspellings that are clearly due to OCR errors.
2. Fix obvious punctuation errors.
3. Correct word splits (e.g., "architec ture" should be "architecture").
4. Do not alter proper names, even if they seem unusual, unless it's an obvious OCR error.
5. Do not add or remove any content; your task is to clean, not to summarize or expand.
6. If you're unsure about a word or phrase, leave it as is.
7. Maintain the original formatting as much as possible.  Remember that it is a page from an old magazine.  Many pages begin with the name of the periodical, the date, and the page number.  This information may be edited slightly for clarity by, for example, separating the date from the page number.

Provide the complete cleaned text, followed by '---' before proceeding with the metadata extraction.
"""

        content_warning_instructions = """
content_warning_instructions:
After analyzing the content, assess whether it contains any language or themes that might warrant a content warning. Consider the following categories:
1. Racist or ethnically insensitive language
2. Sexist or gender-discriminatory content
3. Violence or graphic descriptions
4. Other potentially offensive or outdated terminology

If you identify any such content, briefly describe it using one of the following phrases in a 'contentWarning' field:
- For outdated and/or racist terms: Some of the language in this resource uses outdated and/or racist terminology.
- For imagery and/or descriptions of violence: Some of the language and/or images in this resource depict graphic and violent content.
- For combinations of depictions of violence and outdated terms: Some of the language and/or images in this resource depict graphic and violent content and outdated and/or racist terminology.

If no warning is necessary, set this field to "None".
"""

        if is_cover:
            prompt_text = f"""{ocr_cleaning_instructions}
You are an archivist at the University of Texas at Austin, cataloging a collection of issues of The Southern Architect (1892-1931).
This page is the cover of an issue. The OCR text for this cover is:

{content.strip()}

After cleaning the OCR text, please create the following metadata fields in JSON format:
'tocEntry': Start with "Cover:" followed by a brief description of the cover content, based on the text.
'namedEntities': A list of key entities mentioned in the text, including any significant names, titles, or locations.
'subjectHeadings': Include "Magazine covers" and any other relevant subject headings based on the content.
{content_warning_instructions}

Provide the cleaned text, followed by '---', then ONLY the JSON response, with no additional text before or after."""
        elif is_short:
            prompt_text = f"""{ocr_cleaning_instructions}
You are an archivist at the University of Texas at Austin, cataloging a collection of issues of The Southern Architect (1892-1931).
This page contains an image. The OCR text for this image is short, containing only the following text:

{content.strip()}

After cleaning the OCR text, please create the following metadata fields in JSON format:
'tocEntry': Start with "Photograph:" or "Drawing:" followed by a brief description of what the image likely shows, based on the text.
'namedEntities': A list of key entities mentioned in the text, including architects, locations, and any other significant names.
'subjectHeadings': Include "Architectural photography" or "Architecture--Designs and plans--Presentation drawings" and/or any other relevant subject headings based on the content.
{content_warning_instructions}

Provide the cleaned text, followed by '---', then ONLY the JSON response, with no additional text before or after."""
        else:
            prompt_text = f"""{ocr_cleaning_instructions}
You are an archivist at the University of Texas at Austin, cataloging a collection of issues of The Southern Architect (1892-1931). 
Important: you are working from OCR text that may contain errors, so you have to use your judgment to decipher meaning.  
Audience: you are creating this metadata for mainly architectural historians and architectural history students.  Architectural styles, movements, trends, and other historically important information should be prioritized.  

After cleaning the OCR text, create the following metadata fields in JSON:
'tocEntry': One short phrase that succinctly describes the main content of the page. Always note the type of content that it is - for example, "Advertisement for an architecture firm" or "Article on the latest architectural trends in Dallas, Texas".
'namedEntities': A list of key entities directly relevant to the main content. Include significant people, companies, products, and locations central to the topic. Exclude lists of branch offices or minor mentions.
'subjectHeadings': Two to three Library of Congress Authoritative subject headings that best represent the main themes or topics of the page.
{content_warning_instructions}

Focus on capturing the essence of the page's content rather than exhaustive detail. 

Rules for identifying advertisements:
1. Look for company names prominently displayed
2. Check for product descriptions or benefits
3. Look for contact information or calls to action
4. Be aware of persuasive language or marketing slogans
5. Note any pricing information or special offers

Provide the cleaned text, followed by '---', then ONLY the JSON response, with no additional text before or after."""

        cleaned_content, metadata, input_tokens, output_tokens, total_tokens = make_api_call(client, prompt_text, content, is_short or is_cover)
        
        if cleaned_content and metadata:
            # Split the content at the first occurrence of '{'
            parts = cleaned_content.split('{', 1)
            if len(parts) == 2:
                cleaned_ocr_text = parts[0].strip()
                # Remove the leading '---' if present
                if cleaned_ocr_text.startswith('---'):
                    cleaned_ocr_text = cleaned_ocr_text[3:].strip()
            else:
                cleaned_ocr_text = cleaned_content.strip()
                logging.warning(f"Could not split content for {filename}. Using full content as OCR text.")

            # Postprocess the API response
            metadata = postprocess_api_response(metadata)
            
            if metadata.get('contentWarning', 'None') != 'None':
                logging.info(f"Content warning identified in {filename}: {metadata['contentWarning']}")

            result = [
                folder_name,
                page_number,
                filename,
                cleaned_ocr_text,
                metadata.get('tocEntry', 'N/A'),
                ', '.join(metadata.get('namedEntities', [])),
                ', '.join(metadata.get('subjectHeadings', [])),
                metadata.get('contentWarning', 'None'),
                input_tokens,
                output_tokens,
                total_tokens
            ]
            logging.info(f"Processed {filename} successfully. Result: {result[:4]}...")
            return result
        else:
            logging.warning(f"{folder_name}/{filename}: Failed to get valid response")
            return [folder_name, page_number, filename, "Error: Failed to get valid response", "", "", "", "None", 0, 0, 0]

    except Exception as e:
        logging.error(f"Error processing file {filename}: {str(e)}")
        logging.error(traceback.format_exc())  # Log the full traceback
        return [folder_name, page_number, filename, f"Error: Processing Error - {str(e)}", "", "", "", "None", 0, 0, 0]

def process_folder_recursive(folder_path, client, wb):
    start_time = time.time()
    ws = wb.active
    headers = ['Folder', 'Page Number', 'Page Title', 'Cleaned OCR Text', 'TOC Entry', 
               'Named Entities', 'Subject Headings', 'Content Warning', 'Input Tokens', 
               'Output Tokens', 'Total Tokens']
    ws.append(headers)

    # Create API Stats sheet
    stats_sheet = wb.create_sheet("API Stats")
    stats_sheet.append(["Metric", "Value"])

    # Set column widths
    column_widths = [15, 10, 30, 50, 30, 30, 30, 30, 15, 15, 15]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[ws.cell(row=1, column=i+1).column_letter].width = width

    items_with_issues = []
    total_items = 0
    all_data = []

    def folder_sort_key(x):
        parts = x.split('-')
        return tuple(parts + [''] * (3 - len(parts)))

    # Sort folders
    sorted_folders = sorted(os.listdir(folder_path), key=folder_sort_key, reverse=False)

    for folder_name in sorted_folders:
        folder_full_path = os.path.join(folder_path, folder_name)
        if os.path.isdir(folder_full_path):
            files = sorted(
                [f for f in os.listdir(folder_full_path) if f.endswith('.txt')],
                key=lambda x: int(re.search(r'page(\d+)', x).group(1))
            )
            
            for filename in files:
                total_items += 1
                file_path = os.path.join(folder_full_path, filename)
                
                page_number = int(re.search(r'page(\d+)', filename).group(1))
                
                result = process_file(file_path, client, folder_name, page_number)
                if isinstance(result, list) and len(result) == 11:
                    all_data.append(result)
                    logging.info(f"Added result for {filename} to all_data")
                else:
                    items_with_issues.append(f"{folder_name}/{filename}: Unexpected result format")
                    logging.error(f"Unexpected result format for {filename}: {result}")

    # Sort all_data by folder (reversed) and page number
    all_data.sort(key=lambda x: (folder_sort_key(x[0]), int(x[1])), reverse=False)

    # Add sorted data to worksheet
    for row_data in all_data:
        ws.append(row_data)

    # Apply text wrapping and vertical alignment to all rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Calculate and add totals row
    input_total, output_total, total_total = calculate_totals(ws)
    ws.append(['Totals', '', '', '', '', '', '', '', input_total, output_total, total_total])
    
    # Style the totals row
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    # Create a new sheet for items with issues
    ws_issues = wb.create_sheet(title="Issues")
    ws_issues.append(["Filename"])
    for item in items_with_issues:
        ws_issues.append([item])

 #Add API stats to Excel
    api_summary = {
        "Total API Requests": api_stats.total_requests,
        "Total Input Tokens": api_stats.total_input_tokens,
        "Total Output Tokens": api_stats.total_output_tokens,
        "Total Tokens": api_stats.total_tokens,
        "Execution Time (seconds)": time.time() - start_time,
        "Execution Time (minutes)": (time.time() - start_time) / 60,
        "Execution Time (hours)": (time.time() - start_time) / 3600
    }
    
    for key, value in api_summary.items():
        if "Time" in key:
            stats_sheet.append([key, f"{value:.2f}"])
        else:
            stats_sheet.append([key, value])

    return total_items, len(items_with_issues), api_summary

def main():
    start_time = time.time()
    
    logging.info("Starting the script")
    folder_path = "/Users/hannahmoutran/Desktop/ai-projects-ut-austin-fall-2024/southern_architect/southern-architect-page-level-file-folders-small copy"
    client = anthropic.Anthropic(api_key=os.environ.get("CLAUDE_API_KEY"))
    wb = Workbook()

    logging.info("Processing folder")
    total_items, items_with_issues, api_summary = process_folder_recursive(folder_path, client, wb)

    # Print API usage statistics
    print("\nAPI Usage Statistics:")
    print(f"Total API Requests: {api_summary['Total API Requests']}")
    print(f"Total Input Tokens: {api_summary['Total Input Tokens']}")
    print(f"Total Output Tokens: {api_summary['Total Output Tokens']}")
    print(f"Total Tokens: {api_summary['Total Tokens']}")
    
    total_time = time.time() - start_time
    print(f"\nTotal execution time:")
    print(f"Seconds: {total_time:.2f}")
    print(f"Minutes: {total_time/60:.2f}")
    print(f"Hours: {total_time/3600:.2f}")

    # Generate the output filename with current date
    current_date = datetime.now().strftime("%Y-%m-%d")
    output_file = f"SA-claude-3-sonnet-{current_date}.xlsx"
    output_dir = "/Users/hannahmoutran/Desktop/ai-projects-ut-austin-fall-2024/southern_architect/Testing Output"
    os.makedirs(output_dir, exist_ok=True)
    full_output_path = os.path.join(output_dir, output_file)

    # Save the workbook
    logging.info(f"Attempting to save file to {full_output_path}")
    try:
        wb.save(full_output_path)
        logging.info(f"Results saved to {full_output_path}")
    except Exception as e:
        logging.error(f"Error saving file: {str(e)}", exc_info=True)
    return total_time, api_summary

if __name__ == "__main__":
    execution_time, api_stats = main()
    print(f"\nScript completed in {execution_time:.2f} seconds")
    