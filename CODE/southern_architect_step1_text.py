import os
import json
import logging
from datetime import datetime
from openai import OpenAI
import tenacity
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment
import time
import argparse

# Import custom modules
from model_pricing import calculate_cost, get_model_info
from token_logging import create_token_usage_log, log_individual_response
from batch_processor import BatchProcessor
from prompts import SouthernArchitectPrompts

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Suppress verbose HTTP logging
logging.getLogger("openai").setLevel(logging.WARNING)
logging.getLogger("urllib3").setLevel(logging.WARNING)
logging.getLogger("httpx").setLevel(logging.WARNING)

client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))

class APIStats:
    def __init__(self):
        self.total_requests = 0
        self.total_input_tokens = 0
        self.total_output_tokens = 0
        self.total_cached_tokens = 0
        self.processing_times = []

api_stats = APIStats()

def preprocess_ocr_text(text):
    """Preprocess OCR text to fix common errors."""
    replacements = {
        "Sv'cink i^idam": "Frank Adam",
        # add more common OCR errors here as needed
    }
    
    for error, correction in replacements.items():
        text = text.replace(error, correction)
    
    # Remove unusual characters
    text = re.sub(r'[^\w\s.,;:!?()-]', '', text)
    
    return text

def prepare_batch_requests(all_files, model_name):
    """Prepare all requests for batch processing."""
    batch_requests = []
    custom_id_mapping = {}
    
    for i, (folder_name, page_number, file_path, content) in enumerate(all_files):
        # Preprocess the OCR text
        content = preprocess_ocr_text(content)
        
        # Skip empty content
        if not content.strip():
            continue
            
        # Use the prompts module
        prompt, prompt_type = SouthernArchitectPrompts.determine_prompt_type(content, page_number, file_path)

        # Create request data
        request_data = {
            "model": model_name,
            "messages": [{
                "role": "system", 
                "content": "You are an AI archival expert tasked with cleaning OCR text and extracting metadata from it."
            }, {
                "role": "user",
                "content": f"{prompt}\n\nHere's the content to analyze:\n\n{content.strip()}\n\nNote: This content is {'short (less than 250 characters)' if prompt_type == 'short' else 'of normal length' if prompt_type == 'normal' else 'from a cover page'}."
            }],
            "max_tokens": 3000
        }
        
        batch_requests.append(request_data)
        custom_id_mapping[f"southern_architect_text_{i}"] = {
            "folder_name": folder_name,
            "page_number": page_number,
            "file_path": file_path,
            "content": content,
            "prompt_type": prompt_type,
            "row_number": i + 2  # +2 for header row
        }
    
    return batch_requests, custom_id_mapping

def collect_all_files(input_folder):
    """Collect all text files to process."""
    all_files = []
    
    def folder_sort_key(x):
        parts = x.split('-')
        return tuple(parts + [''] * (3 - len(parts)))

    # Sort folders
    sorted_folders = sorted(os.listdir(input_folder), key=folder_sort_key, reverse=False)
    
    for folder_name in sorted_folders:
        folder_path = os.path.join(input_folder, folder_name)
        if os.path.isdir(folder_path):
            # Sort files within the folder by page number
            text_files = sorted(
                [f for f in os.listdir(folder_path) if f.endswith('.txt')],
                key=lambda x: int(re.search(r'page(\d+)', x).group(1))
            )
            
            for filename in text_files:
                page_number = int(re.search(r'page(\d+)', filename).group(1))
                file_path = os.path.join(folder_path, filename)
                
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                    all_files.append((folder_name, page_number, file_path, content))
                except Exception as e:
                    logging.error(f"Error reading file {file_path}: {e}")
                    continue
    
    return all_files

def postprocess_api_response(response_data):
    """Post-process the API response for consistency - updated for geographic entities."""
    
    # Helper function to convert string to list if needed
    def ensure_list(field_value):
        if isinstance(field_value, str):
            # Split by comma and clean up each item
            items = [item.strip() for item in field_value.split(',') if item.strip()]
            return items
        elif isinstance(field_value, list):
            return field_value
        else:
            return []
    
    # Handle named entities
    if 'namedEntities' in response_data:
        response_data['namedEntities'] = ensure_list(response_data['namedEntities'])
        # Remove duplicates while preserving order
        response_data['namedEntities'] = list(dict.fromkeys(response_data['namedEntities']))
        # Remove any entities that are just single letters or numbers
        response_data['namedEntities'] = [entity for entity in response_data['namedEntities'] 
                                        if len(entity) > 1 or not entity.isalnum()]
    
    # Handle geographic entities
    if 'geographicEntities' in response_data:
        response_data['geographicEntities'] = ensure_list(response_data['geographicEntities'])
        # Remove duplicates while preserving order
        response_data['geographicEntities'] = list(dict.fromkeys(response_data['geographicEntities']))
        # Remove any entities that are just single letters or numbers
        response_data['geographicEntities'] = [entity for entity in response_data['geographicEntities'] 
                                             if len(entity) > 1 or not entity.isalnum()]
    
    # Handle topics field (also ensure it's a list)
    if 'topics' in response_data:
        response_data['topics'] = ensure_list(response_data['topics'])
    
    # Handle subjects field variations - convert all to 'topics'
    if 'subjects' in response_data and 'topics' not in response_data:
        response_data['topics'] = ensure_list(response_data.pop('subjects'))
    elif 'subjectHeadings' in response_data and 'topics' not in response_data:
        response_data['topics'] = ensure_list(response_data.pop('subjectHeadings'))
    
    # Ensure 'contentWarning' field exists and is properly formatted
    if 'contentWarning' not in response_data:
        response_data['contentWarning'] = 'None'
    elif response_data['contentWarning'].lower() == 'none' or response_data['contentWarning'].strip() == '':
        response_data['contentWarning'] = 'None'
    else:
        # Capitalize the first letter and ensure it ends with a period
        response_data['contentWarning'] = response_data['contentWarning'].capitalize().rstrip('.') + '.'
    
    return response_data

def parse_json_response(raw_response):
    """Enhanced JSON parsing with trailing comma handling."""
    try:
        # First, try standard parsing after cleaning markdown
        cleaned_response = re.sub(r'```json\s*|\s*```', '', raw_response)
        cleaned_response = re.sub(r'^[^{]*({.*})[^}]*$', r'\1', cleaned_response, flags=re.DOTALL)
        
        # Remove trailing commas that break JSON parsing
        # This regex finds commas followed by whitespace and then ] or }
        cleaned_response = re.sub(r',(\s*[}\]])', r'\1', cleaned_response)
        
        parsed_json = json.loads(cleaned_response)
        return parsed_json, None
        
    except Exception as e:
        try:
            # Second attempt: extract JSON object and fix trailing commas
            match = re.search(r'{.*}', raw_response, re.DOTALL)
            if not match:
                raise ValueError("No JSON object found in response")
                
            json_str = match.group(0)
            
            # Remove trailing commas more aggressively
            json_str = re.sub(r',(\s*[}\]])', r'\1', json_str)
            
            # Fix common JSON issues
            json_str = re.sub(r'(\w+):', r'"\1":', json_str)  # Add quotes to unquoted keys
            json_str = re.sub(r':\s*([^",\[\{][^,\]\}]*)', r': "\1"', json_str)  # Quote unquoted string values
            
            parsed_json = json.loads(json_str)
            return parsed_json, None
            
        except Exception as e2:
            return None, f"JSON parsing failed: {str(e2)}"

@tenacity.retry(
    wait=tenacity.wait_exponential(multiplier=1, min=4, max=10),
    stop=tenacity.stop_after_attempt(3),
    retry=tenacity.retry_if_exception_type(Exception)
)

def process_single_file(file_path, folder_name, page_number, content, model_name="gpt-4o-2024-08-06"):
    """Process a single text file (for individual processing)."""
    # Preprocess the OCR text
    content = preprocess_ocr_text(content)
    
    # Check if the content is empty or only whitespace
    if not content.strip():
        return {
            "cleanedText": "",
            "tocEntry": "No readable text on this page",
            "namedEntities": [],
            "geographicEntities": [], 
            "topics": [],
            "contentWarning": "None"
        }, "", None, 0
    
    # Use prompts module
    prompt, prompt_type = SouthernArchitectPrompts.determine_prompt_type(content, page_number, file_path)
    
    api_stats.total_requests += 1
    start_time = time.time()
    
    response = client.chat.completions.create(
        model=model_name,
        messages=[{
            "role": "system", 
            "content": "You are an AI archival expert tasked with cleaning OCR text and extracting metadata from it."
        }, {
            "role": "user",
            "content": f"{prompt}\n\nHere's the content to analyze:\n\n{content.strip()}\n\nNote: This content is {'short (less than 250 characters)' if prompt_type == 'short' else 'of normal length' if prompt_type == 'normal' else 'from a cover page'}."
        }],
        max_tokens=3000, 
        temperature=0.3 # Lower temperature for more focused responses, but still creative
    )
    
    processing_time = time.time() - start_time
    api_stats.processing_times.append(processing_time)
    
    api_stats.total_input_tokens += response.usage.prompt_tokens
    api_stats.total_output_tokens += response.usage.completion_tokens
    
    raw_response = response.choices[0].message.content.strip()
    
    # Use the new parsing function
    parsed_json, error = parse_json_response(raw_response)

    if parsed_json:
        # Handle field name variations - convert to 'topics'
        if 'subjects' in parsed_json and 'topics' not in parsed_json:
            parsed_json['topics'] = parsed_json.pop('subjects')
        elif 'subjectHeadings' in parsed_json and 'topics' not in parsed_json:
            parsed_json['topics'] = parsed_json.pop('subjectHeadings')
        
        required_fields = ['cleanedText', 'tocEntry', 'namedEntities', 'geographicEntities', 'topics', 'contentWarning']
        if not all(key in parsed_json for key in required_fields):
            # Handle missing fields gracefully
            for field in required_fields:
                if field not in parsed_json:
                    if field == 'topics':
                        # Try alternative field names
                        parsed_json[field] = parsed_json.get('subjects', parsed_json.get('subjectHeadings', []))
                    elif field in ['namedEntities', 'geographicEntities', 'topics']:
                        parsed_json[field] = []
                    else:
                        parsed_json[field] = ""
        
        # Post-process the response
        parsed_json = postprocess_api_response(parsed_json)
        
        return parsed_json, raw_response, response.usage, processing_time
    else:
        logging.error(f"JSON parsing failed for {file_path}: {error}\nRaw response: {raw_response}")
        # Raise exception to trigger retry
        raise Exception(f"JSON parsing failed: {error}")

def process_folder_individual(all_files, wb, analysis_sheet, raw_sheet, issues_sheet, logs_folder_path, model_name, all_results, output_dir):
    """Process using individual API calls."""
    items_with_issues = 0
    total_processing_time = 0
    
    for i, (folder_name, page_number, file_path, content) in enumerate(all_files):
        row_number = i + 2  # +2 for header row
        filename = os.path.basename(file_path)
        
        print(f"\n Processing file {i+1}/{len(all_files)}")
        print(f"   Folder: {folder_name}")
        print(f"   Page: {page_number}")
        print(f"   File: {filename}")
        print(f"   Progress: {((i+1)/len(all_files))*100:.1f}%")
        
        try:
            response_data, raw_response, usage, processing_time = process_single_file(
                file_path, folder_name, page_number, content, model_name
            )
            total_processing_time += processing_time
            
            # Log individual response
            log_individual_response(
                logs_folder_path=logs_folder_path,
                script_name="southern_architect_text_metadata",
                row_number=row_number,
                barcode=f"{folder_name}_page{page_number}",
                response_text=raw_response,
                model_name=model_name,
                prompt_tokens=usage.prompt_tokens if usage else 0,
                completion_tokens=usage.completion_tokens if usage else 0,
                processing_time=processing_time
            )
            
            print(f"   Processed successfully! Tokens: {(usage.prompt_tokens + usage.completion_tokens) if usage else 0:,}")
            
        except Exception as e:
            logging.error(f"Error processing {file_path}: {str(e)}")
            items_with_issues += 1
            raw_response = f"Processing error: {str(e)}"
            response_data = {
                "cleanedText": raw_response,
                "tocEntry": f"Error: {str(e)}",
                "namedEntities": [],
                "geographicEntities": [],  # NEW
                "topics": [],
                "contentWarning": "None"
            }
            usage = None
            
            # Add to issues sheet
            issues_sheet.append([filename, str(e)])
            
            # Log error
            log_individual_response(
                logs_folder_path=logs_folder_path,
                script_name="southern_architect_text_metadata",
                row_number=row_number,
                barcode=f"{folder_name}_page{page_number}",
                response_text=raw_response,
                model_name=model_name,
                prompt_tokens=0,
                completion_tokens=0,
                processing_time=0
            )
            
            print(f"   ❌ Processing failed: {str(e)}")
        
        analysis_row = [
            folder_name,
            page_number,
            filename,
            response_data.get('cleanedText', raw_response),
            response_data.get('tocEntry', raw_response),
            ', '.join(response_data.get('namedEntities', [])),
            ', '.join(response_data.get('geographicEntities', [])),  # NEW
            ', '.join(response_data.get('topics', [])),
            response_data.get('contentWarning', 'None')
        ]
        analysis_sheet.append(analysis_row)
        
        # Set alignment
        current_row = analysis_sheet.max_row
        for cell in analysis_sheet[current_row]:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Add to raw sheet
        raw_row = [folder_name, page_number, raw_response]
        raw_sheet.append(raw_row)
        
        for cell in raw_sheet[raw_sheet.max_row]:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        entry_result = {
            'folder': folder_name,
            'page_number': page_number,
            'file_path': file_path,
            'analysis': {
                'cleaned_text': response_data.get('cleanedText', raw_response),
                'toc_entry': response_data.get('tocEntry', raw_response),
                'named_entities': response_data.get('namedEntities', []),
                'geographic_entities': response_data.get('geographicEntities', []),  # NEW
                'topics': response_data.get('topics', []),
                'content_warning': response_data.get('contentWarning', 'None'),
                'raw_response': raw_response
            }
        }
        all_results.append(entry_result)
        
        # Add delay between requests
        time.sleep(1)
    
    return (wb, all_results, api_stats, len(all_files), items_with_issues, total_processing_time,
           api_stats.total_input_tokens, api_stats.total_output_tokens, False)  # False for was_batch_processed

def process_folder_with_batch(input_folder, output_dir, model_name="gpt-4o-2024-08-06"):
    """Process folder using batch processing when appropriate."""
    
    # Create logs folder
    logs_folder_path = os.path.join(output_dir, "logs")
    if not os.path.exists(logs_folder_path):
        os.makedirs(logs_folder_path)
    
    # Collect all files
    all_files = collect_all_files(input_folder)
    total_items = len(all_files)
    
    print(f"\nTEXT METADATA EXTRACTION")
    print(f"Found {total_items} text files to process")
    print(f"Starting metadata extraction using {model_name}...")
    print("-" * 50)
    
    # Initialize batch processor and check if we should use batch processing
    processor = BatchProcessor()
    use_batch = processor.should_use_batch(total_items)
    
    print(f"Processing mode: {'BATCH' if use_batch else 'INDIVIDUAL'}")
    
    # Show model pricing info
    model_info = get_model_info(model_name)
    if model_info:
        print(f"Model: {model_name}")
        print(f"Pricing: ${model_info['input_per_1k']:.5f}/1K input, ${model_info['output_per_1k']:.5f}/1K output")
        print(f"Batch discount: {model_info['batch_discount']*100:.0f}%")
    
    # Create new workbook
    wb = Workbook()
    analysis_sheet = wb.active
    analysis_sheet.title = "Analysis"
    
    # UPDATED: Add Geographic Entities column to headers
    analysis_headers = [
        'Folder', 'Page Number', 'Page Title', 'Cleaned OCR Text', 
        'TOC Entry', 'Named Entities', 'Geographic Entities', 'Topics', 'Content Warning'
    ]
    analysis_sheet.append(analysis_headers)
    analysis_sheet.freeze_panes = 'E2'
    
    # UPDATED: Add column width for Geographic Entities
    column_widths = [15, 10, 30, 50, 30, 30, 30, 30, 30]
    for i, width in enumerate(column_widths):
        analysis_sheet.column_dimensions[analysis_sheet.cell(row=1, column=i+1).column_letter].width = width
    
    # Create raw responses sheet
    raw_sheet = wb.create_sheet("Raw Responses")
    raw_headers = ['Folder', 'Page Number', 'API Response']
    raw_sheet.append(raw_headers)
    raw_sheet.freeze_panes = 'A2'
    
    for i, width in enumerate([15, 10, 120]):
        raw_sheet.column_dimensions[raw_sheet.cell(row=1, column=i+1).column_letter].width = width
    
    # Create issues sheet
    issues_sheet = wb.create_sheet("Issues")
    issues_sheet.append(["Filename", "Error"])
    
    all_results = []
    items_with_issues = 0
    
    if use_batch:
        # Batch processing
        print(f"Preparing {total_items} requests for batch processing...")
        
        # Estimate costs
        batch_requests, custom_id_mapping = prepare_batch_requests(all_files, model_name)
        cost_estimate = processor.estimate_batch_cost(batch_requests, model_name)
        
        # Simplified cost display
        print(f"Estimated cost: ${cost_estimate['batch_cost']:.4f} (${cost_estimate['savings']:.4f} savings)")
        
        # Convert to batch format
        formatted_requests = processor.create_batch_requests(batch_requests, "southern_architect_text")
        
        # Submit batch
        batch_id = processor.submit_batch(
            formatted_requests, 
            f"Southern Architect Text Metadata - {total_items} files - {datetime.now().strftime('%Y-%m-%d')}"
        )
        
        # Wait for completion
        batch_results = processor.wait_for_completion(batch_id, max_wait_hours=24, check_interval_minutes=5)
        
        if batch_results:
            # Process batch results
            processed_results = processor.process_batch_results(batch_results, custom_id_mapping)
            
            print(f"Processing batch results...")
            
            # Track tokens for logging
            api_stats.total_input_tokens = processed_results["summary"]["total_prompt_tokens"]
            api_stats.total_output_tokens = processed_results["summary"]["total_completion_tokens"]
            
            # Add results to spreadsheet
            for custom_id, result_data in processed_results["results"].items():
                if custom_id.startswith("southern_architect_text_"):
                    # Extract the index from custom_id
                    parts = custom_id.split("_")
                    if len(parts) >= 4:
                        try:
                            index = int(parts[3])
                            mapping_key = f"southern_architect_text_{index}"
                            
                            if mapping_key in custom_id_mapping:
                                folder_name = custom_id_mapping[mapping_key]["folder_name"]
                                page_number = custom_id_mapping[mapping_key]["page_number"]
                                file_path = custom_id_mapping[mapping_key]["file_path"]
                                row_number = custom_id_mapping[mapping_key]["row_number"]
                                filename = os.path.basename(file_path)
                                
                                if result_data["success"]:
                                    raw_response = result_data["content"]
                                    usage = result_data["usage"]
                                    
                                    # Parse JSON response using the new function
                                    parsed_json, error = parse_json_response(raw_response)

                                    if parsed_json:
                                        # Handle field name variations - convert to 'topics'
                                        if 'subjects' in parsed_json and 'topics' not in parsed_json:
                                            parsed_json['topics'] = parsed_json.pop('subjects')
                                        elif 'subjectHeadings' in parsed_json and 'topics' not in parsed_json:
                                            parsed_json['topics'] = parsed_json.pop('subjectHeadings')
                                        
                                        response_data = postprocess_api_response(parsed_json)
                                    else:
                                        response_data = {
                                            "cleanedText": raw_response,
                                            "tocEntry": f"Error: {error}",
                                            "namedEntities": [],
                                            "geographicEntities": [],  # NEW
                                            "topics": [],
                                            "contentWarning": "None"
                                        }
                                    
                                    # Log individual response
                                    log_individual_response(
                                        logs_folder_path=logs_folder_path,
                                        script_name="southern_architect_text_metadata",
                                        row_number=row_number,
                                        barcode=f"{folder_name}_page{page_number}",
                                        response_text=raw_response,
                                        model_name=model_name,
                                        prompt_tokens=usage.get("prompt_tokens", 0),
                                        completion_tokens=usage.get("completion_tokens", 0),
                                        processing_time=0  # Batch processing doesn't track individual timing
                                    )
                                    
                                else:
                                    # Handle error case
                                    raw_response = f"Error: {result_data['error']}"
                                    items_with_issues += 1
                                    response_data = {
                                        "cleanedText": raw_response,
                                        "tocEntry": f"Error: {result_data['error']}",
                                        "namedEntities": [],
                                        "geographicEntities": [],  # NEW
                                        "topics": [],
                                        "contentWarning": "None"
                                    }
                                    
                                    # Add to issues sheet
                                    issues_sheet.append([filename, result_data['error']])
                                    
                                    # Log error
                                    log_individual_response(
                                        logs_folder_path=logs_folder_path,
                                        script_name="southern_architect_text_metadata",
                                        row_number=row_number,
                                        barcode=f"{folder_name}_page{page_number}",
                                        response_text=raw_response,
                                        model_name=model_name,
                                        prompt_tokens=0,
                                        completion_tokens=0,
                                        processing_time=0
                                    )
                                
                                analysis_row = [
                                    folder_name,
                                    page_number,
                                    filename,
                                    response_data.get('cleanedText', raw_response),
                                    response_data.get('tocEntry', raw_response),
                                    ', '.join(response_data.get('namedEntities', [])),
                                    ', '.join(response_data.get('geographicEntities', [])),  # NEW
                                    ', '.join(response_data.get('topics', [])),
                                    response_data.get('contentWarning', 'None')
                                ]
                                analysis_sheet.append(analysis_row)
                                
                                # Set alignment
                                current_row = analysis_sheet.max_row
                                for cell in analysis_sheet[current_row]:
                                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                                
                                # Add to raw sheet
                                raw_row = [folder_name, page_number, raw_response]
                                raw_sheet.append(raw_row)
                                
                                for cell in raw_sheet[raw_sheet.max_row]:
                                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                                
                                entry_result = {
                                    'folder': folder_name,
                                    'page_number': page_number,
                                    'file_path': file_path,
                                    'analysis': {
                                        'cleaned_text': response_data.get('cleanedText', raw_response),
                                        'toc_entry': response_data.get('tocEntry', raw_response),
                                        'named_entities': response_data.get('namedEntities', []),
                                        'geographic_entities': response_data.get('geographicEntities', []),  # NEW
                                        'topics': response_data.get('topics', []),
                                        'content_warning': response_data.get('contentWarning', 'None'),
                                        'raw_response': raw_response
                                    }
                                }
                                all_results.append(entry_result)
                                
                        except (ValueError, IndexError) as e:
                            logging.error(f"Error processing custom_id {custom_id}: {e}")
                            continue
            
            # Return batch processing metrics
            summary = processed_results["summary"]
            return (wb, all_results, api_stats, total_items, items_with_issues, 0,  # 0 for total_time since batch doesn't track individual timing
                   summary["total_prompt_tokens"], summary["total_completion_tokens"], True)  # True for was_batch_processed
    
    # Fall back to individual processing
    print(f"Using individual processing:")
    return process_folder_individual(all_files, wb, analysis_sheet, raw_sheet, issues_sheet, logs_folder_path, model_name, all_results, output_dir)

def main():
    # Add argument parsing for model selection
    parser = argparse.ArgumentParser(description='Process Southern Architect texts')
    parser.add_argument('--model', default="gpt-4o-2024-08-06", help='Model name to use')
    args = parser.parse_args()
    
    model_name = args.model
    
    # Start timing the entire script execution
    script_start_time = time.time()
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_folder_name = "4_pages"  # Change this to whichever folder you want
    input_folder = os.path.join(script_dir, "image_folders", input_folder_name)
    
    # Create dynamic output folder name
    current_date = datetime.now().strftime("%Y-%m-%d")
    current_time = datetime.now().strftime("%H-%M-%S")
    
    # Create folder name
    folder_name = f"SABN_Metadata_Created_{current_date}_Time_{current_time}"

    # Create the full output directory path
    base_output_dir = os.path.join(script_dir, "output_folders")
    output_dir = os.path.join(base_output_dir, folder_name)
    
    # Create the directory
    os.makedirs(output_dir, exist_ok=True)
    
    # Create metadata folder structure
    metadata_folder = os.path.join(output_dir, "metadata", "collection_metadata")
    os.makedirs(metadata_folder, exist_ok=True)

    print(f"Output directory: {output_dir}")
    
    # Process folder
    (wb, all_results, api_stats, total_items, items_with_issues, total_processing_time,
     total_prompt_tokens, total_completion_tokens, was_batch_processed) = process_folder_with_batch(
        input_folder, output_dir, model_name
    )
    
    # Add API Stats sheet
    api_summary = {
        "total_requests": api_stats.total_requests,
        "total_input_tokens": total_prompt_tokens,
        "total_output_tokens": total_completion_tokens,
        "total_tokens": total_prompt_tokens + total_completion_tokens,
        "processing_mode": "BATCH" if was_batch_processed else "INDIVIDUAL"
    }
    
    all_results.append({"api_stats": api_summary})
    
    stats_sheet = wb.create_sheet("API Stats")
    stats_sheet.append(["Metric", "Value"])
    for key, value in api_summary.items():
        stats_sheet.append([key, value])
    
    # Save files
    excel_path = os.path.join(metadata_folder, "image_workflow.xlsx")
    json_path = os.path.join(metadata_folder, "image_workflow.json")
    
    wb.save(excel_path)
    with open(json_path, 'w') as f:
        json.dump(all_results, f, indent=2)
    
    # Calculate script metrics
    script_duration = time.time() - script_start_time
    
    # Calculate actual cost
    estimated_cost = calculate_cost(
        model_name=model_name,
        prompt_tokens=total_prompt_tokens,
        completion_tokens=total_completion_tokens,
        is_batch=was_batch_processed
    )
    
    # Create logs folder
    logs_folder_path = os.path.join(output_dir, "logs")
    if not os.path.exists(logs_folder_path):
        os.makedirs(logs_folder_path)
    
    # Create standardized token usage log
    create_token_usage_log(
        logs_folder_path=logs_folder_path,
        script_name="southern_architect_text_metadata",
        model_name=model_name,
        total_items=total_items,
        items_with_issues=items_with_issues,
        total_time=total_processing_time,
        total_prompt_tokens=total_prompt_tokens,
        total_completion_tokens=total_completion_tokens,
        additional_metrics={
            "Total script execution time": f"{script_duration:.2f}s",
            "Processing time percentage": f"{(total_processing_time/script_duration)*100:.1f}%" if script_duration > 0 else "0%",
            "Items successfully processed": total_items - items_with_issues,
            "Processing mode": "BATCH" if was_batch_processed else "INDIVIDUAL",
            "Actual cost": f"${estimated_cost:.4f}",
            "Average tokens per item": f"{(total_prompt_tokens + total_completion_tokens)/total_items:.0f}" if total_items > 0 else "0"
        }
    )
    
    # Final summary
    print(f"\n ✅ STEP 1 COMPLETE: Generated text metadata in {os.path.basename(output_dir)}")
    print(f"Excel file, JSON data, and processing logs created")
    print(f"Successfully processed: {total_items - items_with_issues}/{total_items} files")
    print(f"Items with issues: {items_with_issues}")
    print(f"Total script time: {script_duration:.1f}s ({script_duration/60:.1f} minutes)")
    print(f"Tokens: {total_prompt_tokens + total_completion_tokens:,} (Input: {total_prompt_tokens:,}, Output: {total_completion_tokens:,})")
    print(f"Processing mode: {'BATCH' if was_batch_processed else 'INDIVIDUAL'}")
    print(f"Cost estimate: ${estimated_cost:.4f}")
    

if __name__ == "__main__":
    main()