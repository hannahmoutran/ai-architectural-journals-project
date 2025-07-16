"""
Checkpoint and graceful shutdown management for long-running processing scripts.

This module provides functionality for:
- Saving and loading processing checkpoints
- Graceful shutdown handling with signal management
- Excel workbook state management
- Progress recovery and resumption

Usage:
    from checkpoint_manager import CheckpointManager, GracefulShutdown
    
    # Setup shutdown handler
    shutdown_handler = GracefulShutdown()
    
    # Setup checkpoint manager
    checkpoint_mgr = CheckpointManager(output_dir)
    
    # In processing loop
    if shutdown_handler.shutdown_requested:
        checkpoint_mgr.save_checkpoint(results, current_index, api_stats)
        sys.exit(0)
"""

import os
import json
import logging
import signal
import sys
import shutil
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment


class GracefulShutdown:
    """Handles graceful shutdown signals (Ctrl+C, SIGTERM) for long-running processes."""
    
    def __init__(self):
        self.shutdown_requested = False
        signal.signal(signal.SIGINT, self._handle_signal)
        signal.signal(signal.SIGTERM, self._handle_signal)
    
    def _handle_signal(self, signum, frame):
        """Handle shutdown signals and set shutdown flag."""
        logging.info("Received shutdown signal, saving progress...")
        self.shutdown_requested = True


class CheckpointManager:
    """Manages checkpoint saving/loading and Excel workbook state for processing recovery."""
    
    def __init__(self, output_dir: str, checkpoint_filename: str = "processing_checkpoint.json"):
        self.output_dir = output_dir
        self.checkpoint_file = os.path.join(output_dir, checkpoint_filename)
        self.excel_path = os.path.join(output_dir, "text_workflow.xlsx")
    
    def save_checkpoint(self, processed_items: List[Dict], current_index: int, 
                       api_stats: Any, batch_id: Optional[str] = None) -> bool:
        """
        Save current progress to a checkpoint file.
        
        Args:
            processed_items: List of processed results
            current_index: Current processing index
            api_stats: API statistics object with token counts
            batch_id: Optional batch ID for batch processing
            
        Returns:
            bool: True if save successful, False otherwise
        """
        checkpoint_data = {
            'processed_items': len(processed_items),
            'current_index': current_index,
            'timestamp': datetime.now().isoformat(),
            'total_input_tokens': getattr(api_stats, 'total_input_tokens', 0),
            'total_output_tokens': getattr(api_stats, 'total_output_tokens', 0),
            'batch_id': batch_id,
            'results': processed_items
        }
        
        try:
            with open(self.checkpoint_file, 'w') as f:
                json.dump(checkpoint_data, f, indent=2)
            logging.info(f"Checkpoint saved: {len(processed_items)} items processed")
            return True
        except Exception as e:
            logging.error(f"Failed to save checkpoint: {e}")
            return False
    
    def load_checkpoint(self) -> Optional[Dict]:
        """
        Load previous progress from checkpoint file.
        
        Returns:
            dict: Checkpoint data if found, None otherwise
        """
        if os.path.exists(self.checkpoint_file):
            try:
                with open(self.checkpoint_file, 'r') as f:
                    checkpoint_data = json.load(f)
                logging.info(f"Checkpoint loaded: {checkpoint_data['processed_items']} items previously processed")
                return checkpoint_data
            except Exception as e:
                logging.error(f"Failed to load checkpoint: {e}")
                return None
        return None
    
    def save_excel_with_backup(self, wb: Workbook, create_backup: bool = False) -> bool:
        """
        Save Excel workbook with optional backup.
        
        Args:
            wb: OpenPyXL Workbook object
            create_backup: Whether to create backup before overwriting
            
        Returns:
            bool: True if save successful, False otherwise
        """
        try:
            if create_backup and os.path.exists(self.excel_path):
                backup_path = self.excel_path.replace('.xlsx', f'_backup_{int(datetime.now().timestamp())}.xlsx')
                shutil.copy2(self.excel_path, backup_path)
                logging.info(f"Created backup: {backup_path}")
            
            wb.save(self.excel_path)
            return True
        except Exception as e:
            logging.error(f"Failed to save Excel file: {e}")
            return False
    
    def load_existing_workbook(self) -> tuple[Optional[Workbook], bool]:
        """
        Load existing workbook if it exists.
        
        Returns:
            tuple: (Workbook object or None, success boolean)
        """
        if os.path.exists(self.excel_path):
            try:
                wb = load_workbook(self.excel_path)
                logging.info("Loaded existing workbook")
                return wb, True
            except Exception as e:
                logging.error(f"Failed to load existing workbook: {e}")
                return None, False
        return None, False
    
    def restore_workbook_from_checkpoint(self, checkpoint_data: Dict, wb: Workbook, 
                                       analysis_sheet, raw_sheet) -> bool:
        """
        Restore workbook state from checkpoint data.
        
        Args:
            checkpoint_data: Checkpoint data containing results
            wb: Workbook object to restore to
            analysis_sheet: Analysis worksheet
            raw_sheet: Raw responses worksheet
            
        Returns:
            bool: True if restore successful, False otherwise
        """
        try:
            for result in checkpoint_data['results']:
                if 'analysis' in result:
                    # Add to analysis sheet
                    analysis_row = [
                        result['folder'],
                        result['page_number'],
                        os.path.basename(result['file_path']),
                        result['analysis']['cleaned_text'],
                        result['analysis']['toc_entry'],
                        ', '.join(result['analysis']['named_entities']),
                        ', '.join(result['analysis']['subject_headings']),
                        result['analysis']['content_warning']
                    ]
                    analysis_sheet.append(analysis_row)
                    
                    # Set alignment
                    current_row = analysis_sheet.max_row
                    for cell in analysis_sheet[current_row]:
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                    
                    # Add to raw sheet
                    raw_row = [result['folder'], result['page_number'], result['analysis']['raw_response']]
                    raw_sheet.append(raw_row)
                    
                    for cell in raw_sheet[raw_sheet.max_row]:
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            logging.info(f"Restored {len(checkpoint_data['results'])} items to workbook")
            return True
        except Exception as e:
            logging.error(f"Failed to restore workbook from checkpoint: {e}")
            return False
    
    def cleanup_checkpoint(self) -> bool:
        """
        Remove checkpoint file after successful completion.
        
        Returns:
            bool: True if cleanup successful, False otherwise
        """
        try:
            if os.path.exists(self.checkpoint_file):
                os.remove(self.checkpoint_file)
                logging.info("Checkpoint file cleaned up (processing complete)")
                return True
            return True  # File doesn't exist, so cleanup is "successful"
        except Exception as e:
            logging.error(f"Failed to cleanup checkpoint file: {e}")
            return False
    
    def find_resume_folder(self, base_output_dir: str, model_prefix: str) -> Optional[str]:
        """
        Find the most recent output folder with a checkpoint for resuming.
        
        Args:
            base_output_dir: Base directory containing output folders
            model_prefix: Model name prefix to search for
            
        Returns:
            str: Path to resume folder, or None if not found
        """
        try:
            if not os.path.exists(base_output_dir):
                return None
            
            existing_folders = []
            for folder in os.listdir(base_output_dir):
                if folder.startswith(model_prefix + "_TEXT_Created_"):
                    folder_path = os.path.join(base_output_dir, folder)
                    checkpoint_file = os.path.join(folder_path, "processing_checkpoint.json")
                    if os.path.exists(checkpoint_file):
                        existing_folders.append((folder, os.path.getctime(folder_path)))
            
            if existing_folders:
                # Return the most recent folder with a checkpoint
                most_recent_folder = max(existing_folders, key=lambda x: x[1])[0]
                return os.path.join(base_output_dir, most_recent_folder)
            
            return None
        except Exception as e:
            logging.error(f"Error finding resume folder: {e}")
            return None


def create_checkpoint_aware_loop(items: List[Any], checkpoint_mgr: CheckpointManager, 
                                shutdown_handler: GracefulShutdown, 
                                process_func: callable, save_interval: int = 10) -> List[Any]:
    """
    Helper function to create a checkpoint-aware processing loop.
    
    Args:
        items: List of items to process
        checkpoint_mgr: CheckpointManager instance
        shutdown_handler: GracefulShutdown instance
        process_func: Function to process each item
        save_interval: How often to save checkpoints (default: every 10 items)
        
    Returns:
        List of processed results
    """
    results = []
    
    for i, item in enumerate(items):
        # Check for shutdown signal
        if shutdown_handler.shutdown_requested:
            logging.info("Saving progress before shutdown...")
            checkpoint_mgr.save_checkpoint(results, i, None)  # api_stats would be passed in real usage
            print(f"\n💾 Progress saved at item {i}/{len(items)}")
            print(f"🔄 To resume, run with --resume flag")
            sys.exit(0)
        
        # Process the item
        try:
            result = process_func(item, i)
            results.append(result)
        except Exception as e:
            logging.error(f"Error processing item {i}: {e}")
            continue
        
        # Save checkpoint at intervals
        if (i + 1) % save_interval == 0:
            checkpoint_mgr.save_checkpoint(results, i + 1, None)
            print(f"   💾 Checkpoint saved at item {i+1}/{len(items)}")
    
    # Cleanup checkpoint when complete
    checkpoint_mgr.cleanup_checkpoint()
    return results