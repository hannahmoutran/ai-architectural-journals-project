#!/usr/bin/env python3
"""
SA Integrate Reviewer Edits
============================

Reads reviewer decisions exported from sa_html_review.py and applies them back
into the workflow JSON, then generates final output files.

What this script does:
1. Finds the latest decisions export in review/exports/ (or use --decisions)
2. Loads image_workflow.json or text_workflow.json (auto-detected)
3. Backs up the original JSON to original-outputs/ (only on first run)
4. Applies field edits (text_transcription, visual_description, toc_entry,
   content_warning, named_entities, geographic_entities, topics)
5. Applies term decisions for subject headings, geographic headings,
   chronological terms, and custom terms
6. Saves updated workflow JSON
7. Generates metadata/collection_metadata/final_metadata.json
8. Generates metadata/collection_metadata/edit_statistics_report.json
9. Generates metadata/collection_metadata/edit_changelog.json
10. Creates metadata/collection_metadata/final_deliverable.xlsx

Usage:
    python sa_integrate_edits.py
    python sa_integrate_edits.py --decisions path/to/decisions.json
    python sa_integrate_edits.py --folder /path/to/output_folder
    python sa_integrate_edits.py --yes
"""

import os
import sys
import json
import shutil
import argparse
import re
from datetime import datetime
from difflib import SequenceMatcher

script_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, script_dir)

from shared_utilities import find_newest_folder
from sa_workflow_config import FOLDER_CONFIG


def detect_workflow_type(folder_path):
    """Return (workflow_type, json_path) for the folder, or (None, None)."""
    collection_metadata_dir = os.path.join(folder_path, "metadata", "collection_metadata")
    image_json = os.path.join(collection_metadata_dir, "image_workflow.json")
    text_json = os.path.join(collection_metadata_dir, "text_workflow.json")
    if os.path.exists(image_json):
        return "image", image_json
    if os.path.exists(text_json):
        return "text", text_json
    return None, None


class SAEditsIntegrator:
    """Integrates reviewer edits from the HTML review interface back into workflow files."""

    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.folder_name = os.path.basename(folder_path)
        self.collection_metadata_dir = os.path.join(folder_path, "metadata", "collection_metadata")
        self.review_folder = os.path.join(folder_path, "review")
        self.exports_folder = os.path.join(self.review_folder, "exports")
        self.original_outputs_folder = os.path.join(folder_path, "original-outputs")

        self.logs_folder = os.path.join(folder_path, "logs")

        self.workflow_type = None
        self.workflow_json_path = None
        self.workflow_data = None
        self.issue_syntheses_entry = None
        self.decisions_data = None
        self.edit_history = []

        self.stats = {
            'reviewer_name': '',
            'export_timestamp': '',
            'integration_timestamp': '',
            'total_records_in_export': 0,
            'total_records_in_batch': 0,
            'records_with_edits': 0,
            'records_reviewed_only': 0,
            'total_field_edits': 0,
            'edits_by_field': {},
            'text_field_metrics': {},
            'list_field_metrics': {},
            'subject_headings_total': 0,
            'subject_headings_approved': 0,
            'subject_headings_rejected': 0,
            'alt_headings_added': 0,
            'geo_headings_added': 0,
            'chrono_terms_total': 0,
            'chrono_terms_approved': 0,
            'chrono_terms_rejected': 0,
            'custom_terms_added': 0,
        }

    # ------------------------------------------------------------------
    # Finding and loading data
    # ------------------------------------------------------------------

    def find_latest_export(self):
        """Return the path to the most recent decisions JSON in exports/."""
        if not os.path.exists(self.exports_folder):
            print("Error: exports folder not found at " + self.exports_folder)
            print("Run sa_html_review.py first, review records, and export decisions.")
            return None

        json_files = [f for f in os.listdir(self.exports_folder) if f.endswith('.json')]
        if not json_files:
            print("Error: No JSON files found in " + self.exports_folder)
            print("Export decisions from the HTML review interface first.")
            return None

        json_files.sort(
            key=lambda f: os.path.getmtime(os.path.join(self.exports_folder, f)),
            reverse=True
        )
        return os.path.join(self.exports_folder, json_files[0])

    def load_decisions(self, decisions_path):
        """Load reviewer decisions from the exported JSON file."""
        try:
            with open(decisions_path, 'r', encoding='utf-8') as f:
                self.decisions_data = json.load(f)

            self.stats['reviewer_name'] = self.decisions_data.get('reviewer_name', 'Unknown')
            self.stats['export_timestamp'] = self.decisions_data.get('export_timestamp', '')
            decision_list = self.decisions_data.get('decisions', [])
            self.stats['total_records_in_export'] = len(decision_list)
            self.stats['total_records_in_batch'] = self.decisions_data.get(
                'total_records', self.stats['total_records_in_export']
            )

            print("Loaded " + str(self.stats['total_records_in_export']) + " record decisions")
            print("Reviewer: " + self.stats['reviewer_name'])
            print("Export timestamp: " + self.stats['export_timestamp'])
            return True

        except Exception as exc:
            print("Error loading decisions: " + str(exc))
            return False

    def load_workflow_json(self):
        """Auto-detect workflow type and load the JSON file."""
        workflow_type, json_path = detect_workflow_type(self.folder_path)
        if not workflow_type:
            print("Error: Neither image_workflow.json nor text_workflow.json found.")
            return False

        self.workflow_type = workflow_type
        self.workflow_json_path = json_path

        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                raw_data = json.load(f)

            if (raw_data
                    and isinstance(raw_data[-1], dict)
                    and 'issue_syntheses' in raw_data[-1]):
                self.issue_syntheses_entry = raw_data[-1]
                self.workflow_data = raw_data[:-1]
            else:
                self.issue_syntheses_entry = None
                self.workflow_data = raw_data

            print("Loaded " + self.workflow_type + " workflow with "
                  + str(len(self.workflow_data)) + " page records")
            return True

        except Exception as exc:
            print("Error loading workflow JSON: " + str(exc))
            return False

    # ------------------------------------------------------------------
    # Backup
    # ------------------------------------------------------------------

    def backup_original_files(self):
        """Back up all files in the metadata/ subtree to original-outputs/metadata/...

        Preserves the subfolder structure of metadata/ exactly. Only copies files
        that do not already exist in the backup location -- so the first run captures
        the pre-edit originals and subsequent runs leave them untouched.
        """
        os.makedirs(self.original_outputs_folder, exist_ok=True)

        metadata_dir = os.path.join(self.folder_path, "metadata")
        if not os.path.exists(metadata_dir):
            print("  Warning: metadata/ folder not found, nothing to back up.")
            return

        backed_up = 0
        skipped = 0
        for root, _dirs, files in os.walk(metadata_dir):
            rel = os.path.relpath(root, self.folder_path)
            dest_dir = os.path.join(self.original_outputs_folder, rel)
            os.makedirs(dest_dir, exist_ok=True)
            for fname in files:
                src = os.path.join(root, fname)
                dest = os.path.join(dest_dir, fname)
                if os.path.exists(dest):
                    skipped += 1
                else:
                    shutil.copy2(src, dest)
                    print("  Backed up: " + os.path.join(rel, fname))
                    backed_up += 1

        if backed_up == 0 and skipped > 0:
            print("  Backup already exists (" + str(skipped) + " files preserved from first run).")
        elif backed_up > 0:
            print("  Backed up " + str(backed_up) + " file(s) to original-outputs/")

    # ------------------------------------------------------------------
    # Diff helpers
    # ------------------------------------------------------------------

    def _text_diff_metrics(self, original, new_value):
        """Compute character-level diff metrics between two strings."""
        original = str(original) if original else ""
        new_value = str(new_value) if new_value else ""

        matcher = SequenceMatcher(None, original, new_value)
        chars_added = 0
        chars_deleted = 0
        chars_unchanged = 0

        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            if tag == 'equal':
                chars_unchanged += (i2 - i1)
            elif tag == 'delete':
                chars_deleted += (i2 - i1)
            elif tag == 'insert':
                chars_added += (j2 - j1)
            elif tag == 'replace':
                chars_deleted += (i2 - i1)
                chars_added += (j2 - j1)

        return {
            'chars_added': chars_added,
            'chars_deleted': chars_deleted,
            'chars_unchanged': chars_unchanged,
            'chars_total_changed': chars_added + chars_deleted,
            'original_length': len(original),
            'new_length': len(new_value),
            'similarity_ratio': round(matcher.ratio(), 4),
        }

    def _list_diff_metrics(self, original_list, new_list):
        """Compute item-level diff metrics between two lists."""
        if not isinstance(original_list, list):
            original_list = [original_list] if original_list else []
        if not isinstance(new_list, list):
            new_list = [new_list] if new_list else []

        def normalize(item):
            return json.dumps(item, sort_keys=True) if isinstance(item, dict) else str(item)

        orig_set = set(normalize(i) for i in original_list if i)
        new_set = set(normalize(i) for i in new_list if i)

        return {
            'items_added': len(new_set - orig_set),
            'items_removed': len(orig_set - new_set),
            'items_kept': len(orig_set & new_set),
        }

    def _truncate(self, value, max_length=120):
        """Truncate a value for display in the changelog."""
        if value is None:
            return ""
        if isinstance(value, list):
            value = json.dumps(value)
        value = str(value)
        return value[:max_length] + "..." if len(value) > max_length else value

    def _text_diff_hunks(self, original, new_value, context=60):
        """Return a list of diff hunks showing exactly what changed in a text field.

        Each hunk is a dict with keys:
            context_before  -- text immediately before the changed region (from original)
            removed         -- text that was removed (absent for pure insertions)
            added           -- text that was inserted (absent for pure deletions)
            context_after   -- text immediately after the changed region (from new_value)
        """
        original = str(original) if original else ""
        new_value = str(new_value) if new_value else ""

        matcher = SequenceMatcher(None, original, new_value)
        hunks = []
        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            if tag == 'equal':
                continue

            ctx_start = max(0, i1 - context)
            ctx_before = ('...' if ctx_start > 0 else '') + original[ctx_start:i1]

            ctx_end = min(len(new_value), j2 + context)
            ctx_after = new_value[j2:ctx_end] + ('...' if ctx_end < len(new_value) else '')

            hunk = {'context_before': ctx_before, 'context_after': ctx_after}
            if tag in ('delete', 'replace'):
                hunk['removed'] = original[i1:i2]
            if tag in ('insert', 'replace'):
                hunk['added'] = new_value[j1:j2]
            hunks.append(hunk)

        return hunks

    def _build_term_label_map(self, analysis):
        """Build a mapping from term_id string to a human-readable label for one record.

        Covers selected-N, alt-N, geo-N, and chrono-N identifiers, using the same
        ordering logic the HTML review page uses when assigning those IDs.
        """
        label_map = {}

        # selected-N: AI-chosen subject headings
        for idx, term in enumerate(analysis.get('final_selected_terms', [])):
            label = term.get('label', '')
            source = term.get('source', '')
            label_map['selected-' + str(idx)] = label + (' [' + source + ']' if source else '')

        # chrono-N: chronological vocabulary terms
        for idx, term in enumerate(analysis.get('chronological_vocabulary_terms', [])):
            label = term.get('label', '')
            source = term.get('source', '')
            label_map['chrono-' + str(idx)] = label + (' [' + source + ']' if source else '')

        # alt-N: alternative subject headings (non-selected vocab results, grouped by source)
        vocab_results = analysis.get('vocabulary_search_results', {})
        final_terms = analysis.get('final_selected_terms', [])
        selected_pairs = set(
            (t.get('label', '').lower(), t.get('source', '')) for t in final_terms
        )
        sources_order = ['LCSH', 'FAST', 'Getty AAT', 'Getty TGN']
        sources_map = {s: [] for s in sources_order}
        for orig_topic, matches in vocab_results.items():
            for match in matches:
                source = match.get('source', 'Unknown')
                label = match.get('label', '')
                if (label.lower(), source) in selected_pairs:
                    continue
                if source in sources_map:
                    sources_map[source].append({'label': label, 'source': source})
        flat_alt = []
        for source_name in sources_order:
            flat_alt.extend(sources_map[source_name])
        for idx, term in enumerate(flat_alt):
            label = term['label']
            source = term['source']
            label_map['alt-' + str(idx)] = label + (' [' + source + ']' if source else '')

        # geo-N: geographic vocabulary terms
        geo_vocab = analysis.get('geographic_vocabulary_search_results', {})
        flat_geo = []
        for source_entity, matches in geo_vocab.items():
            for match in matches:
                flat_geo.append({
                    'label': match.get('label', ''),
                    'source': match.get('source', 'FAST Geographic'),
                })
        for idx, term in enumerate(flat_geo):
            label = term['label']
            source = term['source']
            label_map['geo-' + str(idx)] = label + (' [' + source + ']' if source else '')

        return label_map

    # ------------------------------------------------------------------
    # Applying edits
    # ------------------------------------------------------------------

    TEXT_FIELDS = {'text_transcription', 'visual_description', 'toc_entry', 'content_warning'}
    LIST_FIELDS = {'named_entities', 'geographic_entities', 'topics'}

    def _apply_field_edit(self, record, field_name, edit_data, record_id):
        """Apply a single field edit to the record's analysis dict."""
        analysis = record.setdefault('analysis', {})
        original_value = edit_data.get('original', '')
        new_value = edit_data.get('value', '')

        analysis[field_name] = new_value

        if field_name in self.TEXT_FIELDS:
            metrics = self._text_diff_metrics(original_value, new_value)
            if field_name not in self.stats['text_field_metrics']:
                self.stats['text_field_metrics'][field_name] = {
                    'edits_count': 0,
                    'chars_added': 0,
                    'chars_deleted': 0,
                    'chars_total_changed': 0,
                    'total_original_length': 0,
                    'total_new_length': 0,
                    'similarity_sum': 0.0,
                }
            fm = self.stats['text_field_metrics'][field_name]
            fm['edits_count'] += 1
            fm['chars_added'] += metrics['chars_added']
            fm['chars_deleted'] += metrics['chars_deleted']
            fm['chars_total_changed'] += metrics['chars_total_changed']
            fm['total_original_length'] += metrics['original_length']
            fm['total_new_length'] += metrics['new_length']
            fm['similarity_sum'] += metrics['similarity_ratio']
        elif field_name in self.LIST_FIELDS:
            metrics = self._list_diff_metrics(original_value, new_value)
            if field_name not in self.stats['list_field_metrics']:
                self.stats['list_field_metrics'][field_name] = {
                    'edits_count': 0,
                    'items_added': 0,
                    'items_removed': 0,
                    'items_kept': 0,
                }
            lm = self.stats['list_field_metrics'][field_name]
            lm['edits_count'] += 1
            lm['items_added'] += metrics['items_added']
            lm['items_removed'] += metrics['items_removed']
            lm['items_kept'] += metrics['items_kept']
        else:
            metrics = {}

        entry = {
            'record_id': record_id,
            'field': field_name,
            'original': self._truncate(original_value),
            'new_value': self._truncate(new_value),
            'edit_type': 'field_edit',
            'metrics': metrics,
        }
        if field_name in self.TEXT_FIELDS:
            diff_hunks = self._text_diff_hunks(original_value, new_value)
            if diff_hunks:
                entry['diff_hunks'] = diff_hunks
        self.edit_history.append(entry)

        self.stats['total_field_edits'] += 1
        self.stats['edits_by_field'][field_name] = (
            self.stats['edits_by_field'].get(field_name, 0) + 1
        )

    def _apply_term_decisions(self, record, term_decisions, custom_terms, record_id, analysis=None):
        """Apply vocabulary term decisions and custom terms to the record."""
        if analysis is None:
            analysis = record.get('analysis', {})
        record.setdefault('analysis', analysis)

        term_label_map = self._build_term_label_map(analysis)

        # Count decisions by prefix to update aggregate stats later
        for term_id, status in term_decisions.items():
            if not isinstance(status, str):
                continue
            if term_id.startswith('selected-') and status == 'rejected':
                self.stats['subject_headings_rejected'] += 1
            elif term_id.startswith('alt-') and status == 'approved':
                self.stats['alt_headings_added'] += 1
            elif term_id.startswith('geo-') and status == 'approved':
                self.stats['geo_headings_added'] += 1
            elif term_id.startswith('chrono-') and status == 'rejected':
                self.stats['chrono_terms_rejected'] += 1

            self.edit_history.append({
                'record_id': record_id,
                'field': 'vocabulary_term',
                'term_id': term_id,
                'term_label': term_label_map.get(term_id, term_id),
                'decision': status,
                'edit_type': 'term_decision',
                'metrics': {},
            })

        # Store decisions on the record for use in final metadata generation
        analysis['reviewer_term_decisions'] = term_decisions

        # Custom terms
        if custom_terms:
            analysis['reviewer_custom_terms'] = custom_terms
            self.stats['custom_terms_added'] += len(custom_terms)
            for term in custom_terms:
                self.edit_history.append({
                    'record_id': record_id,
                    'field': 'custom_term',
                    'original': '',
                    'new_value': term.get('label', '') + ' (' + term.get('source', '') + ')',
                    'edit_type': 'custom_term_added',
                    'metrics': {},
                })

    def apply_all_edits(self):
        """Iterate over all decisions and apply them to workflow_data."""
        if not self.decisions_data or self.workflow_data is None:
            return False

        for decision in self.decisions_data.get('decisions', []):
            record_id = decision.get('record_id')
            record_idx = record_id - 1

            if record_idx < 0 or record_idx >= len(self.workflow_data):
                print("  Warning: record_id " + str(record_id) + " not found in workflow data (skipping)")
                continue

            record = self.workflow_data[record_idx]
            has_edits = False

            # Field edits
            for field_name, edit_data in decision.get('edits', {}).items():
                if edit_data.get('edited', False):
                    self._apply_field_edit(record, field_name, edit_data, record_id)
                    has_edits = True

            # Term decisions and custom terms
            term_decisions = decision.get('term_decisions', {})
            custom_terms = decision.get('custom_terms', [])
            if term_decisions or custom_terms:
                pre_edit_analysis = record.get('analysis', {})
                self._apply_term_decisions(record, term_decisions, custom_terms, record_id, pre_edit_analysis)
                has_edits = True

            # Reviewer metadata
            analysis = record.setdefault('analysis', {})
            analysis['reviewer_reviewed'] = decision.get('reviewed', False)
            analysis['reviewer_notes'] = decision.get('reviewer_notes', '')
            analysis['reviewer_name'] = self.stats['reviewer_name']
            analysis['reviewer_date'] = self.stats['export_timestamp']

            if has_edits:
                self.stats['records_with_edits'] += 1
            elif decision.get('reviewed', False):
                self.stats['records_reviewed_only'] += 1

        # Compute aggregate vocabulary counts from the workflow data
        self._compute_aggregate_vocab_counts()
        return True

    def _compute_aggregate_vocab_counts(self):
        """Tally total subject headings and chronological terms across all records."""
        for record in self.workflow_data:
            analysis = record.get('analysis', {})
            self.stats['subject_headings_total'] += len(analysis.get('final_selected_terms', []))
            self.stats['chrono_terms_total'] += len(analysis.get('chronological_vocabulary_terms', []))

        self.stats['subject_headings_approved'] = (
            self.stats['subject_headings_total'] - self.stats['subject_headings_rejected']
        )
        self.stats['chrono_terms_approved'] = (
            self.stats['chrono_terms_total'] - self.stats['chrono_terms_rejected']
        )

    # ------------------------------------------------------------------
    # Saving workflow JSON
    # ------------------------------------------------------------------

    def save_workflow_json(self):
        """Write the updated workflow data back to disk."""
        output_data = list(self.workflow_data)
        if self.issue_syntheses_entry is not None:
            output_data.append(self.issue_syntheses_entry)

        try:
            with open(self.workflow_json_path, 'w', encoding='utf-8') as f:
                json.dump(output_data, f, indent=2, ensure_ascii=False)
            print("  Saved updated " + os.path.basename(self.workflow_json_path))
            return True
        except Exception as exc:
            print("Error saving workflow JSON: " + str(exc))
            return False

    # ------------------------------------------------------------------
    # Resolving approved geo terms (same index order as HTML review)
    # ------------------------------------------------------------------

    def _get_approved_geo_terms(self, analysis, term_decisions):
        """Return geographic terms approved via 'geo-N' decisions.

        The HTML review page iterates geographic_vocabulary_search_results in
        insertion order, assigning geo-0, geo-1, ... across all source entities
        and their matches. This method replicates that ordering.
        """
        geo_vocab = analysis.get('geographic_vocabulary_search_results', {})
        if not geo_vocab:
            return []

        # Build an ordered flat list matching the HTML page render order
        flat_geo = []
        for source_entity, matches in geo_vocab.items():
            for match in matches:
                flat_geo.append({
                    'source_entity': source_entity,
                    'label': match.get('label', ''),
                    'uri': match.get('uri', ''),
                    'source': match.get('source', 'FAST Geographic'),
                })

        approved = []
        for term_id, status in term_decisions.items():
            if not term_id.startswith('geo-'):
                continue
            if status != 'approved':
                continue
            try:
                idx = int(term_id.split('-')[1])
            except (IndexError, ValueError):
                continue
            if idx < len(flat_geo):
                term = flat_geo[idx]
                approved.append({
                    'label': term['label'],
                    'uri': term['uri'],
                    'source': term['source'],
                    'derived_from_entity': term['source_entity'],
                })

        return approved

    def _get_approved_alt_terms(self, analysis, term_decisions):
        """Return alternative subject headings approved via 'alt-N' decisions.

        The HTML page iterates vocabulary_search_results grouped by source in
        sources_order = [LCSH, FAST, Getty AAT, Getty TGN], assigning alt-0,
        alt-1, ... across the full combined list. This replicates that ordering.
        """
        vocab_results = analysis.get('vocabulary_search_results', {})
        if not vocab_results:
            return []

        final_terms = analysis.get('final_selected_terms', [])
        selected_pairs = set(
            (t.get('label', '').lower(), t.get('source', '')) for t in final_terms
        )

        sources_order = ['LCSH', 'FAST', 'Getty AAT', 'Getty TGN']
        sources_map = {s: [] for s in sources_order}

        for orig_topic, matches in vocab_results.items():
            for match in matches:
                source = match.get('source', 'Unknown')
                label = match.get('label', '')
                if (label.lower(), source) in selected_pairs:
                    continue
                if source in sources_map:
                    sources_map[source].append({
                        'orig_topic': orig_topic,
                        'label': label,
                        'uri': match.get('uri', ''),
                        'source': source,
                    })

        # Build flat list in same order as HTML render
        flat_alt = []
        for source_name in sources_order:
            flat_alt.extend(sources_map[source_name])

        approved = []
        for term_id, status in term_decisions.items():
            if not term_id.startswith('alt-'):
                continue
            if status != 'approved':
                continue
            try:
                idx = int(term_id.split('-')[1])
            except (IndexError, ValueError):
                continue
            if idx < len(flat_alt):
                term = flat_alt[idx]
                approved.append({
                    'label': term['label'],
                    'uri': term['uri'],
                    'source': term['source'],
                    'derived_from_topic': term.get('orig_topic', ''),
                })

        return approved

    # ------------------------------------------------------------------
    # Regenerate derived metadata files
    # ------------------------------------------------------------------

    def regenerate_derived_files(self):
        """Regenerate all metadata files derived from the workflow JSON.

        Imports generation methods from step3, step4, and step5 and calls them
        against the already-saved (updated) workflow JSON. No API calls are made.
        """
        # --- Step 3: page metadata TXTs, issue content index TXTs,
        #             vocabulary mapping report TXT, workflow XLSX (vocab column) ---
        try:
            from southern_architect_step3 import SouthernArchitectVocabularyProcessor
            processor = SouthernArchitectVocabularyProcessor(self.folder_path)
            processor.detect_workflow_type()
            processor.load_json_data()
            processor.create_page_metadata_files()
            processor.create_issue_content_index()
            processor.create_vocabulary_mapping_report({})
            if processor.excel_path and os.path.exists(processor.excel_path):
                data_items = (
                    processor.json_data[:-1]
                    if processor.json_data and 'api_stats' in processor.json_data[-1]
                    else processor.json_data
                )
                selection_results = {i: {} for i in range(len(data_items))}
                processor.update_excel_file(selection_results)
        except Exception as exc:
            print("  Warning: step3 file regeneration failed: " + str(exc))

        # --- Step 4: issue metadata TXTs ---
        # AI-generated issue descriptions are preserved from the stored issue_syntheses;
        # subject headings and geographic terms are re-extracted from the updated workflow.
        if self.issue_syntheses_entry is None:
            print("  No issue syntheses stored -- skipping issue metadata files.")
        else:
            try:
                from southern_architect_step4 import SouthernArchitectIssueSynthesizer
                synthesizer = SouthernArchitectIssueSynthesizer(self.folder_path)
                synthesizer.detect_workflow_type()
                synthesizer.load_json_data()
                issue_metadata_folder = os.path.join(
                    self.folder_path, "metadata", "issue_metadata"
                )
                os.makedirs(issue_metadata_folder, exist_ok=True)
                stored_issues = self.issue_syntheses_entry.get(
                    'issue_syntheses', {}
                ).get('issues', [])
                for stored in stored_issues:
                    issue_name = stored.get('issue_name', '')
                    if not issue_name:
                        continue
                    issue_chosen_terms, issue_geo_terms = (
                        synthesizer.extract_issue_specific_data(issue_name)
                    )
                    synthesis_result = {
                        'issue_description': stored.get('issue_description', ''),
                        'selected_subject_headings': issue_chosen_terms,
                    }
                    synthesizer.create_issue_metadata_file(
                        synthesis_result, issue_name, issue_metadata_folder
                    )
                    if issue_geo_terms:
                        synthesizer.append_geographic_terms_to_file_in_folder(
                            issue_geo_terms, issue_name, issue_metadata_folder
                        )
            except Exception as exc:
                print("  Warning: step4 file regeneration failed: " + str(exc))

        # --- Step 5: entity authority JSON and report TXT ---
        try:
            from southern_architect_step5 import EntityAuthority
            authority = EntityAuthority(self.folder_path)
            authority.detect_workflow_type()
            authority.load_json_data()
            entity_records = authority.extract_all_entities()
            authority.create_authority_file(entity_records)
            authority.create_human_readable_report(entity_records)
        except Exception as exc:
            print("  Warning: step5 file regeneration failed: " + str(exc))

    # ------------------------------------------------------------------
    # Final metadata generation
    # ------------------------------------------------------------------

    def generate_final_metadata(self):
        """Write final_metadata.json with only approved/reviewer-approved data."""
        final_records = []

        for record in self.workflow_data:
            analysis = record.get('analysis', {})
            term_decisions = analysis.get('reviewer_term_decisions', {})

            # Subject headings: keep selected terms that were not rejected
            final_selected = analysis.get('final_selected_terms', [])
            approved_subject_headings = []
            for idx, term in enumerate(final_selected):
                term_id = 'selected-' + str(idx)
                decision = term_decisions.get(term_id)
                if decision == 'rejected':
                    continue
                approved_subject_headings.append({
                    'label': term.get('label', ''),
                    'uri': term.get('uri', ''),
                    'source': term.get('source', ''),
                })

            # Alternative subject headings added by reviewer
            alt_terms = self._get_approved_alt_terms(analysis, term_decisions)
            for term in alt_terms:
                approved_subject_headings.append({
                    'label': term['label'],
                    'uri': term['uri'],
                    'source': term['source'],
                })

            # Custom terms added by reviewer
            custom_terms = analysis.get('reviewer_custom_terms', [])
            for term in custom_terms:
                approved_subject_headings.append({
                    'label': term.get('label', ''),
                    'uri': term.get('uri', ''),
                    'source': term.get('source', 'Manual'),
                })

            # Geographic headings: approved by reviewer from geo vocab
            approved_geo_headings = self._get_approved_geo_terms(analysis, term_decisions)

            # Chronological terms: keep those not rejected
            chrono_vocab = analysis.get('chronological_vocabulary_terms', [])
            approved_chrono = []
            for idx, term in enumerate(chrono_vocab):
                term_id = 'chrono-' + str(idx)
                decision = term_decisions.get(term_id)
                if decision == 'rejected':
                    continue
                approved_chrono.append({
                    'label': term.get('label', ''),
                    'uri': term.get('uri', ''),
                    'source': term.get('source', ''),
                })

            clean_record = {
                'folder': record.get('folder', ''),
                'page_number': record.get('page_number', 0),
                'image_path': record.get('image_path', ''),
                'metadata': {
                    'text_transcription': analysis.get('text_transcription', ''),
                    'visual_description': analysis.get('visual_description', ''),
                    'toc_entry': analysis.get('toc_entry', ''),
                    'named_entities': analysis.get('named_entities', []),
                    'geographic_entities': analysis.get('geographic_entities', []),
                    'topics': analysis.get('topics', []),
                    'content_warning': analysis.get('content_warning', ''),
                    'subject_headings': approved_subject_headings,
                    'geographic_headings': approved_geo_headings,
                    'chronological_terms': approved_chrono,
                    'custom_terms': custom_terms,
                },
                'review_info': {
                    'reviewed': analysis.get('reviewer_reviewed', False),
                    'reviewer_name': analysis.get('reviewer_name', ''),
                    'review_date': analysis.get('reviewer_date', ''),
                    'reviewer_notes': analysis.get('reviewer_notes', ''),
                },
            }
            final_records.append(clean_record)

        output = {
            'generated_timestamp': datetime.now().isoformat(),
            'reviewer_name': self.stats['reviewer_name'],
            'workflow_type': self.workflow_type,
            'total_records': len(final_records),
            'records': final_records,
        }

        out_path = os.path.join(self.collection_metadata_dir, "final_metadata.json")
        try:
            with open(out_path, 'w', encoding='utf-8') as f:
                json.dump(output, f, indent=2, ensure_ascii=False)
            print("  Generated final_metadata.json (" + str(len(final_records)) + " records)")
            return out_path
        except Exception as exc:
            print("  Error generating final_metadata.json: " + str(exc))
            return None

    # ------------------------------------------------------------------
    # Edit statistics report
    # ------------------------------------------------------------------

    def generate_edit_statistics_report(self):
        """Write edit_statistics_report.json with summary metrics."""
        self.stats['integration_timestamp'] = datetime.now().isoformat()

        total = self.stats['total_records_in_export']

        # Text field summaries
        text_summary = {}
        for field_name, metrics in self.stats['text_field_metrics'].items():
            edits = metrics['edits_count']
            unedited = total - edits
            similarity_all = metrics['similarity_sum'] + unedited
            avg_similarity = (similarity_all / total * 100) if total > 0 else 100.0
            text_summary[field_name] = {
                'records_edited': edits,
                'records_total': total,
                'chars_total_changed': metrics['chars_total_changed'],
                'similarity_pct': round(avg_similarity, 2),
            }

        # List field summaries
        list_summary = {}
        for field_name, metrics in self.stats['list_field_metrics'].items():
            list_summary[field_name] = {
                'records_edited': metrics['edits_count'],
                'items_added': metrics['items_added'],
                'items_removed': metrics['items_removed'],
            }

        # Subject headings
        sh_total = self.stats['subject_headings_total']
        sh_approved = self.stats['subject_headings_approved']
        sh_rejected = self.stats['subject_headings_rejected']
        sh_pct = round((sh_approved / sh_total * 100) if sh_total > 0 else 100.0, 2)

        # Chrono terms
        ct_total = self.stats['chrono_terms_total']
        ct_approved = self.stats['chrono_terms_approved']
        ct_rejected = self.stats['chrono_terms_rejected']
        ct_pct = round((ct_approved / ct_total * 100) if ct_total > 0 else 100.0, 2)

        report = {
            'batch_info': {
                'reviewer': self.stats['reviewer_name'],
                'export_timestamp': self.stats['export_timestamp'],
                'integration_timestamp': self.stats['integration_timestamp'],
                'workflow_type': self.workflow_type,
                'total_records_in_batch': self.stats['total_records_in_batch'],
                'records_reviewed': self.stats['total_records_in_export'],
                'records_with_edits': self.stats['records_with_edits'],
                'records_reviewed_only': self.stats['records_reviewed_only'],
            },
            'field_edits': {
                'total_field_edits': self.stats['total_field_edits'],
                'edits_by_field': self.stats['edits_by_field'],
                'text_fields': text_summary,
                'list_fields': list_summary,
            },
            'vocabulary_decisions': {
                'subject_headings': {
                    'total_ai_selected': sh_total,
                    'approved': sh_approved,
                    'rejected': sh_rejected,
                    'approval_rate_pct': sh_pct,
                    'alternative_headings_added': self.stats['alt_headings_added'],
                    'custom_terms_added': self.stats['custom_terms_added'],
                },
                'geographic_headings': {
                    'headings_added_by_reviewer': self.stats['geo_headings_added'],
                },
                'chronological_terms': {
                    'total': ct_total,
                    'approved': ct_approved,
                    'rejected': ct_rejected,
                    'approval_rate_pct': ct_pct,
                },
            },
        }

        os.makedirs(self.logs_folder, exist_ok=True)
        out_path = os.path.join(self.logs_folder, "edit_statistics_report.json")
        try:
            with open(out_path, 'w', encoding='utf-8') as f:
                json.dump(report, f, indent=2, ensure_ascii=False)
            print("  Generated edit_statistics_report.json")
            return out_path
        except Exception as exc:
            print("  Error generating edit_statistics_report.json: " + str(exc))
            return None

    # ------------------------------------------------------------------
    # Edit changelog
    # ------------------------------------------------------------------

    def generate_edit_changelog(self):
        """Write edit_changelog.json with the full before/after edit history."""
        changelog = {
            'generated': self.stats.get('integration_timestamp', datetime.now().isoformat()),
            'reviewer': self.stats['reviewer_name'],
            'total_edits': len(self.edit_history),
            'edits': self.edit_history,
        }
        os.makedirs(self.logs_folder, exist_ok=True)
        out_path = os.path.join(self.logs_folder, "edit_changelog.json")
        try:
            with open(out_path, 'w', encoding='utf-8') as f:
                json.dump(changelog, f, indent=2, ensure_ascii=False)
            print("  Generated edit_changelog.json")
            return out_path
        except Exception as exc:
            print("  Error generating edit_changelog.json: " + str(exc))
            return None

    # ------------------------------------------------------------------
    # Excel deliverable
    # ------------------------------------------------------------------

    def create_final_deliverable_xlsx(self):
        """Create final_deliverable.xlsx with Final Metadata and Edit Statistics sheets."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
        except ImportError:
            print("  Warning: openpyxl not installed. Skipping Excel deliverable.")
            print("  Install with: pip install openpyxl")
            return None

        try:
            wb = Workbook()

            header_fill = PatternFill(start_color="FF2C3E50", end_color="FF2C3E50", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            stats_fill = PatternFill(start_color="FFE8F4F8", end_color="FFE8F4F8", fill_type="solid")
            section_font = Font(bold=True, size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'),
            )
            wrap_top = Alignment(wrap_text=True, vertical='top')
            center_mid = Alignment(horizontal='center', vertical='center')

            # ----------------------------------------
            # Sheet 1: Final Metadata
            # ----------------------------------------
            ws1 = wb.active
            ws1.title = "Final Metadata"

            metadata_headers = [
                "Folder", "Page Number", "TOC Entry", "Text Transcription",
                "Visual Description", "Topics", "Subject Headings",
                "Geographic Entities", "Geographic Headings",
                "Chronological Terms", "Named Entities", "Content Warning",
                "Reviewed", "Reviewer", "Review Date", "Notes",
            ]

            for col_idx, header in enumerate(metadata_headers, 1):
                cell = ws1.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_mid

            final_metadata_path = os.path.join(self.collection_metadata_dir, "final_metadata.json")
            if os.path.exists(final_metadata_path):
                with open(final_metadata_path, 'r', encoding='utf-8') as f:
                    final_data = json.load(f)

                def fmt_list(lst):
                    if isinstance(lst, list):
                        return '; '.join(str(item) for item in lst if item)
                    return str(lst) if lst else ''

                def fmt_term_list(lst):
                    if not isinstance(lst, list):
                        return ''
                    parts = []
                    for item in lst:
                        if isinstance(item, dict):
                            label = item.get('label', '')
                            source = item.get('source', '')
                            uri = item.get('uri', '')
                            if uri:
                                parts.append(label + ' [' + source + '] (' + uri + ')')
                            elif source:
                                parts.append(label + ' [' + source + ']')
                            else:
                                parts.append(label)
                        else:
                            parts.append(str(item))
                    return '; '.join(parts)

                for row_idx, record in enumerate(final_data.get('records', []), 2):
                    metadata = record.get('metadata', {})
                    review_info = record.get('review_info', {})

                    row_data = [
                        record.get('folder', ''),
                        record.get('page_number', ''),
                        metadata.get('toc_entry', ''),
                        metadata.get('text_transcription', ''),
                        metadata.get('visual_description', ''),
                        fmt_list(metadata.get('topics', [])),
                        fmt_term_list(metadata.get('subject_headings', [])),
                        fmt_list(metadata.get('geographic_entities', [])),
                        fmt_term_list(metadata.get('geographic_headings', [])),
                        fmt_term_list(metadata.get('chronological_terms', [])),
                        fmt_list(metadata.get('named_entities', [])),
                        metadata.get('content_warning', ''),
                        'Yes' if review_info.get('reviewed') else 'No',
                        review_info.get('reviewer_name', ''),
                        review_info.get('review_date', ''),
                        review_info.get('reviewer_notes', ''),
                    ]

                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = border
                        cell.alignment = wrap_top

            col_widths = [14, 12, 45, 60, 60, 40, 60, 40, 60, 50, 40, 20, 10, 20, 22, 40]
            for col_idx, width in enumerate(col_widths, 1):
                ws1.column_dimensions[
                    ws1.cell(row=1, column=col_idx).column_letter
                ].width = width

            ws1.freeze_panes = 'A2'

            # ----------------------------------------
            # Sheet 2: Edit Statistics
            # ----------------------------------------
            ws2 = wb.create_sheet("Edit Statistics")

            ws2['A1'] = "REVIEWER EDIT INTEGRATION SUMMARY"
            ws2['A1'].font = Font(size=14, bold=True)
            ws2.merge_cells('A1:C1')

            row = 3

            def write_stat(label, value, fill=True):
                nonlocal row
                c1 = ws2.cell(row=row, column=1, value=label)
                c1.font = Font(bold=True)
                c2 = ws2.cell(row=row, column=2, value=value)
                if fill:
                    c1.fill = stats_fill
                    c2.fill = stats_fill
                row += 1

            def write_section(title):
                nonlocal row
                row += 1
                ws2.cell(row=row, column=1, value=title).font = section_font
                row += 1

            write_stat("Reviewer:", self.stats['reviewer_name'])
            write_stat("Export Timestamp:", self.stats['export_timestamp'])
            write_stat("Integration Timestamp:", self.stats.get('integration_timestamp', ''))
            write_stat("Workflow Type:", self.workflow_type or '')
            write_stat("", "")
            write_stat("Total Records in Batch:", self.stats['total_records_in_batch'])
            write_stat("Records Reviewed:", self.stats['total_records_in_export'])
            write_stat("Records with Edits:", self.stats['records_with_edits'])
            write_stat("Records Reviewed (no edits):", self.stats['records_reviewed_only'])
            write_stat("Total Field Edits:", self.stats['total_field_edits'])

            write_section("EDITS BY FIELD")
            for field_name, count in sorted(self.stats['edits_by_field'].items()):
                ws2.cell(row=row, column=1, value="  " + field_name + ":")
                ws2.cell(row=row, column=2, value=count)
                row += 1

            write_section("TEXT FIELD CHARACTER CHANGES")
            for field_name, metrics in sorted(self.stats['text_field_metrics'].items()):
                edits = metrics['edits_count']
                total_exp = self.stats['total_records_in_export']
                unedited = total_exp - edits
                sim_all = (metrics['similarity_sum'] + unedited) / total_exp * 100 if total_exp > 0 else 100.0
                ws2.cell(row=row, column=1, value="  " + field_name + ":").fill = stats_fill
                ws2.cell(row=row, column=2, value=(
                    str(edits) + " edits, " +
                    str(metrics['chars_total_changed']) + " chars changed, " +
                    str(round(sim_all, 1)) + "% similarity"
                )).fill = stats_fill
                row += 1

            write_section("SUBJECT HEADINGS (AI-selected)")
            sh_total = self.stats['subject_headings_total']
            sh_approved = self.stats['subject_headings_approved']
            sh_pct = round((sh_approved / sh_total * 100) if sh_total > 0 else 100.0, 1)
            write_stat("  Total AI-selected:", sh_total)
            write_stat("  Approved:", str(sh_approved) + " (" + str(sh_pct) + "%)")
            write_stat("  Rejected:", self.stats['subject_headings_rejected'])
            write_stat("  Alternative headings added:", self.stats['alt_headings_added'])
            write_stat("  Custom terms added:", self.stats['custom_terms_added'])

            write_section("GEOGRAPHIC HEADINGS")
            write_stat("  Geographic headings added by reviewer:", self.stats['geo_headings_added'])

            write_section("CHRONOLOGICAL TERMS")
            ct_total = self.stats['chrono_terms_total']
            ct_approved = self.stats['chrono_terms_approved']
            ct_pct = round((ct_approved / ct_total * 100) if ct_total > 0 else 100.0, 1)
            write_stat("  Total:", ct_total)
            write_stat("  Approved:", str(ct_approved) + " (" + str(ct_pct) + "%)")
            write_stat("  Rejected:", self.stats['chrono_terms_rejected'])

            ws2.column_dimensions['A'].width = 38
            ws2.column_dimensions['B'].width = 55

            # Save
            out_path = os.path.join(self.collection_metadata_dir, "final_deliverable.xlsx")
            wb.save(out_path)
            print("  Created final_deliverable.xlsx")
            return out_path

        except Exception as exc:
            print("  Error creating Excel deliverable: " + str(exc))
            import traceback
            traceback.print_exc()
            return None

    # ------------------------------------------------------------------
    # Summary print
    # ------------------------------------------------------------------

    def print_summary(self):
        """Print a human-readable integration summary to stdout."""
        print("")
        print("=" * 60)
        print("INTEGRATION SUMMARY")
        print("=" * 60)
        print("Reviewer: " + self.stats['reviewer_name'])
        print("Records processed: " + str(self.stats['total_records_in_export']))
        print("Records with edits: " + str(self.stats['records_with_edits']))
        print("Records reviewed (no edits): " + str(self.stats['records_reviewed_only']))
        print("")
        print("Field edits: " + str(self.stats['total_field_edits']))
        if self.stats['edits_by_field']:
            for field_name, count in sorted(self.stats['edits_by_field'].items()):
                print("  " + field_name + ": " + str(count))

        sh_total = self.stats['subject_headings_total']
        sh_approved = self.stats['subject_headings_approved']
        sh_pct = round((sh_approved / sh_total * 100) if sh_total > 0 else 100.0, 1)
        print("")
        print("Subject headings:")
        print("  Total AI-selected: " + str(sh_total))
        print("  Approved: " + str(sh_approved) + " (" + str(sh_pct) + "%)")
        print("  Rejected: " + str(self.stats['subject_headings_rejected']))
        print("  Alternative headings added: " + str(self.stats['alt_headings_added']))
        print("  Custom terms added: " + str(self.stats['custom_terms_added']))

        print("")
        print("Geographic headings added by reviewer: " + str(self.stats['geo_headings_added']))

        ct_total = self.stats['chrono_terms_total']
        ct_approved = self.stats['chrono_terms_approved']
        ct_pct = round((ct_approved / ct_total * 100) if ct_total > 0 else 100.0, 1)
        print("")
        print("Chronological terms:")
        print("  Total: " + str(ct_total))
        print("  Approved: " + str(ct_approved) + " (" + str(ct_pct) + "%)")
        print("  Rejected: " + str(self.stats['chrono_terms_rejected']))

    # ------------------------------------------------------------------
    # Main run method
    # ------------------------------------------------------------------

    def run(self, decisions_path=None):
        """Execute the full integration pipeline."""
        print("")
        print("=" * 60)
        print("SA Integrate Reviewer Edits")
        print("=" * 60)
        print("Folder: " + self.folder_name)

        # Step 1: Find decisions file
        print("")
        print("1. Locating reviewer decisions export...")
        if decisions_path:
            if not os.path.exists(decisions_path):
                print("Error: Decisions file not found: " + decisions_path)
                return False
            print("  Using: " + os.path.basename(decisions_path))
        else:
            decisions_path = self.find_latest_export()
            if not decisions_path:
                return False
            print("  Found latest: " + os.path.basename(decisions_path))

        # Step 2: Load decisions
        print("")
        print("2. Loading reviewer decisions...")
        if not self.load_decisions(decisions_path):
            return False

        # Step 3: Load workflow JSON
        print("")
        print("3. Loading workflow data...")
        if not self.load_workflow_json():
            return False

        # Step 4: Backup originals
        print("")
        print("4. Backing up original files...")
        self.backup_original_files()

        # Step 5: Apply edits
        print("")
        print("5. Applying edits...")
        if not self.apply_all_edits():
            return False
        print("  Applied edits to " + str(self.stats['records_with_edits']) + " records")

        # Step 6: Save updated workflow JSON
        print("")
        print("6. Saving updated workflow JSON...")
        if not self.save_workflow_json():
            return False

        # Step 7: Regenerate all derived metadata files
        print("")
        print("7. Regenerating derived metadata files...")
        self.regenerate_derived_files()

        # Step 8: Generate final metadata
        print("")
        print("8. Generating final_metadata.json...")
        self.generate_final_metadata()

        # Step 9: Reports
        print("")
        print("9. Generating reports...")
        self.generate_edit_statistics_report()
        self.generate_edit_changelog()

        # Step 10: Excel deliverable
        print("")
        print("10. Creating Excel deliverable...")
        self.create_final_deliverable_xlsx()

        self.print_summary()

        print("")
        print("=" * 60)
        print("INTEGRATION COMPLETE")
        print("=" * 60)
        print("Collection metadata: " + self.collection_metadata_dir)
        print("  " + os.path.basename(self.workflow_json_path) + " (updated)")
        print("  southern_architect_entity_authority.json (regenerated)")
        print("  entity_authority_report.txt (regenerated)")
        print("  vocabulary_mapping_report.txt (regenerated)")
        print("  final_metadata.json")
        print("  final_deliverable.xlsx")
        print("")
        page_metadata_dir = os.path.join(self.folder_path, "metadata", "page_metadata")
        print("Page metadata: " + page_metadata_dir)
        print("  *_metadata.txt (all regenerated)")
        print("")
        issue_metadata_dir = os.path.join(self.folder_path, "metadata", "issue_metadata")
        print("Issue metadata: " + issue_metadata_dir)
        print("  *_Issue_Content_Index.txt (all regenerated)")
        print("  *_Issue_Metadata.txt (all regenerated)")
        print("")
        print("Log files: " + self.logs_folder)
        print("  edit_statistics_report.json")
        print("  edit_changelog.json")
        print("")
        print("Originals backed up to: " + self.original_outputs_folder)
        print("=" * 60)

        return True


# ------------------------------------------------------------------
# Interactive folder/export selection helpers
# ------------------------------------------------------------------

def list_available_folders(base_dir):
    """Return a list of (name, path) tuples for available output folders, newest first."""
    if not os.path.exists(base_dir):
        return []
    folders = []
    for item in os.listdir(base_dir):
        item_path = os.path.join(base_dir, item)
        if os.path.isdir(item_path) and item.startswith("SABN_"):
            folders.append((item, item_path))
    folders.sort(key=lambda x: os.path.getmtime(x[1]), reverse=True)
    return folders


def list_available_exports(exports_folder):
    """Return a list of (name, path) tuples for available JSON exports, newest first."""
    if not os.path.exists(exports_folder):
        return []
    exports = []
    for item in os.listdir(exports_folder):
        if item.endswith('.json'):
            item_path = os.path.join(exports_folder, item)
            exports.append((item, item_path))
    exports.sort(key=lambda x: os.path.getmtime(x[1]), reverse=True)
    return exports


def prompt_for_folder(base_dir):
    """Interactively prompt user to select an output folder."""
    folders = list_available_folders(base_dir)
    if not folders:
        print("No SA output folders found in: " + base_dir)
        return None

    print("")
    print("Available output folders:")
    print("-" * 60)
    for idx, (name, path) in enumerate(folders, 1):
        mtime = datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d %H:%M")
        marker = " (newest)" if idx == 1 else ""
        print("  " + str(idx) + ". " + name + marker)
        print("       Modified: " + mtime)
    print("-" * 60)
    print("Enter a number to select, or press Enter for the newest folder.")

    while True:
        choice = input("\nSelect folder: ").strip()
        if choice == "":
            return folders[0][1]
        if choice.isdigit():
            idx = int(choice) - 1
            if 0 <= idx < len(folders):
                return folders[idx][1]
            print("Invalid choice. Enter 1-" + str(len(folders)) + ".")
            continue
        if os.path.isdir(choice):
            return choice
        print("Invalid input. Enter a number or a valid folder path.")


def prompt_for_decisions(folder_path):
    """Interactively prompt user to select an export JSON file."""
    exports_folder = os.path.join(folder_path, "review", "exports")
    exports = list_available_exports(exports_folder)

    if not exports:
        print("")
        print("No JSON export files found in: " + exports_folder)
        print("Export decisions from the HTML review interface and place the file there.")
        print("Or enter a full path to the decisions JSON file.")

        while True:
            path = input("\nPath to decisions JSON (or 'q' to quit): ").strip()
            if path.lower() == 'q':
                return None
            if os.path.isfile(path) and path.endswith('.json'):
                return path
            print("Invalid path.")

    print("")
    print("Available decisions exports:")
    print("-" * 60)
    for idx, (name, path) in enumerate(exports, 1):
        mtime = datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d %H:%M")
        marker = " (newest)" if idx == 1 else ""
        print("  " + str(idx) + ". " + name + marker)
        print("       Modified: " + mtime)
    print("-" * 60)
    print("Enter a number to select, or press Enter for the newest export.")

    while True:
        choice = input("\nSelect export: ").strip()
        if choice == "":
            return exports[0][1]
        if choice.isdigit():
            idx = int(choice) - 1
            if 0 <= idx < len(exports):
                return exports[idx][1]
            print("Invalid choice. Enter 1-" + str(len(exports)) + ".")
            continue
        if os.path.isfile(choice) and choice.endswith('.json'):
            return choice
        print("Invalid input. Enter a number or a valid JSON file path.")


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description='Apply reviewer edits from sa_html_review.py back into the SA workflow JSON.'
    )
    parser.add_argument(
        '--decisions', '-d',
        help='Path to a specific reviewer decisions JSON file.'
    )
    parser.add_argument(
        '--folder', '-f',
        help='Path to the output folder (defaults to interactive prompt).'
    )
    parser.add_argument(
        '--yes', '-y',
        action='store_true',
        help='Skip the confirmation prompt.'
    )
    args = parser.parse_args()

    print("SA Integrate Reviewer Edits")
    print("=" * 60)

    base_output_dir = os.path.join(script_dir, FOLDER_CONFIG['output_dir'])

    # Resolve output folder
    if args.folder:
        folder_path = args.folder
        if not os.path.isdir(folder_path):
            print("Error: Folder not found: " + folder_path)
            return 1
        print("Using folder: " + os.path.basename(folder_path))
    else:
        folder_path = prompt_for_folder(base_output_dir)
        if not folder_path:
            return 1
        print("Selected folder: " + os.path.basename(folder_path))

    # Resolve decisions file
    if args.decisions:
        decisions_path = args.decisions
        if not os.path.exists(decisions_path):
            print("Error: Decisions file not found: " + decisions_path)
            return 1
        print("Using decisions: " + os.path.basename(decisions_path))
    else:
        decisions_path = prompt_for_decisions(folder_path)
        if not decisions_path:
            print("Operation cancelled.")
            return 0
        print("Selected decisions: " + os.path.basename(decisions_path))

    # Confirmation
    if not args.yes:
        print("")
        print("-" * 60)
        print("Summary:")
        print("  Output folder: " + os.path.basename(folder_path))
        print("  Decisions file: " + os.path.basename(decisions_path))
        print("-" * 60)
        response = input(
            "\nThis will modify the workflow JSON. Original files will be backed up. Continue? (yes/no): "
        ).strip().lower()
        if response not in ('yes', 'y'):
            print("Operation cancelled.")
            return 0

    integrator = SAEditsIntegrator(folder_path)
    success = integrator.run(decisions_path=decisions_path)
    return 0 if success else 1


if __name__ == "__main__":
    sys.exit(main())
