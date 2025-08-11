# Controlled Vocabulary API querying - LCSH, FAST, Getty AAT/TGN Subject Headings

import os
import json
import logging
import time
import requests
import xml.etree.ElementTree as ET
from datetime import datetime
from typing import List, Dict, Any, Tuple
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from collections import defaultdict
import re
from shared_utilities import find_newest_folder

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def create_vocab_api_usage_log(logs_folder_path: str, script_name: str, total_topics: int, 
                              api_stats: Dict[str, Any]) -> bool:
    """Create vocabulary API usage log."""
    try:
        log_filename = f"{script_name}_vocab_api_usage_log.txt"
        log_path = os.path.join(logs_folder_path, log_filename)
        
        with open(log_path, 'w', encoding='utf-8') as f:
            f.write("SOUTHERN ARCHITECT VOCABULARY API USAGE LOG\n")
            f.write("=" * 50 + "\n\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Script: {script_name}\n\n")
            
            # Summary statistics
            f.write("SUMMARY STATISTICS:\n")
            f.write("-" * 20 + "\n")
            f.write(f"Total topics processed: {total_topics}\n")
            f.write(f"Total API requests made: {api_stats.get('total_requests', 0)}\n")
            f.write(f"Total processing time: {api_stats.get('total_time', 0):.2f} seconds\n")
            f.write(f"Average time per topic: {api_stats.get('avg_time_per_topic', 0):.2f} seconds\n\n")
            
            # API-specific statistics
            for api_name, stats in api_stats.get('api_breakdown', {}).items():
                f.write(f"{api_name.upper()} API STATISTICS:\n")
                f.write("-" * (len(api_name) + 16) + "\n")
                f.write(f"Requests made: {stats.get('requests', 0)}\n")
                f.write(f"Successful responses: {stats.get('successful', 0)}\n")
                f.write(f"Failed responses: {stats.get('failed', 0)}\n")
                f.write(f"Success rate: {stats.get('success_rate', 0):.1f}%\n")
                f.write(f"Total processing time: {stats.get('total_time', 0):.2f} seconds\n")
                f.write(f"Average time per request: {stats.get('avg_time', 0):.2f} seconds\n")
                f.write(f"Terms found: {stats.get('terms_found', 0)}\n")
                f.write(f"Cache hits: {stats.get('cache_hits', 0)}\n\n")
            
            # Topic-level results summary
            f.write("TOPIC PROCESSING RESULTS:\n")
            f.write("-" * 25 + "\n")
            topic_results = api_stats.get('topic_results', {})
            for topic, result in topic_results.items():
                f.write(f"Topic: {topic}\n")
                f.write(f"  Total terms found: {result.get('total_terms', 0)}\n")
                for api_name, count in result.get('terms_by_api', {}).items():
                    f.write(f"  {api_name}: {count} terms\n")
                f.write(f"  Processing time: {result.get('processing_time', 0):.2f}s\n\n")
        
        print(f"Vocabulary API usage log created: {log_path}")
        return True
        
    except Exception as e:
        logging.error(f"Error creating vocabulary API usage log: {e}")
        return False

def log_individual_vocab_response(logs_folder_path: str, script_name: str, topic: str, 
                                 api_name: str, query: str, response_data: List[Dict], 
                                 processing_time: float, error: str = None) -> bool:
    """Log individual vocabulary API response."""
    try:
        log_filename = f"{script_name}_vocab_full_responses_log.txt"
        log_path = os.path.join(logs_folder_path, log_filename)
        
        # Create file with header if it doesn't exist
        if not os.path.exists(log_path):
            with open(log_path, 'w', encoding='utf-8') as f:
                f.write("SOUTHERN ARCHITECT VOCABULARY API DETAILED RESPONSES LOG\n")
                f.write("=" * 60 + "\n\n")
        
        with open(log_path, 'a', encoding='utf-8') as f:
            f.write(f"TIMESTAMP: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"TOPIC: {topic}\n")
            f.write(f"API: {api_name}\n")
            f.write(f"QUERY: {query}\n")
            f.write(f"PROCESSING TIME: {processing_time:.3f}s\n")
            
            if error:
                f.write(f"ERROR: {error}\n")
                f.write("RESPONSE: Failed\n")
            else:
                f.write(f"TERMS FOUND: {len(response_data)}\n")
                f.write("RESPONSE:\n")
                for i, term in enumerate(response_data, 1):
                    label = term.get('label', 'N/A')
                    uri = term.get('uri', 'N/A')
                    source = term.get('source', 'N/A')
                    f.write(f"  {i}. {label}\n")
                    f.write(f"     URI: {uri}\n")
                    f.write(f"     Source: {source}\n")
            
            f.write("-" * 60 + "\n\n")
        
        return True
        
    except Exception as e:
        logging.error(f"Error logging individual vocabulary response: {e}")
        return False

class APIStatsTracker:
    """Track API statistics across all vocabulary services."""
    
    def __init__(self):
        self.start_time = time.time()
        self.api_breakdown = {
            'LCSH': {'requests': 0, 'successful': 0, 'failed': 0, 'total_time': 0, 'terms_found': 0, 'cache_hits': 0},
            'FAST': {'requests': 0, 'successful': 0, 'failed': 0, 'total_time': 0, 'terms_found': 0, 'cache_hits': 0},
            'FAST Geographic': {'requests': 0, 'successful': 0, 'failed': 0, 'total_time': 0, 'terms_found': 0, 'cache_hits': 0},
            'Getty AAT': {'requests': 0, 'successful': 0, 'failed': 0, 'total_time': 0, 'terms_found': 0, 'cache_hits': 0},
            'Getty TGN': {'requests': 0, 'successful': 0, 'failed': 0, 'total_time': 0, 'terms_found': 0, 'cache_hits': 0}
        }
        self.topic_results = {}
        self.geographic_results = {}  # Track geographic results separately
        self.total_requests = 0
    
    def record_api_call(self, api_name: str, topic: str, query: str, success: bool, 
                       processing_time: float, terms_found: int, from_cache: bool = False):
        """Record an API call."""
        if api_name in self.api_breakdown:
            stats = self.api_breakdown[api_name]
            stats['requests'] += 1
            stats['total_time'] += processing_time
            
            if from_cache:
                stats['cache_hits'] += 1
            
            if success:
                stats['successful'] += 1
                stats['terms_found'] += terms_found
            else:
                stats['failed'] += 1
        
        # Track topic-level results
        if topic not in self.topic_results:
            self.topic_results[topic] = {
                'total_terms': 0,
                'terms_by_api': defaultdict(int),
                'processing_time': 0
            }
        
        self.topic_results[topic]['terms_by_api'][api_name] += terms_found
        self.topic_results[topic]['total_terms'] += terms_found
        self.topic_results[topic]['processing_time'] += processing_time
        
        self.total_requests += 1
    
    def get_summary_stats(self, total_topics: int) -> Dict[str, Any]:
        """Get summary statistics."""
        total_time = time.time() - self.start_time
        
        # Calculate success rates for each API
        for api_name in self.api_breakdown:
            stats = self.api_breakdown[api_name]
            total_api_requests = stats['requests']
            if total_api_requests > 0:
                stats['success_rate'] = (stats['successful'] / total_api_requests) * 100
                stats['avg_time'] = stats['total_time'] / total_api_requests
            else:
                stats['success_rate'] = 0
                stats['avg_time'] = 0
        
        return {
            'total_requests': self.total_requests,
            'total_time': total_time,
            'avg_time_per_topic': total_time / total_topics if total_topics > 0 else 0,
            'api_breakdown': self.api_breakdown,
            'topic_results': self.topic_results
        }

class FASTTermFinder:
    """FAST term finder"""
    
    def __init__(self, wskey=None, stats_tracker=None, logs_folder_path=None):
        # Updated API endpoints
        self.suggest_url = "https://fast.oclc.org/searchfast/fastsuggest"
        self.search_url = "http://fast.oclc.org/search"
        
        self.headers = {
            'User-Agent': 'Python-FAST-Term-Finder/1.0 (Educational/Research Use)'
        }
        self.max_results = 3  # Strict limit to 3 terms
        self.max_geo_results = 1  # Strict limit to 1 geographic term
        self.request_delay = 0.5
        self.cache = {}
        self.wskey = wskey  # OCLC WSKey for authentication (optional)
        self.stats_tracker = stats_tracker
        self.logs_folder_path = logs_folder_path
        
    def search(self, query: str, topic: str = None) -> List[Dict[str, str]]:
        """Search FAST using the updated API format with logging."""
        cache_key = f"fast_new_{query}"
        from_cache = cache_key in self.cache
        
        start_time = time.time()
        
        if from_cache:
            results = self.cache[cache_key]
            processing_time = time.time() - start_time
            
            # Log cache hit
            if self.stats_tracker:
                self.stats_tracker.record_api_call(
                    'FAST', topic or query, query, True, processing_time, len(results), from_cache=True
                )
            
            if self.logs_folder_path:
                log_individual_vocab_response(
                    self.logs_folder_path, "southern_architect_step2", topic or query, 
                    "FAST", f"{query} (CACHED)", results, processing_time
                )
            
            return results
        
        results = []
        error_msg = None
        
        try:
            # Try method 1: New suggest API without WSKey (usually works)
            results = self._try_suggest_api(query)
            
            # Try method 2: If no results and multi-word, try key terms
            if not results and ' ' in query:
                words = query.split()
                for word in words:
                    if len(word) > 3:  # Skip short words
                        word_results = self._try_suggest_api(word)
                        # Filter word results for relevance to original query
                        for result in word_results:
                            if self._is_relevant_to_query(result, query):
                                results.append(result)
                        if len(results) >= self.max_results:  # Stop when we have enough
                            break
            
            # Remove duplicates based on URI and STRICT LIMIT TO 3
            seen_uris = set()
            unique_results = []
            for result in results:
                if result['uri'] not in seen_uris and len(unique_results) < self.max_results:
                    unique_results.append(result)
                    seen_uris.add(result['uri'])
            
            results = unique_results
            success = True
            
        except Exception as e:
            error_msg = str(e)
            success = False
            results = []
        
        processing_time = time.time() - start_time
        
        # Cache results
        self.cache[cache_key] = results
        
        # Log the API call
        if self.stats_tracker:
            self.stats_tracker.record_api_call(
                'FAST', topic or query, query, success, processing_time, len(results)
            )
        
        if self.logs_folder_path:
            log_individual_vocab_response(
                self.logs_folder_path, "southern_architect_step2", topic or query, 
                "FAST", query, results, processing_time, error_msg
            )
        
        time.sleep(self.request_delay)
        return results
    
    def search_geographic(self, query: str, topic: str = None) -> List[Dict[str, str]]:
        """Search FAST for geographic entities with additional metadata."""
        cache_key = f"fast_geo_{query}"
        from_cache = cache_key in self.cache
        
        start_time = time.time()
        
        if from_cache:
            results = self.cache[cache_key]
            processing_time = time.time() - start_time
            
            if self.stats_tracker:
                self.stats_tracker.record_api_call(
                    'FAST Geographic', topic or query, query, True, processing_time, len(results), from_cache=True
                )
            
            if self.logs_folder_path:
                log_individual_vocab_response(
                    self.logs_folder_path, "southern_architect_step2", topic or query, 
                    "FAST Geographic", f"{query} (CACHED)", results, processing_time
                )
            
            return results
        
        results = []
        error_msg = None
        
        try:
            # Try method 1: Suggest API focusing on geographic facets
            params = {
                'query': query,
                'queryReturn': 'suggestall,idroot,auth,type,tag',
                'suggest': 'autoSubject',
                'rows': self.max_geo_results * 2,  # Use max_geo_results instead of max_results
                'sort': 'usage desc',
                'facet': 'type',
                'facet.field': 'type',
                'fq': 'type:Geographic'  # Filter for geographic entities
            }
            
            if self.wskey:
                params['wskey'] = self.wskey
            
            resp = requests.get(self.suggest_url, params=params, headers=self.headers, timeout=10)
            resp.raise_for_status()
            
            data = resp.json()
            results = self._parse_geographic_response(data)
            
            # If no results with geographic filter, try broader search
            if not results:
                params.pop('fq', None)  # Remove geographic filter
                resp = requests.get(self.suggest_url, params=params, headers=self.headers, timeout=10)
                resp.raise_for_status()
                data = resp.json()
                
                # Parse and filter for geographic relevance
                all_results = self._parse_geographic_response(data)
                results = [r for r in all_results if self._is_geographic_relevant(r, query)][:self.max_geo_results] 
            
            success = True
            
        except Exception as e:
            error_msg = str(e)
            success = False
            results = []
        
        processing_time = time.time() - start_time
        
        # Cache results
        self.cache[cache_key] = results
        
        # Log the API call
        if self.stats_tracker:
            self.stats_tracker.record_api_call(
                'FAST Geographic', topic or query, query, success, processing_time, len(results)
            )
        
        if self.logs_folder_path:
            log_individual_vocab_response(
                self.logs_folder_path, "southern_architect_step2", topic or query, 
                "FAST Geographic", query, results, processing_time, error_msg
            )
        
        time.sleep(self.request_delay)
        return results
    
    def _try_suggest_api(self, query: str) -> List[Dict[str, str]]:
        """Try the suggest API with the working parameters."""
        params = {
            'query': query,
            'queryReturn': 'suggestall,idroot,auth,type',
            'suggest': 'autoSubject',
            'rows': self.max_results * 2,  # Get a few more to filter
            'sort': 'usage desc'
        }
        
        # Add WSKey if available
        if self.wskey:
            params['wskey'] = self.wskey
        
        resp = requests.get(self.suggest_url, params=params, headers=self.headers, timeout=10)
        resp.raise_for_status()
        
        data = resp.json()
        return self._parse_suggest_response(data)
    
    def _parse_suggest_response(self, data: dict) -> List[Dict[str, str]]:
        """Parse the suggest API response format."""
        results = []
        
        if 'response' in data and 'docs' in data['response']:
            docs = data['response']['docs']
            
            for doc in docs:
                if len(results) >= self.max_results:
                    break
                
                # Extract fields from response
                suggest_all = doc.get('suggestall', '')
                id_root = doc.get('idroot', '')
                auth = doc.get('auth', '')
                doc_type = doc.get('type', '')
                
                # Handle fields that might be lists or strings
                if isinstance(id_root, list) and id_root:
                    id_root = id_root[0]
                elif not isinstance(id_root, str):
                    id_root = str(id_root) if id_root else ''
                
                if isinstance(suggest_all, list) and suggest_all:
                    suggest_all = suggest_all[0]
                elif not isinstance(suggest_all, str):
                    suggest_all = str(suggest_all) if suggest_all else ''
                
                if isinstance(auth, list) and auth:
                    auth = auth[0]
                elif not isinstance(auth, str):
                    auth = str(auth) if auth else ''
                
                if isinstance(doc_type, list) and doc_type:
                    doc_type = doc_type[0]
                elif not isinstance(doc_type, str):
                    doc_type = str(doc_type) if doc_type else ''
                
                # Only include if we have the essential fields
                if suggest_all and id_root:
                    results.append({
                        'label': auth if auth else suggest_all,
                        'uri': f"http://id.worldcat.org/fast/{id_root}",
                        'type': doc_type,
                        'source': 'FAST',
                        'idroot': id_root
                    })
        
        return results
    
    def _parse_geographic_response(self, data: dict) -> List[Dict[str, str]]:
        """Parse the geographic API response format with additional metadata."""
        results = []
        
        if 'response' in data and 'docs' in data['response']:
            docs = data['response']['docs']
            
            for doc in docs:
                if len(results) >= self.max_geo_results:  # CHANGED: Use max_geo_results
                    break
                
                # Extract fields from response
                suggest_all = doc.get('suggestall', '')
                id_root = doc.get('idroot', '')
                auth = doc.get('auth', '')
                doc_type = doc.get('type', '')
                tag = doc.get('tag', '')
                
                # Handle fields that might be lists or strings
                if isinstance(id_root, list) and id_root:
                    id_root = id_root[0]
                elif not isinstance(id_root, str):
                    id_root = str(id_root) if id_root else ''
                
                if isinstance(suggest_all, list) and suggest_all:
                    suggest_all = suggest_all[0]
                elif not isinstance(suggest_all, str):
                    suggest_all = str(suggest_all) if suggest_all else ''
                
                if isinstance(auth, list) and auth:
                    auth = auth[0]
                elif not isinstance(auth, str):
                    auth = str(auth) if auth else ''
                
                if isinstance(doc_type, list) and doc_type:
                    doc_type = doc_type[0]
                elif not isinstance(doc_type, str):
                    doc_type = str(doc_type) if doc_type else ''
                
                if isinstance(tag, list):
                    tag = ', '.join(str(t) for t in tag if t)
                elif not isinstance(tag, str):
                    tag = str(tag) if tag else ''
                
                # Only include if we have the essential fields
                if suggest_all and id_root:
                    result = {
                        'label': auth if auth else suggest_all,
                        'uri': f"http://id.worldcat.org/fast/{id_root}",
                        'type': doc_type,
                        'source': 'FAST Geographic',
                        'idroot': id_root
                    }
                    
                    # Add additional metadata for geographic entities
                    if tag:
                        result['tag'] = tag
                        result['sources_and_links'] = tag  # For vocabulary mapping report
                    
                    results.append(result)
        
        return results
    
    def _is_relevant_to_query(self, result: Dict[str, str], original_query: str) -> bool:
        """Check if a result from word search is relevant to original query."""
        query_words = set(w.lower() for w in original_query.split())
        result_words = set(w.lower() for w in result['label'].split())
        
        # Check for word overlap - any shared words indicate relevance
        label_overlap = len(query_words.intersection(result_words))
        
        # Include if there's any word overlap
        return label_overlap > 0
    
    def _is_geographic_relevant(self, result: Dict[str, str], original_query: str) -> bool:
        """Check if a result is geographically relevant."""
        query_words = set(w.lower() for w in original_query.split())
        result_words = set(w.lower() for w in result['label'].split())
        
        # Check for word overlap - any shared words indicate relevance
        label_overlap = len(query_words.intersection(result_words))
        
        # Also check if the type field indicates it's geographic
        type_field = result.get('type', '').lower()
        is_geographic_type = 'geographic' in type_field
        
        # Include if there's word overlap OR if it's explicitly marked as geographic
        return label_overlap > 0 or is_geographic_type
    
    def find_terms(self, topics: List[str]) -> Dict[str, List[Dict[str, str]]]:
        """Find FAST terms for multiple topics with the updated API."""
        if not topics:
            return {}
            
        results = {}
        
        for topic in topics:
            if not topic or topic.strip() == "":
                continue
                
            topic = topic.strip()
            
            # Skip if already processed
            if topic in results:
                continue
                
            # Search for terms - STRICT LIMIT TO 3
            fast_results = self.search(topic, topic)[:self.max_results]
            results[topic] = fast_results
            
            if fast_results:
                print(f"   Found {len(fast_results)} FAST terms for '{topic}'")
            else:
                print(f"   No FAST terms found for '{topic}'")
        
        return results

    def find_geographic_terms(self, geographic_entities: List[str]) -> Dict[str, List[Dict[str, str]]]:
        """Find FAST geographic terms for multiple entities."""
        if not geographic_entities:
            return {}
            
        results = {}
        
        for entity in geographic_entities:
            if not entity or entity.strip() == "":
                continue
                
            entity = entity.strip()
            
            # Skip if already processed
            if entity in results:
                continue
            
            # Extract the searchable part by removing the type in parentheses
            # e.g., "Baltimore--Maryland (City)" becomes "Baltimore--Maryland"
            search_term = entity
            if '(' in entity and entity.endswith(')'):
                # Find the last opening parenthesis and remove everything from there
                paren_index = entity.rfind('(')
                if paren_index > 0:
                    search_term = entity[:paren_index].strip()
                        
            # Search for geographic terms using the cleaned search term - LIMITED TO 1 TERM
            fast_results = self.search_geographic(search_term, entity)[:self.max_geo_results]  
            results[entity] = fast_results  # Store under the original entity name
            
            if fast_results:
                print(f"   Found FAST geographic term for '{entity}'")  
            else:
                print(f"   No FAST geographic term found for '{entity}'")

        return results

class GettyTermFinder:
    """Getty Vocabularies term finder (AAT, ULAN, TGN) with broader search strategy and logging."""
    
    def __init__(self, stats_tracker=None, logs_folder_path=None):
        self.base_urls = {
            'AAT': 'http://vocabsservices.getty.edu/AATService.asmx/AATGetTermMatch',
            'ULAN': 'http://vocabsservices.getty.edu/ULANService.asmx/ULANGetTermMatch',
            'TGN': 'http://vocabsservices.getty.edu/TGNService.asmx/TGNGetTermMatch'
        }
        self.headers = {
            'User-Agent': 'Python-Getty-Term-Finder/1.0 (Educational/Research Use)'
        }
        self.max_results = 3  # Strict limit to 3 terms
        self.request_delay = 0.5
        self.cache = {}
        self.stats_tracker = stats_tracker
        self.logs_folder_path = logs_folder_path
    
    def search_aat(self, query: str, topic: str = None) -> List[Dict[str, str]]:
        """Search AAT (Art & Architecture Thesaurus) for terms with broader matching and logging."""
        cache_key = f"AAT_{query}"
        from_cache = cache_key in self.cache
        
        start_time = time.time()
        
        if from_cache:
            results = self.cache[cache_key]
            processing_time = time.time() - start_time
            
            if self.stats_tracker:
                self.stats_tracker.record_api_call(
                    'Getty AAT', topic or query, query, True, processing_time, len(results), from_cache=True
                )
            
            if self.logs_folder_path:
                log_individual_vocab_response(
                    self.logs_folder_path, "southern_architect_step2", topic or query, 
                    "Getty AAT", f"{query} (CACHED)", results, processing_time
                )
            
            return results
        
        all_results = []
        error_msg = None
        
        try:
            # Strategy 1: Exact term search
            params = {
                'term': query,
                'logop': 'and',
                'notes': ''
            }
            
            resp = requests.get(self.base_urls['AAT'], params=params, headers=self.headers, timeout=10)
            resp.raise_for_status()
            
            root = ET.fromstring(resp.content)
            
            # Parse XML response
            for subject in root.findall('.//Subject'):
                if len(all_results) >= self.max_results:
                    break
                    
                subject_id = subject.find('Subject_ID')
                preferred_term = subject.find('Preferred_Term')
                
                if subject_id is not None and preferred_term is not None:
                    all_results.append({
                        'label': preferred_term.text,
                        'uri': f"http://vocab.getty.edu/aat/{subject_id.text}",
                        'source': 'Getty AAT',
                        'subject_id': subject_id.text
                    })
            
            # Strategy 2: If no results and multi-word query, try key terms
            if not all_results and ' ' in query:
                words = query.split()
                for word in words:
                    if len(word) > 3 and len(all_results) < self.max_results:  # Skip short words
                        params_word = {
                            'term': word,
                            'logop': 'and', 
                            'notes': ''
                        }
                        
                        try:
                            resp_word = requests.get(self.base_urls['AAT'], params=params_word, headers=self.headers, timeout=10)
                            resp_word.raise_for_status()
                            root_word = ET.fromstring(resp_word.content)
                            
                            for subject in root_word.findall('.//Subject'):
                                if len(all_results) >= self.max_results:
                                    break
                                    
                                subject_id = subject.find('Subject_ID')
                                preferred_term = subject.find('Preferred_Term')
                                
                                if subject_id is not None and preferred_term is not None:
                                    # Check if this result is relevant to original query
                                    term_lower = preferred_term.text.lower()
                                    query_words = set(w.lower() for w in query.split())
                                    term_words = set(w.lower() for w in preferred_term.text.split())
                                    
                                    # Include if there's word overlap or architectural relevance
                                    if (query_words.intersection(term_words) or 
                                        any(arch_word in term_lower for arch_word in ['architecture', 'architectural', 'building', 'style'])):
                                        all_results.append({
                                            'label': preferred_term.text,
                                            'uri': f"http://vocab.getty.edu/aat/{subject_id.text}",
                                            'source': 'Getty AAT',
                                            'subject_id': subject_id.text
                                        })
                        except:
                            continue  # Skip word if it fails
            
            success = True
            
        except Exception as e:
            error_msg = str(e)
            success = False
        
        processing_time = time.time() - start_time
        
        # Remove duplicates and STRICT LIMIT TO 3
        seen_uris = set()
        unique_results = []
        for result in all_results:
            if result['uri'] not in seen_uris and len(unique_results) < self.max_results:
                unique_results.append(result)
                seen_uris.add(result['uri'])
        
        results = unique_results
        self.cache[cache_key] = results
        
        # Log the API call
        if self.stats_tracker:
            self.stats_tracker.record_api_call(
                'Getty AAT', topic or query, query, success, processing_time, len(results)
            )
        
        if self.logs_folder_path:
            log_individual_vocab_response(
                self.logs_folder_path, "southern_architect_step2", topic or query, 
                "Getty AAT", query, results, processing_time, error_msg
            )
        
        time.sleep(self.request_delay)
        return results
    
    def search_tgn(self, query: str, topic: str = None) -> List[Dict[str, str]]:
        """Search TGN (Thesaurus of Geographic Names) for terms with broader matching and logging."""
        cache_key = f"TGN_{query}"
        from_cache = cache_key in self.cache
        
        start_time = time.time()
        
        if from_cache:
            results = self.cache[cache_key]
            processing_time = time.time() - start_time
            
            if self.stats_tracker:
                self.stats_tracker.record_api_call(
                    'Getty TGN', topic or query, query, True, processing_time, len(results), from_cache=True
                )
            
            if self.logs_folder_path:
                log_individual_vocab_response(
                    self.logs_folder_path, "southern_architect_step2", topic or query, 
                    "Getty TGN", f"{query} (CACHED)", results, processing_time
                )
            
            return results
        
        all_results = []
        error_msg = None
        
        try:
            # Strategy 1: Exact search
            params = {
                'name': query,
                'placetypeid': '',
                'nationid': ''
            }
            
            resp = requests.get(self.base_urls['TGN'], params=params, headers=self.headers, timeout=10)
            resp.raise_for_status()
            
            root = ET.fromstring(resp.content)
            
            # Parse XML response
            for subject in root.findall('.//Subject'):
                if len(all_results) >= self.max_results:
                    break
                    
                subject_id = subject.find('Subject_ID')
                preferred_term = subject.find('Preferred_Term')
                
                if subject_id is not None and preferred_term is not None:
                    all_results.append({
                        'label': preferred_term.text,
                        'uri': f"http://vocab.getty.edu/tgn/{subject_id.text}",
                        'source': 'Getty TGN',
                        'subject_id': subject_id.text
                    })
            
            # Strategy 2: If no results and contains geographic terms, try broader search
            if not all_results:
                geographic_words = ['american', 'southern', 'northern', 'eastern', 'western', 'city', 'state', 'county']
                query_words = query.lower().split()
                
                for word in query_words:
                    if len(all_results) >= self.max_results:
                        break
                        
                    if word in geographic_words or len(word) > 4:
                        params_word = {
                            'name': word,
                            'placetypeid': '',
                            'nationid': ''
                        }
                        
                        try:
                            resp_word = requests.get(self.base_urls['TGN'], params=params_word, headers=self.headers, timeout=10)
                            resp_word.raise_for_status()
                            root_word = ET.fromstring(resp_word.content)
                            
                            for subject in root_word.findall('.//Subject'):
                                if len(all_results) >= self.max_results:
                                    break
                                    
                                subject_id = subject.find('Subject_ID')
                                preferred_term = subject.find('Preferred_Term')
                                
                                if subject_id is not None and preferred_term is not None:
                                    all_results.append({
                                        'label': preferred_term.text,
                                        'uri': f"http://vocab.getty.edu/tgn/{subject_id.text}",
                                        'source': 'Getty TGN',
                                        'subject_id': subject_id.text
                                    })
                        except:
                            continue
                        
                        if all_results:  # Stop after first successful word search
                            break
            
            success = True
            
        except Exception as e:
            error_msg = str(e)
            success = False
        
        processing_time = time.time() - start_time
        
        # Remove duplicates and LIMIT TO 3
        seen_uris = set()
        unique_results = []
        for result in all_results:
            if result['uri'] not in seen_uris and len(unique_results) < self.max_results:
                unique_results.append(result)
                seen_uris.add(result['uri'])
        
        results = unique_results
        self.cache[cache_key] = results
        
        # Log the API call
        if self.stats_tracker:
            self.stats_tracker.record_api_call(
                'Getty TGN', topic or query, query, success, processing_time, len(results)
            )
        
        if self.logs_folder_path:
            log_individual_vocab_response(
                self.logs_folder_path, "southern_architect_step2", topic or query, 
                "Getty TGN", query, results, processing_time, error_msg
            )
        
        time.sleep(self.request_delay)
        return results
    
    def find_terms(self, topics: List[str]) -> Dict[str, List[Dict[str, str]]]:
        """Find Getty terms for multiple topics."""
        if not topics:
            return {}
            
        results = {}
        
        for topic in topics:
            if not topic or topic.strip() == "":
                continue
                
            topic = topic.strip()
            
            # Skip if already processed
            if topic in results:
                continue
                
            # Search AAT and TGN
            all_results = []
            
            # Search AAT (Art & Architecture Thesaurus)
            aat_results = self.search_aat(topic, topic)
            all_results.extend(aat_results)
            
            # Search TGN (Thesaurus of Geographic Names) - if we have room
            if len(all_results) < self.max_results:
                tgn_results = self.search_tgn(topic, topic)
                # Add TGN results up to the limit
                remaining_slots = self.max_results - len(all_results)
                all_results.extend(tgn_results[:remaining_slots])
            
            # LIMIT TO 3 TOTAL
            results[topic] = all_results[:self.max_results]
            
            if all_results:
                print(f"   Found {len(all_results)} Getty terms for '{topic}'")
            else:
                print(f"   No Getty terms found for '{topic}'")
        
        return results

class LOCAuthorizedTermFinder:
    """Enhanced LOC term finder with rate limiting, error handling, and logging."""
    
    def __init__(self, stats_tracker=None, logs_folder_path=None):
        self.base_url = "https://id.loc.gov/authorities/subjects/suggest2"
        self.headers = {
            'User-Agent': 'Python-LOC-Term-Finder/1.0 (Educational/Research Use)'
        }
        self.lcsh_authorized_headings = "http://id.loc.gov/authorities/subjects/collection_LCSHAuthorizedHeadings"
        self.max_results = 3  # Limit to 3 terms
        self.request_delay = 0.5
        self.cache = {}
        self.stats_tracker = stats_tracker
        self.logs_folder_path = logs_folder_path
        
    def search(self, query: str, search_type: str, topic: str = None) -> List[Dict[str, str]]:
        """Search LOC for authorized terms with logging."""
        cache_key = f"{query}_{search_type}"
        from_cache = cache_key in self.cache
        
        start_time = time.time()
        
        if from_cache:
            results = self.cache[cache_key]
            processing_time = time.time() - start_time
            
            if self.stats_tracker:
                self.stats_tracker.record_api_call(
                    'LCSH', topic or query, f"{query} ({search_type})", True, processing_time, len(results), from_cache=True
                )
            
            if self.logs_folder_path:
                log_individual_vocab_response(
                    self.logs_folder_path, "southern_architect_step2", topic or query, 
                    "LCSH", f"{query} ({search_type}) (CACHED)", results, processing_time
                )
            
            return results
        
        params = {
            'q': query,
            'searchtype': search_type,
            'count': self.max_results,
            'memberOf': self.lcsh_authorized_headings
        }
        
        results = []
        error_msg = None
        success = False
        
        try:
            resp = requests.get(self.base_url, params=params, headers=self.headers, timeout=10)
            resp.raise_for_status()
            hits = resp.json().get('hits', [])
            
            for h in hits:
                if len(results) >= self.max_results:
                    break
                if h.get('aLabel') and h.get('uri'):
                    results.append({
                        'label': h['aLabel'], 
                        'uri': h['uri'],
                        'source': 'LCSH'
                    })
            
            success = True
            
        except requests.exceptions.RequestException as e:
            error_msg = str(e)
            success = False
        
        processing_time = time.time() - start_time
        
        self.cache[cache_key] = results
        
        # Log the API call
        if self.stats_tracker:
            self.stats_tracker.record_api_call(
                'LCSH', topic or query, f"{query} ({search_type})", success, processing_time, len(results)
            )
        
        if self.logs_folder_path:
            log_individual_vocab_response(
                self.logs_folder_path, "southern_architect_step2", topic or query, 
                "LCSH", f"{query} ({search_type})", results, processing_time, error_msg
            )
        
        time.sleep(self.request_delay)
        return results
    
    def find_terms(self, topics: List[str]) -> Dict[str, List[Dict[str, str]]]:
        """Find LCSH terms for multiple topics with logging."""
        if not topics:
            return {}
            
        results = {}
        
        for topic in topics:
            if not topic or topic.strip() == "":
                continue
                
            topic = topic.strip()
            
            # Skip if already processed
            if topic in results:
                continue
                
            # Try keyword search first
            keyword_results = self.search(topic, "keyword", topic)
            
            # If we need more results, try left-anchored search
            if len(keyword_results) < self.max_results:
                leftanchored_results = self.search(topic, "leftanchored", topic)
                
                # Merge results, avoiding duplicates
                existing_uris = {r['uri'] for r in keyword_results}
                for result in leftanchored_results:
                    if len(keyword_results) >= self.max_results:
                        break
                    if result['uri'] not in existing_uris:
                        keyword_results.append(result)
            
            # LIMIT TO 3
            results[topic] = keyword_results[:self.max_results]
            
            if keyword_results:
                print(f"   Found {len(keyword_results)} LCSH terms for '{topic}'")
            else:
                print(f"   No LCSH terms found for '{topic}'")
        
        return results

class SouthernArchitectEnhancer:
    """Main class for enhancing Southern Architect results with multi-vocabulary terms including LCSH."""
    
    def __init__(self, folder_path: str):
        self.folder_path = folder_path
        
        # Create logs folder
        self.logs_folder_path = os.path.join(folder_path, "logs")
        if not os.path.exists(self.logs_folder_path):
            os.makedirs(self.logs_folder_path)
        
        # Initialize stats tracker
        self.stats_tracker = APIStatsTracker()
        
        # Initialize finders with stats tracking and logging
        self.lcsh_finder = LOCAuthorizedTermFinder(self.stats_tracker, self.logs_folder_path)
        self.fast_finder = FASTTermFinder(stats_tracker=self.stats_tracker, logs_folder_path=self.logs_folder_path)
        self.getty_finder = GettyTermFinder(self.stats_tracker, self.logs_folder_path)
        
        self.workflow_type = None
        self.json_data = None
        self.excel_path = None
        self.max_terms_per_vocabulary = 3  # For topics
        self.max_geo_terms_per_vocabulary = 1  # For geographic entities
        self.max_total_terms = 12  # 3 terms × 4 vocabularies = 12 max total for topics
    
    def detect_workflow_type(self) -> bool:
        """Detect whether this is a text or image workflow folder."""
        # Check for expected files in the metadata/collection_metadata subfolder
        metadata_dir = os.path.join(self.folder_path, "metadata", "collection_metadata")
        text_files = ['text_workflow.xlsx', 'text_workflow.json']
        image_files = ['image_workflow.xlsx', 'image_workflow.json']
        
        has_text_files = all(os.path.exists(os.path.join(metadata_dir, f)) for f in text_files)
        has_image_files = all(os.path.exists(os.path.join(metadata_dir, f)) for f in image_files)
            
        if has_text_files and not has_image_files:
            self.workflow_type = 'text'
            self.excel_path = os.path.join(self.folder_path, "metadata", "collection_metadata", "text_workflow.xlsx")
            return True
        elif has_image_files and not has_text_files:
            self.workflow_type = 'image'
            self.excel_path = os.path.join(self.folder_path, "metadata", "collection_metadata", "image_workflow.xlsx")
            return True
        elif has_text_files and has_image_files:
            logging.error("Both text and image workflow files found. Please specify workflow type.")
            return False
        else:
            logging.error("No recognized workflow files found in the folder.")
            return False
    
    def load_json_data(self) -> bool:
        """Load the JSON data from the appropriate workflow file."""
        # Save the enhanced JSON in the collection_metadata folder
        json_filename = f"{self.workflow_type}_workflow.json"
        collection_metadata_dir = os.path.join(self.folder_path, "metadata", "collection_metadata")
        json_path = os.path.join(collection_metadata_dir, json_filename)

        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                self.json_data = json.load(f)
            print(f"Loaded JSON data from {json_filename}")
            return True
        except Exception as e:
            logging.error(f"Error loading JSON data: {e}")
            return False
    
    def extract_subject_headings(self) -> tuple[List[str], List[str]]:
        """Extract all unique topics and geographic entities from the JSON data."""
        all_subjects = set()
        all_geographic_entities = set()
        
        # Skip the last item if it's API stats
        data_items = self.json_data[:-1] if self.json_data and 'api_stats' in self.json_data[-1] else self.json_data
        
        for item in data_items:
            if 'analysis' in item:
                # Extract topics (subjects)
                if 'topics' in item['analysis']:
                    subjects = item['analysis']['topics']
                    if isinstance(subjects, list):
                        for subject in subjects:
                            if subject and subject.strip():
                                all_subjects.add(subject.strip())
                    elif isinstance(subjects, str) and subjects.strip():
                        # Handle comma-separated string format
                        for subject in subjects.split(','):
                            if subject.strip():
                                all_subjects.add(subject.strip())
                
                # Extract geographic entities
                if 'geographic_entities' in item['analysis']:
                    geo_entities = item['analysis']['geographic_entities']
                    if isinstance(geo_entities, list):
                        for entity in geo_entities:
                            if entity and entity.strip():
                                all_geographic_entities.add(entity.strip())
                    elif isinstance(geo_entities, str) and geo_entities.strip():
                        # Handle comma-separated string format
                        for entity in geo_entities.split(','):
                            if entity.strip():
                                all_geographic_entities.add(entity.strip())
        
        return sorted(list(all_subjects)), sorted(list(all_geographic_entities))
    
    def extract_chronological_terms_from_issues(self) -> List[str]:
        """Extract chronological terms dynamically from issue folder names."""
        
        # Skip API stats for processing
        data_items = self.json_data[:-1] if self.json_data and 'api_stats' in self.json_data[-1] else self.json_data
        
        # Collect all unique years from folder names
        years_found = set()
        for item in data_items:
            folder_name = item.get('folder', '')
            # Extract 4-digit year from folder name (e.g., "1923-10-24" -> "1923")
            year_match = re.search(r'\b(18\d{2}|19\d{2})\b', folder_name)
            if year_match:
                years_found.add(int(year_match.group(1)))
        
        if not years_found:
            print("No years found in folder names")
            return []
        
        print(f"Found years in collection: {sorted(years_found)}")
        
        # Generate chronological terms
        chronological_terms = set()
        decades_found = set()
        centuries_found = set()
        
        for year in years_found:
            # Determine decade and century
            decade = (year // 10) * 10  # e.g., 1923 -> 1920
            century = 19 if year < 1900 else 20
            
            decades_found.add(decade)
            centuries_found.add(century)
        
        # Map decades to LCSH spelled-out format
        decade_mapping = {
            1890: "Eighteen nineties",
            1900: "Nineteen hundreds (Decade)",  # 1900-1909
            1910: "Nineteen tens", 
            1920: "Nineteen twenties",
            1930: "Nineteen thirties"
        }
        
        # Add decade terms
        for decade in decades_found:
            lcsh_decade = decade_mapping.get(decade)
            if lcsh_decade:
                chronological_terms.add(lcsh_decade)
                print(f"  Adding decade term: {lcsh_decade}")
        
        # Add architecture-specific century terms
        for century in centuries_found:
            if century == 19:
                chronological_terms.add("Architecture--United States--History--19th century")
                print(f"  Adding architecture century term: Architecture--United States--History--19th century")
            elif century == 20:
                chronological_terms.add("Architecture--United States--History--20th century") 
                print(f"  Adding architecture century term: Architecture--United States--History--20th century")
        
        return list(chronological_terms)
    def process_multi_vocabulary_lookup(self, subjects: List[str], geographic_entities: List[str], chronological_terms: List[str]) -> Tuple[Dict[str, str], Dict[str, List[Dict[str, str]]], Dict[str, str], Dict[str, List[Dict[str, str]]], Dict[str, str], Dict[str, List[Dict[str, str]]]]:
        """Process multi-vocabulary lookup for subjects, geographic entities, and chronological terms."""
        
        print(f"\nProcessing multi-vocabulary lookup...")
        print(f"API calls will be logged to: {self.logs_folder_path}")
        
        # Process subjects
        subject_to_terms_excel = {}
        subject_to_terms_json = {}
        
        for i, subject in enumerate(subjects, 1):
            
            # Search all vocabularies for THIS SPECIFIC subject
            all_terms = []
            
            # Search LCSH FIRST - LIMIT TO 3 TERMS
            lcsh_results = self.lcsh_finder.find_terms([subject])
            if subject in lcsh_results:
                lcsh_terms = lcsh_results[subject][:self.max_terms_per_vocabulary]
                all_terms.extend(lcsh_terms)
            
            # Search FAST - LIMIT TO 3 TERMS
            fast_results = self.fast_finder.find_terms([subject])
            if subject in fast_results:
                fast_terms = fast_results[subject][:self.max_terms_per_vocabulary]
                all_terms.extend(fast_terms)
            
            # Search Getty - LIMIT TO 3 TERMS
            getty_results = self.getty_finder.find_terms([subject])
            if subject in getty_results:
                getty_terms = getty_results[subject][:self.max_terms_per_vocabulary]
                all_terms.extend(getty_terms)
                        
            # Format results for this subject
            formatted_terms_excel = self.format_results_for_excel(all_terms)
            formatted_terms_json = self.format_results_for_json(all_terms)
            
            # Store results for this specific subject
            subject_to_terms_excel[subject] = formatted_terms_excel
            subject_to_terms_json[subject] = formatted_terms_json
        
        # Process geographic entities
        geographic_to_terms_excel = {}
        geographic_to_terms_json = {}
        
        for i, entity in enumerate(geographic_entities, 1):
            print(f"\nProcessing geographic entity {i}/{len(geographic_entities)}: '{entity}'")
            
            # Search FAST Geographic only - LIMITED TO 3 TERMS
            fast_geo_results = self.fast_finder.find_geographic_terms([entity])
            if entity in fast_geo_results:
                geo_terms = fast_geo_results[entity][:self.max_geo_terms_per_vocabulary]
            else:
                geo_terms = []
                print(f"     No FAST Geographic term found")

            # Format results for this geographic entity
            formatted_terms_excel = self.format_results_for_excel(geo_terms)
            formatted_terms_json = self.format_results_for_json(geo_terms)
            
            # Store results for this specific geographic entity
            geographic_to_terms_excel[entity] = formatted_terms_excel
            geographic_to_terms_json[entity] = formatted_terms_json
        
            # Process chronological terms - search LCSH for actual URIs with quoted searches
            chronological_to_terms_excel = {}
            chronological_to_terms_json = {}

            for chronological_term in chronological_terms:
                print(f"\nProcessing chronological term: '{chronological_term}'")
                
                # Search LCSH for this chronological term with quotes for exact match
                quoted_term = f'"{chronological_term}"'  # Add quotes for exact matching
                lcsh_results = self.lcsh_finder.find_terms([quoted_term])
                
                if quoted_term in lcsh_results and lcsh_results[quoted_term]:
                    # Found LCSH term with URI using quoted search
                    lcsh_terms = lcsh_results[quoted_term][:1]  # Only take 1 term
                    
                    # Update the label to remove quotes from display
                    for term in lcsh_terms:
                        term['label'] = chronological_term  # Use original unquoted term for display
                    
                    # Format results for this chronological term
                    formatted_terms_excel = self.format_results_for_excel(lcsh_terms)
                    formatted_terms_json = self.format_results_for_json(lcsh_terms)
                    
                    chronological_to_terms_excel[chronological_term] = formatted_terms_excel
                    chronological_to_terms_json[chronological_term] = formatted_terms_json
                    
                    print(f"   Found LCSH chronological term with URI (quoted search): {chronological_term}")
                else:
                    # If quoted search fails, try without quotes as fallback
                    lcsh_results_unquoted = self.lcsh_finder.find_terms([chronological_term])
                    
                    if chronological_term in lcsh_results_unquoted and lcsh_results_unquoted[chronological_term]:
                        # Found with unquoted search
                        lcsh_terms = lcsh_results_unquoted[chronological_term][:1]  # Only take 1 term
                        
                        # Format results for this chronological term
                        formatted_terms_excel = self.format_results_for_excel(lcsh_terms)
                        formatted_terms_json = self.format_results_for_json(lcsh_terms)
                        
                        chronological_to_terms_excel[chronological_term] = formatted_terms_excel
                        chronological_to_terms_json[chronological_term] = formatted_terms_json
                        
                        print(f"   Found LCSH chronological term with URI (unquoted fallback): {chronological_term}")
                    else:
                        # Fallback: create term without URI if not found in LCSH at all
                        term_object = {
                            'label': chronological_term,
                            'uri': '',  # No URI available
                            'source': 'LCSH Chronological (generated)',
                            'description': f'Generated chronological term for Southern Architect collection',
                            'type': 'chronological'
                        }
                        
                        # Format for Excel and JSON
                        formatted_excel = f"{chronological_term} [{term_object['source']}]"
                        formatted_json = [term_object]
                        
                        chronological_to_terms_excel[chronological_term] = formatted_terms_excel
                        chronological_to_terms_json[chronological_term] = formatted_json
                        
                        print(f"   Created chronological term without URI: {chronological_term}")
        
        return (subject_to_terms_excel, subject_to_terms_json, 
            geographic_to_terms_excel, geographic_to_terms_json,
            chronological_to_terms_excel, chronological_to_terms_json)
    
    def format_results_for_excel(self, terms: List[Dict[str, str]]) -> str:
        """Format results for spreadsheet display with labels, URIs, and sources."""
        if not terms:
            return ""
        
        formatted_terms = []
        for term in terms:
            source = term.get('source', 'Unknown')
            label = term.get('label', '')
            uri = term.get('uri', '')
            
            if label and uri:
                formatted_terms.append(f"{label} ({uri}) [{source}]")
            elif label:
                formatted_terms.append(f"{label} [{source}]")
        
        return "; ".join(formatted_terms)
    
    def format_results_for_json(self, terms: List[Dict[str, str]]) -> List[Dict[str, str]]:
        """Format results for JSON storage with full structure."""
        if not terms:
            return []
        
        formatted_terms = []
        seen_uris = set()
        
        for term in terms:
            uri = term.get('uri', '')
            if uri and uri not in seen_uris:
                formatted_terms.append({
                    'label': term.get('label', ''),
                    'uri': uri,
                    'source': term.get('source', 'Unknown'),
                    'description': term.get('description', ''),
                    'qid': term.get('qid', ''),
                    'subject_id': term.get('subject_id', ''),
                    'type': term.get('type', ''),
                    'tag': term.get('tag', '')
                })
                seen_uris.add(uri)
        
        return formatted_terms
    
    def enhance_excel_file(self, subject_to_terms: Dict[str, str], geographic_to_terms: Dict[str, str], chronological_to_terms: Dict[str, str]) -> bool:
        """Add vocabulary terms columns to the Excel file for both topics, geographic entities, and chronological terms."""
        try:
            # Load the existing workbook
            wb = load_workbook(self.excel_path)
            analysis_sheet = wb['Analysis']
            
            # Determine the column structure based on workflow type
            if self.workflow_type == 'text':
                subject_col = 8   # Topics column
                geo_col = 7       # Geographic Entities column  
                insert_col = 9    # Insert vocabulary terms after Content Warning (column 9)
            else:  # image workflow
                subject_col = 9   # Topics column
                geo_col = 8       # Geographic Entities column
                insert_col = 10   # Insert vocabulary terms after Content Warning (column 10)
            
            # Insert THREE new columns (one for topic vocab terms, one for geographic vocab terms, one for chronological terms)
            analysis_sheet.insert_cols(insert_col, 3)
            
            # Add headers
            topic_vocab_header = analysis_sheet.cell(row=1, column=insert_col)
            topic_vocab_header.value = "Topic Vocabulary Terms (Max 3 per vocab per topic)"
            topic_vocab_header.alignment = Alignment(vertical='top', wrap_text=True)
            
            geo_vocab_header = analysis_sheet.cell(row=1, column=insert_col + 1)
            geo_vocab_header.value = "Geographic Vocabulary Terms (FAST only)"
            geo_vocab_header.alignment = Alignment(vertical='top', wrap_text=True)
            
            # NEW: Add chronological terms header
            chrono_vocab_header = analysis_sheet.cell(row=1, column=insert_col + 2)
            chrono_vocab_header.value = "Chronological Terms (LCSH)"
            chrono_vocab_header.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Set column widths
            topic_col_letter = topic_vocab_header.column_letter
            geo_col_letter = geo_vocab_header.column_letter
            chrono_col_letter = chrono_vocab_header.column_letter
            analysis_sheet.column_dimensions[topic_col_letter].width = 50
            analysis_sheet.column_dimensions[geo_col_letter].width = 50
            analysis_sheet.column_dimensions[chrono_col_letter].width = 50
            
            # Process each data row
            processed_rows = 0
            for row_num in range(2, analysis_sheet.max_row + 1):
                # Get the topics and geographic entities from the current row
                subject_cell = analysis_sheet.cell(row=row_num, column=subject_col)
                geo_cell = analysis_sheet.cell(row=row_num, column=geo_col)
                
                topics = subject_cell.value or ""
                geo_entities = geo_cell.value or ""
                
                # Process topics vocabulary terms
                topic_vocab_terms = []
                if topics and topics.strip():
                    subjects = [s.strip() for s in topics.split(',') if s.strip()]
                    print(f"Row {row_num-1}: Processing {len(subjects)} topics: {subjects}")
                    
                    for subject in subjects:
                        if subject in subject_to_terms and subject_to_terms[subject]:
                            terms = [t.strip() for t in subject_to_terms[subject].split(';') if t.strip()]
                            topic_vocab_terms.extend(terms)
                            print(f"  - '{subject}': {len(terms)} terms")
                        else:
                            print(f"  - '{subject}': No terms found")
                
                # Process geographic entities vocabulary terms
                geo_vocab_terms = []
                if geo_entities and geo_entities.strip():
                    entities = [e.strip() for e in geo_entities.split(',') if e.strip()]
                    print(f"Row {row_num-1}: Processing {len(entities)} geographic entities: {entities}")
                    
                    for entity in entities:
                        if entity in geographic_to_terms and geographic_to_terms[entity]:
                            terms = [t.strip() for t in geographic_to_terms[entity].split(';') if t.strip()]
                            geo_vocab_terms.extend(terms)
                            print(f"  - '{entity}': {len(terms)} geographic terms")
                        else:
                            print(f"  - '{entity}': No geographic terms found")
                
                # NEW: Process chronological terms - these apply to ALL rows since they're collection-wide
                chrono_vocab_terms = []
                for chrono_term in chronological_to_terms.values():
                    if chrono_term:  # Only add non-empty terms
                        terms = [t.strip() for t in chrono_term.split(';') if t.strip()]
                        chrono_vocab_terms.extend(terms)
                
                # Remove duplicates while preserving order for all three types
                unique_topic_terms = []
                seen_topic = set()
                for term in topic_vocab_terms:
                    if term not in seen_topic:
                        unique_topic_terms.append(term)
                        seen_topic.add(term)
                
                unique_geo_terms = []
                seen_geo = set()
                for term in geo_vocab_terms:
                    if term not in seen_geo:
                        unique_geo_terms.append(term)
                        seen_geo.add(term)
                
                unique_chrono_terms = []
                seen_chrono = set()
                for term in chrono_vocab_terms:
                    if term not in seen_chrono:
                        unique_chrono_terms.append(term)
                        seen_chrono.add(term)
                
                print(f"  → Total unique topic terms: {len(unique_topic_terms)}")
                print(f"  → Total unique geographic terms: {len(unique_geo_terms)}")
                print(f"  → Total unique chronological terms: {len(unique_chrono_terms)}")
                
                # Set the vocabulary terms cell values
                topic_vocab_cell = analysis_sheet.cell(row=row_num, column=insert_col)
                topic_vocab_cell.value = "; ".join(unique_topic_terms) if unique_topic_terms else ""
                topic_vocab_cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                geo_vocab_cell = analysis_sheet.cell(row=row_num, column=insert_col + 1)
                geo_vocab_cell.value = "; ".join(unique_geo_terms) if unique_geo_terms else ""
                geo_vocab_cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                # NEW: Set chronological terms cell value
                chrono_vocab_cell = analysis_sheet.cell(row=row_num, column=insert_col + 2)
                chrono_vocab_cell.value = "; ".join(unique_chrono_terms) if unique_chrono_terms else ""
                chrono_vocab_cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                if unique_topic_terms or unique_geo_terms or unique_chrono_terms:
                    processed_rows += 1
            
            # Save the enhanced workbook
            wb.save(self.excel_path)
            print(f"Enhanced Excel file saved with vocabulary terms in {processed_rows} rows")
            return True
            
        except Exception as e:
            logging.error(f"Error enhancing Excel file: {e}")
            return False
    
    def enhance_json_file(self, subject_to_terms_json: Dict[str, List[Dict[str, str]]], 
                    geographic_to_terms_json: Dict[str, List[Dict[str, str]]],
                    chronological_to_terms_json: Dict[str, List[Dict[str, str]]]) -> bool:
        """Add vocabulary search results to JSON file with topic-to-terms, geographic-to-terms, and chronological terms mapping."""
        try:
            # Skip the last item if it's API stats
            data_items = self.json_data[:-1] if self.json_data and 'api_stats' in self.json_data[-1] else self.json_data
            api_stats = self.json_data[-1] if self.json_data and 'api_stats' in self.json_data[-1] else None
            
            enhanced_items = []
            processed_items = 0
            
            # NEW: Prepare collection-wide chronological terms (same for all items)
            all_chronological_terms = []
            for term_list in chronological_to_terms_json.values():
                all_chronological_terms.extend(term_list)
            
            for item in data_items:
                if 'analysis' in item:
                    # Get topics for this item
                    topics = item['analysis'].get('topics', [])
                    geographic_entities = item['analysis'].get('geographic_entities', [])
                    
                    # Normalize topics to list format
                    if isinstance(topics, str):
                        subjects = [s.strip() for s in topics.split(',') if s.strip()]
                    else:
                        subjects = topics if isinstance(topics, list) else []
                    
                    # Normalize geographic entities to list format
                    if isinstance(geographic_entities, str):
                        geo_entities = [e.strip() for e in geographic_entities.split(',') if e.strip()]
                    else:
                        geo_entities = geographic_entities if isinstance(geographic_entities, list) else []
                    
                    # Create topic-to-terms mapping for this page
                    topic_to_terms = {}
                    for subject in subjects:
                        if subject in subject_to_terms_json and subject_to_terms_json[subject]:
                            topic_to_terms[subject] = subject_to_terms_json[subject].copy()
                    
                    # Create geographic-entity-to-terms mapping for this page
                    geographic_to_terms = {}
                    for entity in geo_entities:
                        if entity in geographic_to_terms_json and geographic_to_terms_json[entity]:
                            geographic_to_terms[entity] = geographic_to_terms_json[entity].copy()
                    
                    # Add all mappings to the analysis
                    item['analysis']['vocabulary_search_results'] = topic_to_terms
                    item['analysis']['geographic_vocabulary_search_results'] = geographic_to_terms
                    # NEW: Add chronological terms to all items (they apply to the whole collection)
                    item['analysis']['chronological_vocabulary_terms'] = all_chronological_terms.copy()
                    
                    if topic_to_terms or geographic_to_terms or all_chronological_terms:
                        processed_items += 1
                
                enhanced_items.append(item)
            
            # Add API stats back if it existed
            if api_stats:
                enhanced_items.append(api_stats)
            
            # Save the enhanced JSON
            json_filename = f"{self.workflow_type}_workflow.json"
            # Saves in collection_metadata folder
            collection_metadata_dir = os.path.join(self.folder_path, "metadata", "collection_metadata")
            json_path = os.path.join(collection_metadata_dir, json_filename)
            
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(enhanced_items, f, indent=2, ensure_ascii=False)
            
            return True
            
        except Exception as e:
            logging.error(f"Error enhancing JSON file: {e}")
            return False
    
    def create_vocabulary_report(self, subject_to_terms: Dict[str, str], 
                            geographic_to_terms: Dict[str, str], 
                            chronological_to_terms: Dict[str, str]) -> bool:
        """Create a detailed vocabulary mapping report organized by page for topics, geographic entities, and chronological terms."""
        try:
            # Save vocabulary report in the collection_metadata folder
            collection_metadata_dir = os.path.join(self.folder_path, "metadata", "collection_metadata")
            report_path = os.path.join(collection_metadata_dir, "vocabulary_mapping_report.txt")
            
            # Create mappings of page number to topics/geographic entities and folder info
            page_to_topics = defaultdict(list)
            page_to_geographic = defaultdict(list)
            page_to_folder = {}  
            
            # Load the existing workbook to get page information
            wb = load_workbook(self.excel_path)
            analysis_sheet = wb['Analysis']
            
            # Determine the column structure based on workflow type
            if self.workflow_type == 'text':
                page_col = 2     # Page Number column (1-indexed: Folder=1, Page Number=2)
                subject_col = 8  # Topics column (1-indexed: ..., Topics=8)
                geo_col = 7      # Geographic Entities column (1-indexed: ..., Geographic Entities=7)
            else:  # image workflow
                page_col = 2     # Page Number column (1-indexed: Folder=1, Page Number=2)
                subject_col = 9  # Topics column (1-indexed: ..., Topics=9)
                geo_col = 8      # Geographic Entities column (1-indexed: ..., Geographic Entities=8)
            
            # Build mapping of page to topics and geographic entities
            for row_num in range(2, analysis_sheet.max_row + 1):
                page_cell = analysis_sheet.cell(row=row_num, column=page_col)
                subject_cell = analysis_sheet.cell(row=row_num, column=subject_col)
                geo_cell = analysis_sheet.cell(row=row_num, column=geo_col)
                
                page_number = page_cell.value or "Unknown"
                folder_cell = analysis_sheet.cell(row=row_num, column=1)  # Folder column
                folder_name = folder_cell.value or "Unknown"
                page_to_folder[page_number] = folder_name  # Store folder for this page
                topics = subject_cell.value or ""
                geo_entities = geo_cell.value or ""
                
                # Process topics
                if topics and topics.strip():
                    subjects = [s.strip() for s in topics.split(',') if s.strip()]
                    for subject in subjects:
                        if subject:
                            page_to_topics[page_number].append(subject)
                
                # Process geographic entities
                if geo_entities and geo_entities.strip():
                    entities = [e.strip() for e in geo_entities.split(',') if e.strip()]
                    for entity in entities:
                        if entity:
                            page_to_geographic[page_number].append(entity)
            
            wb.close()
            
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write("SOUTHERN ARCHITECT VOCABULARY MAPPING REPORT\n")
                f.write("=" * 50 + "\n\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Workflow Type: {self.workflow_type.upper()}\n")
                f.write(f"Total Topics Processed: {len(subject_to_terms)}\n")
                f.write(f"Total Geographic Entities Processed: {len(geographic_to_terms)}\n")
                # NEW: Add chronological terms count
                f.write(f"Total Chronological Terms Generated: {len(chronological_to_terms)}\n")
                f.write(f"Terms Per Vocabulary Limit: {self.max_terms_per_vocabulary}\n")
                f.write(f"Max Total Terms Per Topic: {self.max_total_terms}\n\n")
                
                f.write("VOCABULARIES SEARCHED:\n")
                f.write("FOR TOPICS:\n")
                f.write("- LCSH (Library of Congress Subject Headings)\n")
                f.write("- FAST (Faceted Application of Subject Terminology)\n")
                f.write("- Getty AAT (Art & Architecture Thesaurus)\n")
                f.write("- Getty TGN (Thesaurus of Geographic Names)\n")
                f.write("FOR GEOGRAPHIC ENTITIES:\n")
                f.write("- FAST Geographic (Faceted Application of Subject Terminology - Geographic)\n")
                # NEW: Add chronological section
                f.write("FOR CHRONOLOGICAL COVERAGE:\n")
                f.write("- LCSH (Library of Congress Subject Headings - Generated from issue dates)\n\n")
                
                # Statistics
                topics_with_terms = sum(1 for terms in subject_to_terms.values() if terms)
                topics_without_terms = len(subject_to_terms) - topics_with_terms
                geo_with_terms = sum(1 for terms in geographic_to_terms.values() if terms)
                geo_without_terms = len(geographic_to_terms) - geo_with_terms
                # NEW: Add chronological statistics
                chrono_with_terms = sum(1 for terms in chronological_to_terms.values() if terms)
                chrono_without_terms = len(chronological_to_terms) - chrono_with_terms
                
                # Get all pages that have either topics or geographic entities
                all_pages = set(list(page_to_topics.keys()) + list(page_to_geographic.keys()))
                total_pages = len(all_pages)
                
                f.write("STATISTICS:\n")
                f.write(f"- Total pages processed: {total_pages}\n")
                f.write(f"- Topics with vocabulary terms: {topics_with_terms}\n")
                f.write(f"- Topics without vocabulary terms: {topics_without_terms}\n")
                if len(subject_to_terms) > 0:
                    f.write(f"- Topic success rate: {(topics_with_terms/len(subject_to_terms)*100):.1f}%\n")
                else:
                    f.write("- Topic success rate: 0%\n")
                f.write(f"- Geographic entities with vocabulary terms: {geo_with_terms}\n")
                f.write(f"- Geographic entities without vocabulary terms: {geo_without_terms}\n")
                if len(geographic_to_terms) > 0:
                    f.write(f"- Geographic success rate: {(geo_with_terms/len(geographic_to_terms)*100):.1f}%\n")
                else:
                    f.write("- Geographic success rate: 0%\n")
                # NEW: Add chronological statistics
                f.write(f"- Chronological terms with vocabulary URIs: {chrono_with_terms}\n")
                f.write(f"- Chronological terms without vocabulary URIs: {chrono_without_terms}\n")
                if len(chronological_to_terms) > 0:
                    f.write(f"- Chronological success rate: {(chrono_with_terms/len(chronological_to_terms)*100):.1f}%\n\n")
                else:
                    f.write("- Chronological success rate: 0%\n\n")
                
                # Count terms by source for topics
                topic_source_counts = {'LCSH': 0, 'FAST': 0, 'Getty AAT': 0, 'Getty TGN': 0}
                for terms_str in subject_to_terms.values():
                    if terms_str:
                        for term in terms_str.split(';'):
                            if '[LCSH]' in term:
                                topic_source_counts['LCSH'] += 1
                            elif '[FAST]' in term:
                                topic_source_counts['FAST'] += 1
                            elif '[Getty AAT]' in term:
                                topic_source_counts['Getty AAT'] += 1
                            elif '[Getty TGN]' in term:
                                topic_source_counts['Getty TGN'] += 1
                
                # Count terms by source for geographic entities
                geo_source_counts = {'FAST Geographic': 0}
                for terms_str in geographic_to_terms.values():
                    if terms_str:
                        for term in terms_str.split(';'):
                            if '[FAST Geographic]' in term:
                                geo_source_counts['FAST Geographic'] += 1
                
                # NEW: Count terms by source for chronological terms
                chrono_source_counts = {'LCSH': 0, 'LCSH Chronological (generated)': 0}
                for terms_str in chronological_to_terms.values():
                    if terms_str:
                        for term in terms_str.split(';'):
                            if '[LCSH]' in term:
                                chrono_source_counts['LCSH'] += 1
                            elif '[LCSH Chronological (generated)]' in term:
                                chrono_source_counts['LCSH Chronological (generated)'] += 1
                
                f.write("TOPIC TERMS BY SOURCE:\n")
                for source, count in topic_source_counts.items():
                    f.write(f"- {source}: {count} terms\n")
                f.write("\nGEOGRAPHIC TERMS BY SOURCE:\n")
                for source, count in geo_source_counts.items():
                    f.write(f"- {source}: {count} terms\n")
                # NEW: Add chronological terms by source
                f.write("\nCHRONOLOGICAL TERMS BY SOURCE:\n")
                for source, count in chrono_source_counts.items():
                    f.write(f"- {source}: {count} terms\n")
                f.write("\n")
                
                # NEW: Add chronological terms section (collection-wide)
                f.write("CHRONOLOGICAL TERMS (COLLECTION-WIDE):\n")
                f.write("-" * 40 + "\n")
                if chronological_to_terms:
                    for chrono_term, vocab_terms in sorted(chronological_to_terms.items()):
                        f.write(f"Chronological Term: {chrono_term}\n")
                        if vocab_terms:
                            f.write(f"LCSH Terms: {vocab_terms}\n")
                        else:
                            f.write("LCSH Terms: No terms found\n")
                        f.write("\n")
                else:
                    f.write("No chronological terms generated for this collection.\n")
                f.write("\n")
                
                # Page statistics
                f.write("PAGE STATISTICS:\n")
                pages_with_topic_terms = 0
                pages_with_geo_terms = 0
                pages_with_any_terms = 0
                
                for page_num in all_pages:
                    topics_on_page = page_to_topics.get(page_num, [])
                    geo_entities_on_page = page_to_geographic.get(page_num, [])
                    
                    page_has_topic_terms = any(subject_to_terms.get(topic, '') for topic in topics_on_page)
                    page_has_geo_terms = any(geographic_to_terms.get(entity, '') for entity in geo_entities_on_page)
                    
                    if page_has_topic_terms:
                        pages_with_topic_terms += 1
                    if page_has_geo_terms:
                        pages_with_geo_terms += 1
                    if page_has_topic_terms or page_has_geo_terms:
                        pages_with_any_terms += 1
                
                f.write(f"- Pages with topic vocabulary terms: {pages_with_topic_terms}/{total_pages}\n")
                f.write(f"- Pages with geographic vocabulary terms: {pages_with_geo_terms}/{total_pages}\n")
                f.write(f"- Pages with any vocabulary terms: {pages_with_any_terms}/{total_pages}\n")
                # NEW: Add note about chronological terms
                f.write(f"- Pages with chronological terms: {total_pages}/{total_pages} (applied to all pages)\n")
                if total_pages > 0:
                    f.write(f"- Overall page success rate: {(pages_with_any_terms/total_pages*100):.1f}%\n")
                f.write("\n")
                
                # Rest of the existing detailed mappings code stays the same...
                # (The detailed page-by-page breakdown section doesn't need changes since 
                # chronological terms are collection-wide, not page-specific)
                
                # Detailed mappings organized by page
                f.write("DETAILED MAPPINGS BY PAGE:\n")
                f.write("-" * 40 + "\n\n")
                
                # Sort pages numerically if possible
                try:
                    sorted_pages = sorted(all_pages, key=lambda x: int(x) if str(x).isdigit() else float('inf'))
                except:
                    sorted_pages = sorted(all_pages, key=str)
                
                for page_num in sorted_pages:
                    topics_on_page = page_to_topics.get(page_num, [])
                    geo_entities_on_page = page_to_geographic.get(page_num, [])
                    
                    folder_name = page_to_folder.get(page_num, "Unknown")
                    f.write(f"PAGE {page_num} (ISSUE: {folder_name}):\n")
                    
                    if not topics_on_page and not geo_entities_on_page:
                        f.write("No topics or geographic entities found for this page\n")
                        # NEW: Still mention chronological terms apply
                        f.write(f"Collection chronological terms apply: {len(chronological_to_terms)} terms\n\n")
                        continue
                    
                    # Show topics for this page
                    if topics_on_page:
                        unique_topics = list(set(topics_on_page))
                        f.write(f"Topics on this page ({len(unique_topics)}): {', '.join(sorted(unique_topics))}\n")
                    
                    # Show geographic entities for this page
                    if geo_entities_on_page:
                        unique_geo = list(set(geo_entities_on_page))
                        f.write(f"Geographic entities on this page ({len(unique_geo)}): {', '.join(sorted(unique_geo))}\n")
                    
                    # NEW: Show chronological terms (same for all pages)
                    f.write(f"Collection chronological terms ({len(chronological_to_terms)}): {', '.join(sorted(chronological_to_terms.keys()))}\n")
                    
                    f.write("\n")
                    
                    # Show vocabulary terms for each unique topic on this page
                    if topics_on_page:
                        f.write("TOPIC VOCABULARY TERMS:\n")
                        unique_topics_on_page = sorted(list(set(topics_on_page)))
                        
                        for topic in unique_topics_on_page:
                            vocab_terms = subject_to_terms.get(topic, '')
                            f.write(f"  Topic: {topic}\n")
                            if vocab_terms:
                                f.write(f"  Terms: {vocab_terms}\n")
                            else:
                                f.write(f"  Terms: No terms found\n")
                            f.write("\n")
                    
                    # Show vocabulary terms for each unique geographic entity on this page
                    if geo_entities_on_page:
                        f.write("GEOGRAPHIC VOCABULARY TERMS:\n")
                        unique_geo_entities_on_page = sorted(list(set(geo_entities_on_page)))
                        
                        for entity in unique_geo_entities_on_page:
                            vocab_terms = geographic_to_terms.get(entity, '')
                            f.write(f"  Geographic Entity: {entity}\n")
                            if vocab_terms:
                                f.write(f"  Terms: {vocab_terms}\n")
                            else:
                                f.write(f"  Terms: No terms found\n")
                            f.write("\n")
                    
                    # NEW: Show chronological vocabulary terms (same for all pages)
                    if chronological_to_terms:
                        f.write("CHRONOLOGICAL VOCABULARY TERMS (COLLECTION-WIDE):\n")
                        for chrono_term, vocab_terms in sorted(chronological_to_terms.items()):
                            f.write(f"  Chronological Term: {chrono_term}\n")
                            if vocab_terms:
                                f.write(f"  Terms: {vocab_terms}\n")
                            else:
                                f.write(f"  Terms: No terms found\n")
                            f.write("\n")
                    
                    # Page summary
                    unique_topics_on_page = list(set(topics_on_page)) if topics_on_page else []
                    unique_geo_entities_on_page = list(set(geo_entities_on_page)) if geo_entities_on_page else []
                    
                    topics_with_terms_on_page = sum(1 for topic in unique_topics_on_page if subject_to_terms.get(topic, ''))
                    geo_with_terms_on_page = sum(1 for entity in unique_geo_entities_on_page if geographic_to_terms.get(entity, ''))
                    # NEW: Add chronological terms to page summary
                    chrono_with_terms_count = sum(1 for terms in chronological_to_terms.values() if terms)
                    
                    f.write(f"Page Summary:\n")
                    if len(unique_topics_on_page) > 0:
                        f.write(f"  - Topics with vocabulary terms: {topics_with_terms_on_page}/{len(unique_topics_on_page)}\n")
                    if len(unique_geo_entities_on_page) > 0:
                        f.write(f"  - Geographic entities with vocabulary terms: {geo_with_terms_on_page}/{len(unique_geo_entities_on_page)}\n")
                    # NEW: Add chronological terms to summary
                    f.write(f"  - Chronological terms available: {chrono_with_terms_count}/{len(chronological_to_terms)}\n")
                    
                    total_items_on_page = len(unique_topics_on_page) + len(unique_geo_entities_on_page)
                    total_with_terms_on_page = topics_with_terms_on_page + geo_with_terms_on_page
                    if total_items_on_page > 0:
                        f.write(f"  - Page-specific success: {total_with_terms_on_page}/{total_items_on_page} ({(total_with_terms_on_page/total_items_on_page*100):.1f}%)\n")
                    f.write(f"  - Collection chronological coverage: {chrono_with_terms_count} terms available\n")
                    
                    f.write("\n" + "-" * 40 + "\n\n")
                
                # Rest of the existing reference sections...
                # Summary by topic (alphabetical) at the end for reference
                f.write("ALPHABETICAL TOPIC REFERENCE:\n")
                f.write("-" * 40 + "\n\n")
                
                for subject, vocab_terms in sorted(subject_to_terms.items()):
                    # Find which pages this topic appears on
                    pages_with_topic = []
                    for page_num, topics_on_page in page_to_topics.items():
                        if subject in topics_on_page:
                            pages_with_topic.append(str(page_num))
                    
                    f.write(f"Topic: {subject}\n")
                    if pages_with_topic:
                        sorted_pages = sorted(pages_with_topic, key=lambda x: int(x) if x.isdigit() else float('inf'))
                        f.write(f"Appears on pages: {', '.join(sorted_pages)}\n")
                    else:
                        f.write("Appears on pages: None\n")
                    
                    if vocab_terms:
                        f.write(f"Terms: {vocab_terms}\n")
                    else:
                        f.write("Terms: No terms found\n")
                    f.write("\n")
                
                # Summary by geographic entity (alphabetical) at the end for reference
                f.write("ALPHABETICAL GEOGRAPHIC ENTITY REFERENCE:\n")
                f.write("-" * 45 + "\n\n")
                
                for entity, vocab_terms in sorted(geographic_to_terms.items()):
                    # Find which pages this geographic entity appears on
                    pages_with_entity = []
                    for page_num, entities_on_page in page_to_geographic.items():
                        if entity in entities_on_page:
                            pages_with_entity.append(str(page_num))
                    
                    f.write(f"Geographic Entity: {entity}\n")
                    if pages_with_entity:
                        sorted_pages = sorted(pages_with_entity, key=lambda x: int(x) if x.isdigit() else float('inf'))
                        f.write(f"Appears on pages: {', '.join(sorted_pages)}\n")
                    else:
                        f.write("Appears on pages: None\n")
                    
                    if vocab_terms:
                        f.write(f"Terms: {vocab_terms}\n")
                        # Add note about additional metadata for FAST Geographic terms
                        if '[FAST Geographic]' in vocab_terms:
                            f.write(f"Note: FAST Geographic terms include sources and links metadata in detailed API logs\n")
                    else:
                        f.write("Terms: No terms found\n")
                    f.write("\n")
                
                # NEW: Summary by chronological term (alphabetical) at the end for reference
                f.write("ALPHABETICAL CHRONOLOGICAL TERM REFERENCE:\n")
                f.write("-" * 50 + "\n\n")
                
                for chrono_term, vocab_terms in sorted(chronological_to_terms.items()):
                    f.write(f"Chronological Term: {chrono_term}\n")
                    f.write(f"Applies to: All pages (collection-wide coverage)\n")
                    
                    if vocab_terms:
                        f.write(f"LCSH Terms: {vocab_terms}\n")
                    else:
                        f.write("LCSH Terms: No terms found\n")
                    f.write("\n")
            
            return True
            
        except Exception as e:
            logging.error(f"Error creating vocabulary report: {e}")
            return False
        
    def run(self) -> bool:
        """Main execution method with comprehensive API logging for topics and geographic entities."""
        print(f"\nSTEP 2 - MULTI-VOCABULARY ENHANCEMENT")
        print(f"Processing folder: {self.folder_path}")
        print(f"Maximum {self.max_terms_per_vocabulary} terms per vocabulary")
        print(f"Maximum {self.max_total_terms} terms total per subject")
        print(f"API log: {self.logs_folder_path}")
        print("-" * 50)
        
        # Detect workflow type
        if not self.detect_workflow_type():
            return False
        
        print(f"Detected workflow type: {self.workflow_type.upper()}")
        
        # Load JSON data
        if not self.load_json_data():
            return False
        
        # Extract topics and geographic entities
        subjects, geographic_entities = self.extract_subject_headings()

        # NEW: Extract chronological terms from actual issue dates
        chronological_terms = self.extract_chronological_terms_from_issues()

        if not subjects and not geographic_entities and not chronological_terms:
            print("No topics, geographic entities, or chronological terms found in the data")
            return False

        print(f"Found {len(subjects)} unique topics")
        print(f"Found {len(chronological_terms)} chronological terms") 
        print(f"Found {len(geographic_entities)} unique geographic entities")
        
        # Process multi-vocabulary lookup with comprehensive logging
        (subject_to_terms_excel, subject_to_terms_json, 
        geographic_to_terms_excel, geographic_to_terms_json,
        chronological_to_terms_excel, chronological_to_terms_json) = self.process_multi_vocabulary_lookup(
            subjects, geographic_entities, chronological_terms
        )
        
        # Enhance Excel file
        if not self.enhance_excel_file(subject_to_terms_excel, geographic_to_terms_excel, chronological_to_terms_excel):
            return False

        # Enhance JSON file
        if not self.enhance_json_file(subject_to_terms_json, geographic_to_terms_json, chronological_to_terms_json):
            return False

        # Create vocabulary report
        self.create_vocabulary_report(subject_to_terms_excel, geographic_to_terms_excel, chronological_to_terms_excel)
        
        # Generate comprehensive API usage logs
        total_items = len(subjects) + len(geographic_entities)
        api_summary_stats = self.stats_tracker.get_summary_stats(total_items)
        create_vocab_api_usage_log(
            logs_folder_path=self.logs_folder_path,
            script_name="southern_architect_step2",
            total_topics=total_items,
            api_stats=api_summary_stats
        )
        
        # Final summary
        subjects_with_terms = sum(1 for terms in subject_to_terms_excel.values() if terms)
        geo_with_terms = sum(1 for terms in geographic_to_terms_excel.values() if terms)

        print(f"\n✅ STEP 2 COMPLETE: Enhanced with vocabulary terms in {os.path.basename(self.folder_path)}")
        print(f"Updated Excel/JSON files, vocabulary report, and API logs created")
        print(f"Topics with vocabulary terms: {subjects_with_terms}/{len(subjects)}")
        print(f"Geographic entities with vocabulary terms: {geo_with_terms}/{len(geographic_entities)}")
        if subjects:
            print(f"Topic success rate: {(subjects_with_terms/len(subjects)*100):.1f}%")
        if geographic_entities:
            print(f"Geographic success rate: {(geo_with_terms/len(geographic_entities)*100):.1f}%")
        print(f"Limited to {self.max_terms_per_vocabulary} terms per vocabulary")
        print(f"Maximum {self.max_total_terms} terms total per subject")
        print(f"Total API requests made: {api_summary_stats['total_requests']}")
        print(f"Processing time: {api_summary_stats['total_time']:.1f}s")
       
        
        # Show API breakdown
        print(f"\nAPI BREAKDOWN:")
        for api_name, stats in api_summary_stats['api_breakdown'].items():
            requests = stats['requests']
            success_rate = stats['success_rate']
            terms_found = stats['terms_found']
            cache_hits = stats['cache_hits']
            print(f"   {api_name}: {requests} requests, {success_rate:.1f}% success, {terms_found} terms, {cache_hits} cache hits")
        
        return True

def main():
    
    # Default base directory for Southern Architect output folders
    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_output_dir = os.path.join(script_dir, "output_folders")
    
    # Default folder path (newest folder if not specified)
    folder_path = find_newest_folder(base_output_dir)
    if not folder_path:
        print(f"No folders found in: {base_output_dir}")
        return 1
    print(f"Auto-selected newest folder: {os.path.basename(folder_path)}")

    # Create and run the enhancer with comprehensive API logging
    enhancer = SouthernArchitectEnhancer(folder_path)
    success = enhancer.run()
    
    if not success:
        print("Multi-vocabulary enhancement failed")
        return 1
    
    return 0

if __name__ == "__main__":
    exit(main())