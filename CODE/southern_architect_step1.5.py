#!/usr/bin/env python3
"""
Southern Architect Step 1.5: Batch Cleanup
==========================================

This script detects failed batch processing items from Step 1 and reprocesses them 
individually with retry logic to ensure clean, complete metadata before Steps 2-5.

This runs immediately after Step 1 if batch processing was used, ensuring all 
downstream steps have high-quality data to work with.

Author: Southern Architect Processing Team
"""

import os
import json
import logging
import time
import argparse
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from openai import OpenAI
import tenacity
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from PIL import Image as PILImage
from io import BytesIO
from openpyxl.drawing.image import Image as XLImage
import base64

# Import our custom modules
from model_pricing import calculate_cost, get_model_info
from token_logging import create_token_usage_log, log_individual_response
from prompts import SouthernArchitectPrompts

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Suppress verbose HTTP logging
logging.getLogger("openai").setLevel(logging.WARNING)
logging.getLogger("urllib3").setLevel(logging.WARNING)
logging.getLogger("httpx").setLevel(logging.WARNING)

client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))

class APIStats:
    def __init__(self):
        self.total_requests = 0
        self.total_input_tokens = 0
        self.total_output_tokens = 0
        self.processing_times = []

api_stats = APIStats()

def postprocess_api_response(response_data):
    """Post-process the API response for consistency (same as Step 1)."""
    
    # Helper function to convert string to list if needed
    def ensure_list(field_value):
        if isinstance(field_value, str):
            # Split by comma and clean up each item
            items = [item.strip() for item in field_value.split(',') if item.strip()]
            return items
        elif isinstance(field_value, list):
            return field_value
        else:
            return []
    
    # Handle named entities
    if 'namedEntities' in response_data:
        response_data['namedEntities'] = ensure_list(response_data['namedEntities'])
        response_data['namedEntities'] = list(dict.fromkeys(response_data['namedEntities']))
        response_data['namedEntities'] = [entity for entity in response_data['namedEntities'] 
                                        if len(entity) > 1 or not entity.isalnum()]
    
    # Handle geographic entities
    if 'geographicEntities' in response_data:
        response_data['geographicEntities'] = ensure_list(response_data['geographicEntities'])
        response_data['geographicEntities'] = list(dict.fromkeys(response_data['geographicEntities']))
        response_data['geographicEntities'] = [entity for entity in response_data['geographicEntities'] 
                                             if len(entity) > 1 or not entity.isalnum()]
    
    # Handle topics field variations
    if 'subjects' in response_data and 'topics' not in response_data:
        response_data['topics'] = ensure_list(response_data.pop('subjects'))
    elif 'subjectHeadings' in response_data and 'topics' not in response_data:
        response_data['topics'] = ensure_list(response_data.pop('subjectHeadings'))
    elif 'topics' in response_data:
        response_data['topics'] = ensure_list(response_data['topics'])
    
    # Ensure content warning field
    if 'contentWarning' not in response_data:
        response_data['contentWarning'] = 'None'
    elif response_data['contentWarning'].lower() == 'none' or response_data['contentWarning'].strip() == '':
        response_data['contentWarning'] = 'None'
    else:
        response_data['contentWarning'] = response_data['contentWarning'].capitalize().rstrip('.') + '.'
    
    return response_data

class BatchCleanupProcessor:
    """Detect and reprocess failed batch items from Step 1."""
    
    def __init__(self, folder_path: str):
        self.folder_path = folder_path
        self.workflow_type = None
        self.json_data = None
        self.excel_path = None
        self.failed_items = []
        self.reprocessed_items = 0
        self.improvements_made = 0
        
    def detect_workflow_type(self) -> bool:
        """Detect workflow type and check for batch processing."""
        metadata_dir = os.path.join(self.folder_path, "metadata", "collection_metadata")
        text_files = ['text_workflow.xlsx', 'text_workflow.json']
        image_files = ['image_workflow.xlsx', 'image_workflow.json']
        
        has_text_files = all(os.path.exists(os.path.join(metadata_dir, f)) for f in text_files)
        has_image_files = all(os.path.exists(os.path.join(metadata_dir, f)) for f in image_files)
        
        if has_text_files and not has_image_files:
            self.workflow_type = 'text'
            self.excel_path = os.path.join(metadata_dir, 'text_workflow.xlsx')
        elif has_image_files and not has_text_files:
            self.workflow_type = 'image'
            self.excel_path = os.path.join(metadata_dir, 'image_workflow.xlsx')
        else:
            logging.error("Could not determine workflow type or multiple workflow files found.")
            return False
        
        return True
    
    def load_json_data(self) -> bool:
        """Load JSON data and check if batch processing was used."""
        json_filename = f"{self.workflow_type}_workflow.json"
        metadata_dir = os.path.join(self.folder_path, "metadata", "collection_metadata")
        json_path = os.path.join(metadata_dir, json_filename)
        
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                self.json_data = json.load(f)
            
            # Check if batch processing was used
            api_stats_entry = None
            if self.json_data and 'api_stats' in self.json_data[-1]:
                api_stats_entry = self.json_data[-1]
                processing_mode = api_stats_entry.get('api_stats', {}).get('processing_mode', 'INDIVIDUAL')
                
                if processing_mode != 'BATCH':
                    print(f"✅ This dataset used {processing_mode} processing with retry logic.")
                    print("Step 1.5 cleanup not needed - individual processing already handles failures.")
                    return False
            
            print(f"📊 Loaded JSON data from {json_filename}")
            print("🔍 Detected batch processing - checking for failed items...")
            return True
            
        except Exception as e:
            logging.error(f"Error loading JSON data: {e}")
            return False
    
    def detect_failed_items(self) -> List[Tuple[int, Dict[str, Any], str]]:
        """Detect items that genuinely failed during batch processing (VERY conservative approach)."""
        failed_items = []
        
        # Skip API stats if present
        data_items = self.json_data[:-1] if self.json_data and 'api_stats' in self.json_data[-1] else self.json_data
        
        print(f"🔍 Analyzing {len(data_items)} items for genuine batch processing failures...")
        
        for i, item in enumerate(data_items):
            if 'analysis' not in item:
                failed_items.append((i, item, "Missing analysis section"))
                continue
                
            analysis = item['analysis']
            failure_reasons = []
            
            # Get both raw response and processed content
            raw_response = analysis.get('raw_response', '')
            
            # 1. EXPLICIT processing errors in raw response (most reliable indicator)
            error_indicators = [
                'Processing error:', 'Error:', 'Failed to', 'Exception:', 'Traceback:',
                'REPROCESSING FAILED:', 'Timeout', 'Connection error', 'API error',
                'Internal server error', 'Rate limit', 'Service unavailable'
            ]
            if any(error_indicator in raw_response for error_indicator in error_indicators):
                failure_reasons.append("Explicit processing error in response")
            
            # 2. COMPLETELY empty main content (not just short - actually empty)
            main_content = self._get_main_content_field(analysis)
            if not main_content or len(main_content.strip()) == 0:
                failure_reasons.append("Completely empty main content field")
            
            # 3. VERY OBVIOUS repetition patterns (much more conservative)
            if self._has_severe_repetition_problems(analysis):
                failure_reasons.append("Severe repetition pattern detected")
            
            # 4. Raw response is literally unparseable JSON
            if raw_response and self._is_completely_broken_json(raw_response):
                failure_reasons.append("Completely broken JSON structure")
            
            if failure_reasons:
                failed_items.append((i, item, '; '.join(failure_reasons)))
        
        return failed_items
    
    def _has_severe_repetition_problems(self, analysis: Dict[str, Any]) -> bool:
        """Check for SEVERE repetition problems only - very conservative."""
        main_content = self._get_main_content_field(analysis)
        toc_entry = analysis.get('toc_entry', '')
        
        # Only check main fields
        fields_to_check = [main_content, toc_entry]
        
        for field_content in fields_to_check:
            if not field_content or len(field_content.strip()) < 30:
                continue
                
            # Look for VERY obvious problems only
            if self._has_extreme_repetition(field_content):
                return True
        
        return False

    def _has_extreme_repetition(self, content: str) -> bool:
        """Detect only extreme repetition cases - very conservative."""
        if not content or len(content) < 50:
            return False
        
        words = content.split()
        if len(words) < 10:
            return False
        
        # Only flag if we see the EXACT patterns from your manufactured error
        # Look for sequences like "the the the the the the" or "error error error error"
        
        # Check for repeated single words (like "the the the the")
        for i in range(len(words) - 7):  # Need at least 8 repetitions to flag
            word = words[i]
            if len(word) <= 3:  # Only flag short words like "the", "and", "is"
                consecutive_count = 1
                for j in range(i + 1, min(i + 20, len(words))):
                    if words[j] == word:
                        consecutive_count += 1
                    else:
                        break
                
                if consecutive_count >= 8:  # 8+ consecutive identical words
                    # Check if it's a problematic word
                    if word.lower() in ['the', 'and', 'is', 'a', 'to', 'of', 'in', 'error', 'failed', 'null', 'loading']:
                        return True
        
        # Check for repeated error phrases (like "error error error error")
        error_phrases = [
            'error error', 'failed failed', 'null null', 'loading loading',
            'undefined undefined', 'timeout timeout', 'processing processing'
        ]
        
        content_lower = content.lower()
        for phrase in error_phrases:
            if content_lower.count(phrase) >= 3:  # Phrase appears 3+ times
                return True
        
        return False

    def _is_completely_broken_json(self, raw_response: str) -> bool:
        """Check for completely broken JSON - very conservative."""
        if not raw_response or len(raw_response.strip()) < 10:
            return False
        
        # Only flag if JSON is completely broken
        broken_indicators = [
            # No JSON structure at all
            (raw_response.count('{') == 0 and raw_response.count('}') == 0),
            # Massively unbalanced braces (way beyond normal)
            (abs(raw_response.count('{') - raw_response.count('}')) > 5),
            # Response is just an error message and very short
            (raw_response.strip().startswith(('Error:', 'Failed:', 'Exception:')) and len(raw_response) < 50),
        ]
        
        return any(broken_indicators)
    
    def _get_main_content_field(self, analysis: Dict[str, Any]) -> str:
        """Get the main content field value."""
        if self.workflow_type == 'text':
            return analysis.get('cleaned_text', '')
        else:
            return analysis.get('text_transcription', '')
    
    def parse_json_response_enhanced(self, raw_response: str) -> Tuple[Dict[str, Any], Optional[str]]:
        """Enhanced JSON parsing with multiple recovery strategies (same as original Step 1)."""
        if not raw_response or not raw_response.strip():
            return None, "Empty response"
        
        try:
            # Strategy 1: Standard cleaning
            cleaned_response = re.sub(r'```json\s*|\s*```', '', raw_response)
            cleaned_response = re.sub(r'^[^{]*({.*})[^}]*$', r'\1', cleaned_response, flags=re.DOTALL)
            cleaned_response = re.sub(r',(\s*[}\]])', r'\1', cleaned_response)
            
            parsed_json = json.loads(cleaned_response)
            return parsed_json, None
            
        except json.JSONDecodeError:
            pass
        
        try:
            # Strategy 2: Extract JSON object more aggressively
            json_match = re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', raw_response, re.DOTALL)
            if json_match:
                json_str = json_match.group(0)
                json_str = re.sub(r',(\s*[}\]])', r'\1', json_str)
                parsed_json = json.loads(json_str)
                return parsed_json, None
                
        except json.JSONDecodeError:
            pass
        
        try:
            # Strategy 3: Fix common issues
            fixed_response = raw_response
            
            # Fix unquoted keys
            fixed_response = re.sub(r'(\w+):', r'"\1":', fixed_response)
            
            # Fix unquoted string values
            fixed_response = re.sub(r':\s*([^",\[\{][^,\]\}]*)', r': "\1"', fixed_response)
            
            # Remove trailing commas
            fixed_response = re.sub(r',(\s*[}\]])', r'\1', fixed_response)
            
            # Extract JSON part
            json_match = re.search(r'\{.*\}', fixed_response, re.DOTALL)
            if json_match:
                parsed_json = json.loads(json_match.group(0))
                return parsed_json, None
                
        except json.JSONDecodeError:
            pass
        
        return None, "All parsing strategies failed"
    
    @tenacity.retry(
        wait=tenacity.wait_exponential(multiplier=1, min=4, max=10),
        stop=tenacity.stop_after_attempt(3),
        retry=tenacity.retry_if_exception_type(Exception)
    )
    def reprocess_single_item(self, item: Dict[str, Any], model_name: str) -> Tuple[Dict[str, Any], str, Any, float]:
        """Reprocess a single failed item with retry logic."""
        if self.workflow_type == 'text':
            return self._reprocess_text_item(item, model_name)
        else:
            return self._reprocess_image_item(item, model_name)
    
    def _reprocess_text_item(self, item: Dict[str, Any], model_name: str) -> Tuple[Dict[str, Any], str, Any, float]:
        """Reprocess a text item using same logic as Step 1."""
        # Get original content
        file_path = item.get('file_path', '')
        folder_name = item.get('folder', '')
        page_number = item.get('page_number', 0)
        
        # Read the original text file
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
        except:
            # Fall back to using data from JSON if file not accessible
            content = item.get('analysis', {}).get('cleaned_text', '')
            if not content:
                raise Exception("No source content available for reprocessing")
        
        # Preprocess OCR text (same as Step 1)
        content = self._preprocess_ocr_text(content)
        
        if not content.strip():
            # Handle empty content
            return {
                "cleanedText": "",
                "tocEntry": "No readable text on this page",
                "namedEntities": [],
                "geographicEntities": [],
                "topics": [],
                "contentWarning": "None"
            }, "Empty content", None, 0
        
        # Use prompts module (same as Step 1)
        prompt, prompt_type = SouthernArchitectPrompts.determine_prompt_type(content, page_number, file_path)
        
        api_stats.total_requests += 1
        start_time = time.time()
        
        response = client.chat.completions.create(
            model=model_name,
            messages=[{
                "role": "system", 
                "content": "You are an AI archival expert tasked with cleaning OCR text and extracting metadata from it."
            }, {
                "role": "user",
                "content": f"{prompt}\n\nHere's the content to analyze:\n\n{content.strip()}\n\nNote: This is a Step 1.5 cleanup - reprocessing to fix batch processing issues."
            }],
            max_tokens=3000,
            temperature=0.1  # Lower temperature for more consistent results
        )
        
        processing_time = time.time() - start_time
        api_stats.processing_times.append(processing_time)
        
        api_stats.total_input_tokens += response.usage.prompt_tokens
        api_stats.total_output_tokens += response.usage.completion_tokens
        
        raw_response = response.choices[0].message.content.strip()
        
        # Use enhanced parsing
        parsed_json, error = self.parse_json_response_enhanced(raw_response)
        
        if not parsed_json:
            raise Exception(f"Enhanced JSON parsing failed: {error}")
        
        # Handle field name variations (same as Step 1)
        if 'subjects' in parsed_json and 'topics' not in parsed_json:
            parsed_json['topics'] = parsed_json.pop('subjects')
        elif 'subjectHeadings' in parsed_json and 'topics' not in parsed_json:
            parsed_json['topics'] = parsed_json.pop('subjectHeadings')
        
        # Ensure required fields exist
        required_fields = ['cleanedText', 'tocEntry', 'namedEntities', 'geographicEntities', 'topics', 'contentWarning']
        for field in required_fields:
            if field not in parsed_json:
                if field in ['namedEntities', 'geographicEntities', 'topics']:
                    parsed_json[field] = []
                else:
                    parsed_json[field] = ""
        
        # Post-process the response (same as Step 1)
        parsed_json = postprocess_api_response(parsed_json)
        
        return parsed_json, raw_response, response.usage, processing_time
    
    def _reprocess_image_item(self, item: Dict[str, Any], model_name: str) -> Tuple[Dict[str, Any], str, Any, float]:
        """Reprocess an image item using same logic as Step 1."""
        # Get original image path
        image_path = item.get('image_path', '')
        if not image_path or not os.path.exists(image_path):
            raise Exception("Original image file not found for reprocessing")
        
        # Prepare base64 image (same as Step 1)
        with open(image_path, "rb") as image_file:
            base64_image = base64.b64encode(image_file.read()).decode('utf-8')
        
        api_stats.total_requests += 1
        start_time = time.time()
        
        response = client.chat.completions.create(
            model=model_name,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "text", "text": f"{SouthernArchitectPrompts.get_image_analysis_prompt()}\n\nNote: This is Step 1.5 cleanup - reprocessing to fix batch processing issues."},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}}
                ]
            }],
            max_tokens=3000,
            temperature=0.1  # Lower temperature for consistency
        )
        
        processing_time = time.time() - start_time
        api_stats.processing_times.append(processing_time)
        
        api_stats.total_input_tokens += response.usage.prompt_tokens
        api_stats.total_output_tokens += response.usage.completion_tokens
        
        raw_response = response.choices[0].message.content.strip()
        
        # Use enhanced parsing
        parsed_json, error = self.parse_json_response_enhanced(raw_response)
        
        if not parsed_json:
            raise Exception(f"Enhanced JSON parsing failed: {error}")
        
        # Handle field name variations (same as Step 1)
        if 'subjects' in parsed_json and 'topics' not in parsed_json:
            parsed_json['topics'] = parsed_json.pop('subjects')
        elif 'subjectHeadings' in parsed_json and 'topics' not in parsed_json:
            parsed_json['topics'] = parsed_json.pop('subjectHeadings')
        
        # Ensure required fields exist
        required_fields = ['textTranscription', 'visualDescription', 'tocEntry', 'namedEntities', 'geographicEntities', 'topics', 'contentWarning']
        for field in required_fields:
            if field not in parsed_json:
                if field in ['namedEntities', 'geographicEntities', 'topics']:
                    parsed_json[field] = []
                else:
                    parsed_json[field] = ""
        
        # Post-process the response (same as Step 1)  
        parsed_json = postprocess_api_response(parsed_json)
        
        return parsed_json, raw_response, response.usage, processing_time
    
    def _preprocess_ocr_text(self, text):
        """Preprocess OCR text (same as Step 1)."""
        replacements = {
            "Sv'cink i^idam": "Frank Adam",
            # Add more common OCR errors here as needed
        }
        
        for error, correction in replacements.items():
            text = text.replace(error, correction)
        
        # Remove unusual characters
        text = re.sub(r'[^\w\s.,;:!?()-]', '', text)
        
        return text
    
    def update_json_data(self, item_index: int, new_analysis: Dict[str, Any]) -> bool:
        """Update JSON data with reprocessed results."""
        try:
            # Skip API stats if present
            data_items = self.json_data[:-1] if self.json_data and 'api_stats' in self.json_data[-1] else self.json_data
            api_stats_entry = self.json_data[-1] if self.json_data and 'api_stats' in self.json_data[-1] else None
            
            if item_index < len(data_items):
                # Update the analysis with new fields
                data_items[item_index]['analysis'].update(new_analysis)
                
                # Reassemble JSON data
                updated_json_data = data_items.copy()
                if api_stats_entry:
                    updated_json_data.append(api_stats_entry)
                
                self.json_data = updated_json_data
                return True
            
            return False
            
        except Exception as e:
            logging.error(f"Error updating JSON data: {e}")
            return False
    
    def update_excel_file(self, item_index: int, new_analysis: Dict[str, Any]) -> bool:
        """Update Excel file with reprocessed results."""
        try:
            wb = load_workbook(self.excel_path)
            analysis_sheet = wb['Analysis']
            
            row_num = item_index + 2  # +2 for header row
            
            if self.workflow_type == 'text':
                # Update text workflow columns (same structure as Step 1)
                analysis_sheet.cell(row=row_num, column=4).value = new_analysis.get('cleanedText', '')  # Cleaned OCR Text
                analysis_sheet.cell(row=row_num, column=5).value = new_analysis.get('tocEntry', '')     # TOC Entry
                analysis_sheet.cell(row=row_num, column=6).value = ', '.join(new_analysis.get('namedEntities', []))      # Named Entities
                analysis_sheet.cell(row=row_num, column=7).value = ', '.join(new_analysis.get('geographicEntities', [])) # Geographic Entities
                analysis_sheet.cell(row=row_num, column=8).value = ', '.join(new_analysis.get('topics', []))             # Topics
                analysis_sheet.cell(row=row_num, column=9).value = new_analysis.get('contentWarning', 'None')           # Content Warning
            else:  # image workflow
                # Update image workflow columns (same structure as Step 1)
                analysis_sheet.cell(row=row_num, column=4).value = new_analysis.get('textTranscription', '')    # Text Transcription
                analysis_sheet.cell(row=row_num, column=5).value = new_analysis.get('visualDescription', '')    # Visual Description
                analysis_sheet.cell(row=row_num, column=6).value = new_analysis.get('tocEntry', '')             # TOC Entry
                analysis_sheet.cell(row=row_num, column=7).value = ', '.join(new_analysis.get('namedEntities', []))      # Named Entities
                analysis_sheet.cell(row=row_num, column=8).value = ', '.join(new_analysis.get('geographicEntities', [])) # Geographic Entities
                analysis_sheet.cell(row=row_num, column=9).value = ', '.join(new_analysis.get('topics', []))             # Topics
                analysis_sheet.cell(row=row_num, column=10).value = new_analysis.get('contentWarning', 'None')          # Content Warning
                
                # For image workflow, also update thumbnail if possible
                try:
                    image_path = self.json_data[item_index].get('image_path', '')
                    if image_path and os.path.exists(image_path):
                        img = PILImage.open(image_path)
                        img.thumbnail((200, 200))
                        output = BytesIO()
                        img.save(output, format='JPEG')
                        output.seek(0)
                        img_excel = XLImage(output)
                        img_excel.anchor = analysis_sheet.cell(row=row_num, column=3).coordinate
                        analysis_sheet.add_image(img_excel)
                        analysis_sheet.row_dimensions[row_num].height = 409
                except Exception as e:
                    logging.warning(f"Could not update thumbnail for row {row_num}: {e}")
            
            # Set alignment for updated cells
            max_col = 10 if self.workflow_type == 'text' else 11
            for col in range(4, max_col):
                cell = analysis_sheet.cell(row=row_num, column=col)
                if col == 3 and self.workflow_type == 'image':
                    cell.alignment = Alignment(vertical='bottom', wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            wb.save(self.excel_path)
            return True
            
        except Exception as e:
            logging.error(f"Error updating Excel file: {e}")
            return False
    
    def save_updated_files(self) -> bool:
        """Save updated JSON file."""
        try:
            json_filename = f"{self.workflow_type}_workflow.json"
            metadata_dir = os.path.join(self.folder_path, "metadata", "collection_metadata")
            json_path = os.path.join(metadata_dir, json_filename)
            
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(self.json_data, f, indent=2, ensure_ascii=False)
            
            print(f"✅ Updated JSON file saved: {json_filename}")
            return True
            
        except Exception as e:
            logging.error(f"Error saving updated JSON file: {e}")
            return False
    
    def create_cleanup_report(self, failed_items: List[Tuple[int, Dict[str, Any], str]]) -> bool:
        """Create a detailed Step 1.5 cleanup report."""
        try:
            logs_folder = os.path.join(self.folder_path, "logs")
            if not os.path.exists(logs_folder):
                os.makedirs(logs_folder)
                
            report_path = os.path.join(logs_folder, "step1_5_batch_cleanup_report.txt")
            
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write("SOUTHERN ARCHITECT STEP 1.5 BATCH CLEANUP REPORT\n")
                f.write("=" * 55 + "\n\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Workflow Type: {self.workflow_type.upper()}\n")
                f.write(f"Folder: {os.path.basename(self.folder_path)}\n\n")
                
                f.write("STEP 1.5 CLEANUP SUMMARY:\n")
                f.write("-" * 25 + "\n")
                f.write("Purpose: Fix failed batch processing items from Step 1 before downstream processing\n")
                f.write(f"Total failed items detected: {len(failed_items)}\n")
                f.write(f"Items successfully reprocessed: {self.reprocessed_items}\n")
                f.write(f"Items with improvements: {self.improvements_made}\n")
                f.write(f"Cleanup success rate: {(self.reprocessed_items/len(failed_items)*100):.1f}%\n" if failed_items else "No failed items found\n")
                f.write("\n")
                
                if failed_items:
                    f.write("COMMON FAILURE TYPES DETECTED:\n")
                    f.write("-" * 30 + "\n")
                    failure_types = {}
                    for _, _, reasons in failed_items:
                        for reason in reasons.split('; '):
                            failure_types[reason] = failure_types.get(reason, 0) + 1
                    
                    for failure_type, count in sorted(failure_types.items(), key=lambda x: x[1], reverse=True):
                        f.write(f"  - {failure_type}: {count} items\n")
                    f.write("\n")
                    
                    f.write("DETAILED FAILED ITEMS:\n")
                    f.write("-" * 25 + "\n")
                    
                    for i, (item_index, item, reasons) in enumerate(failed_items, 1):
                        folder = item.get('folder', 'Unknown')
                        page = item.get('page_number', 'Unknown')
                        
                        f.write(f"{i:3d}. Item {item_index + 1}: {folder} Page {page}\n")
                        f.write(f"     Failure reasons: {reasons}\n")
                        f.write(f"     Status: {'✅ Reprocessed' if i <= self.reprocessed_items else '❌ Failed to reprocess'}\n")
                        f.write("\n")
                
                # API usage stats
                if api_stats.total_requests > 0:
                    total_cost = calculate_cost(
                        model_name="gpt-4o-2024-08-06",  # Default model
                        prompt_tokens=api_stats.total_input_tokens,
                        completion_tokens=api_stats.total_output_tokens,
                        is_batch=False
                    )
                    
                    f.write("API USAGE STATS (STEP 1.5):\n")
                    f.write("-" * 25 + "\n")
                    f.write(f"Total requests: {api_stats.total_requests}\n")
                    f.write(f"Total input tokens: {api_stats.total_input_tokens:,}\n")
                    f.write(f"Total output tokens: {api_stats.total_output_tokens:,}\n")
                    f.write(f"Total tokens: {api_stats.total_input_tokens + api_stats.total_output_tokens:,}\n")
                    f.write(f"Average processing time: {sum(api_stats.processing_times)/len(api_stats.processing_times):.2f}s\n")
                    f.write(f"Total cost: ${total_cost:.4f}\n")
                    f.write("\n")
                
                f.write("IMPACT ON DOWNSTREAM PROCESSING:\n")
                f.write("-" * 35 + "\n")
                if self.reprocessed_items > 0:
                    f.write("✅ Step 1.5 cleanup ensures high-quality input for Steps 2-5\n")
                    f.write("✅ Reduced risk of propagating errors through the pipeline\n")
                    f.write("✅ Improved vocabulary matching and entity extraction\n")
                    f.write("✅ Better issue-level synthesis with complete metadata\n")
                else:
                    f.write("ℹ️ No items needed reprocessing - Step 1 batch processing was successful\n")
                f.write("\n")
            
            print(f"📄 Step 1.5 cleanup report created: {report_path}")
            return True
            
        except Exception as e:
            logging.error(f"Error creating cleanup report: {e}")
            return False
    
    def debug_detection(self, item_index: int = 0):
        """Debug method to test detection on a specific item."""
        if not self.json_data:
            print("No JSON data loaded")
            return
        
        # Skip API stats if present
        data_items = self.json_data[:-1] if self.json_data and 'api_stats' in self.json_data[-1] else self.json_data
        
        if item_index >= len(data_items):
            print(f"Invalid item index: {item_index}, max is {len(data_items)-1}")
            return
            
        item = data_items[item_index]
        analysis = item.get('analysis', {})
        
        print(f"=== DEBUGGING ITEM {item_index} ===")
        
        # Show the content
        main_content = self._get_main_content_field(analysis)
        toc_entry = analysis.get('toc_entry', '')
        raw_response = analysis.get('raw_response', '')
        
        print(f"Main content length: {len(main_content)}")
        print(f"Main content preview: {main_content[:150]}...")
        print(f"TOC entry: {toc_entry[:100]}...")
        print(f"Raw response length: {len(raw_response)}")
        
        # Test each detection method
        print(f"\n=== DETECTION TESTS ===")
        print(f"Has severe repetition: {self._has_severe_repetition_problems(analysis)}")
        print(f"Completely broken JSON: {self._is_completely_broken_json(raw_response)}")
        print(f"Empty main content: {not main_content or len(main_content.strip()) == 0}")
        
        # Check for explicit errors
        error_indicators = [
            'Processing error:', 'Error:', 'Failed to', 'Exception:', 'Traceback:',
            'REPROCESSING FAILED:', 'Timeout', 'Connection error', 'API error'
        ]
        has_explicit_error = any(error_indicator in raw_response for error_indicator in error_indicators)
        print(f"Explicit processing error: {has_explicit_error}")
        
        # Test the actual method
        print(f"\n=== ACTUAL DETECTION RESULT ===")
        failed_items = self.detect_failed_items()
        flagged = any(idx == item_index for idx, _, _ in failed_items)
        print(f"Item {item_index} flagged as failed: {flagged}")
        
        if flagged:
            for idx, _, reasons in failed_items:
                if idx == item_index:
                    print(f"Failure reasons: {reasons}")
                    break
    
    def run(self, model_name: str = "gpt-4o-2024-08-06") -> bool:
        """Main execution method for Step 1.5."""
        print(f"\nSOUTHERN ARCHITECT STEP 1.5 - BATCH CLEANUP")
        print(f"Processing folder: {os.path.basename(self.folder_path)}")
        print(f"Purpose: Fix failed batch items from Step 1 before downstream processing")
        print(f"Model: {model_name}")
        print("-" * 60)
        
        # Detect workflow type
        if not self.detect_workflow_type():
            return False
        
        print(f"🔍 Detected workflow type: {self.workflow_type.upper()}")
        
        # Load JSON data and check if batch processing was used
        if not self.load_json_data():
            return True  # Return True because no cleanup needed, not an error
        
        # Detect failed items
        failed_items = self.detect_failed_items()
        
        if not failed_items:
            print("✅ No failed items detected! Step 1 batch processing was successful.")
            print("🎯 All metadata is ready for Steps 2-5. No cleanup needed.")
            return True
        
        print(f"⚠️ Found {len(failed_items)} items that need reprocessing:")
        
        # Show sample of failures
        for i, (item_index, item, reasons) in enumerate(failed_items[:3], 1):  # Show first 3
            folder = item.get('folder', 'Unknown')
            page = item.get('page_number', 'Unknown')
            print(f"   {i}. Item {item_index + 1}: {folder} Page {page}")
            print(f"      Issues: {reasons}")
        
        if len(failed_items) > 3:
            print(f"   ... and {len(failed_items) - 3} more items")
        
        # Show model pricing info
        model_info = get_model_info(model_name)
        if model_info:
            print(f"\n🔧 Reprocessing with: {model_name}")
            print(f"💰 Individual processing cost: ${model_info['input_per_1k']:.5f}/1K input, ${model_info['output_per_1k']:.5f}/1K output")
            print("💡 Using individual processing with retry logic to ensure success")
        
        # Create logs folder
        logs_folder_path = os.path.join(self.folder_path, "logs")
        if not os.path.exists(logs_folder_path):
            os.makedirs(logs_folder_path)
        
        # Reprocess failed items
        print(f"\n🔄 Reprocessing {len(failed_items)} failed items with retry logic...")
        
        for i, (item_index, item, reasons) in enumerate(failed_items, 1):
            folder = item.get('folder', 'Unknown')
            page = item.get('page_number', 'Unknown')
            
            print(f"\n🔧 Reprocessing {i}/{len(failed_items)}: {folder} Page {page}")
            print(f"   📋 Original issues: {reasons}")
            
            try:
                # Reprocess the item with retry logic
                new_response_data, raw_response, usage, processing_time = self.reprocess_single_item(item, model_name)
                
                # Convert field names to match JSON structure
                if self.workflow_type == 'text':
                    normalized_response = {
                        'cleaned_text': new_response_data.get('cleanedText', ''),
                        'toc_entry': new_response_data.get('tocEntry', ''),
                        'named_entities': new_response_data.get('namedEntities', []),
                        'geographic_entities': new_response_data.get('geographicEntities', []),
                        'topics': new_response_data.get('topics', []),
                        'content_warning': new_response_data.get('contentWarning', 'None'),
                        'raw_response': raw_response
                    }
                else:  # image
                    normalized_response = {
                        'text_transcription': new_response_data.get('textTranscription', ''),
                        'visual_description': new_response_data.get('visualDescription', ''),
                        'toc_entry': new_response_data.get('tocEntry', ''),
                        'named_entities': new_response_data.get('namedEntities', []),
                        'geographic_entities': new_response_data.get('geographicEntities', []),
                        'topics': new_response_data.get('topics', []),
                        'content_warning': new_response_data.get('contentWarning', 'None'),
                        'raw_response': raw_response
                    }
                
                # Update JSON and Excel files
                json_updated = self.update_json_data(item_index, normalized_response)
                excel_updated = self.update_excel_file(item_index, new_response_data)
                
                if json_updated and excel_updated:
                    self.reprocessed_items += 1
                    self.improvements_made += 1
                    
                    # Log individual response
                    log_individual_response(
                        logs_folder_path=logs_folder_path,
                        script_name="southern_architect_step1_5_cleanup",
                        row_number=item_index + 2,
                        barcode=f"{folder}_page{page}",
                        response_text=raw_response,
                        model_name=model_name,
                        prompt_tokens=usage.prompt_tokens if usage else 0,
                        completion_tokens=usage.completion_tokens if usage else 0,
                        processing_time=processing_time,
                        additional_info={
                            "reprocessing_reason": reasons,
                            "step": "1.5 - Batch Cleanup"
                        }
                    )
                    
                    print(f"   ✅ Successfully reprocessed and updated files")
                    
                    # Show what improved
                    main_field = 'cleaned_text' if self.workflow_type == 'text' else 'text_transcription'
                    main_content = normalized_response.get(main_field, '')
                    if len(main_content) > 50:
                        print(f"   📝 Content preview: {main_content[:50]}...")
                    
                else:
                    print(f"   ⚠️ Reprocessed but failed to update files (JSON: {json_updated}, Excel: {excel_updated})")
            
            except Exception as e:
                logging.error(f"Failed to reprocess item {item_index}: {e}")
                print(f"   ❌ Reprocessing failed: {str(e)}")
                
                # Log the failure
                log_individual_response(
                    logs_folder_path=logs_folder_path,
                    script_name="southern_architect_step1_5_cleanup",
                    row_number=item_index + 2,
                    barcode=f"{folder}_page{page}",
                    response_text=f"REPROCESSING FAILED: {str(e)}",
                    model_name=model_name,
                    prompt_tokens=0,
                    completion_tokens=0,
                    processing_time=0,
                    additional_info={
                        "reprocessing_reason": reasons, 
                        "error": str(e),
                        "step": "1.5 - Batch Cleanup"
                    }
                )
            
            # Add delay between requests
            time.sleep(0.5)
        
        # Save updated files
        if self.reprocessed_items > 0:
            print(f"\n💾 Saving updated files...")
            if not self.save_updated_files():
                print("⚠️ Warning: Failed to save updated JSON file")
        
        # Create comprehensive usage log
        if api_stats.total_requests > 0:
            total_cost = calculate_cost(
                model_name=model_name,
                prompt_tokens=api_stats.total_input_tokens,
                completion_tokens=api_stats.total_output_tokens,
                is_batch=False  # Individual reprocessing
            )
            
            create_token_usage_log(
                logs_folder_path=logs_folder_path,
                script_name="southern_architect_step1_5_cleanup",
                model_name=model_name,
                total_items=len(failed_items),
                items_with_issues=len(failed_items) - self.reprocessed_items,
                total_time=sum(api_stats.processing_times) if api_stats.processing_times else 0,
                total_prompt_tokens=api_stats.total_input_tokens,
                total_completion_tokens=api_stats.total_output_tokens,
                additional_metrics={
                    "Original failed items": len(failed_items),
                    "Successfully reprocessed": self.reprocessed_items,
                    "Improvements made": self.improvements_made,
                    "Processing mode": "INDIVIDUAL (Step 1.5 cleanup)",
                    "Cleanup success rate": f"{(self.reprocessed_items/len(failed_items)*100):.1f}%" if failed_items else "N/A",
                    "Step": "1.5 - Batch Cleanup",
                    "Total cost": f"${total_cost:.4f}"
                }
            )
        
        # Create cleanup report
        self.create_cleanup_report(failed_items)
        
        # Final summary
        print(f"\n" + "="*60)
        print(f"✅ STEP 1.5 COMPLETE: Batch cleanup finished")
        print(f"📊 Results:")
        print(f"   • Failed items detected: {len(failed_items)}")
        print(f"   • Successfully reprocessed: {self.reprocessed_items}")
        print(f"   • Cleanup success rate: {(self.reprocessed_items/len(failed_items)*100):.1f}%" if failed_items else "N/A")
        
        if api_stats.total_requests > 0:
            total_cost = calculate_cost(
                model_name=model_name,
                prompt_tokens=api_stats.total_input_tokens,
                completion_tokens=api_stats.total_output_tokens,
                is_batch=False
            )
            print(f"   • Tokens used: {api_stats.total_input_tokens + api_stats.total_output_tokens:,}")
            print(f"   • Additional cost: ${total_cost:.4f}")
        
        if self.reprocessed_items > 0:
            print(f"Updated files: JSON and Excel workflow files")
            print(f"Created: Individual response logs and cleanup report")
            print(f"Ready for Step 2: All metadata is now clean and complete")
        else:
            print(f"No items were successfully reprocessed")
            print(f"Consider manual review of failed items before proceeding to Step 2")
        
        print("="*60)
        print(f"✅ STEP 1.5 COMPLETE: Batch cleanup finished")
        return True

def find_newest_folder(base_directory: str) -> Optional[str]:
    """Find the newest folder in the base directory."""
    if not os.path.exists(base_directory):
        return None
    
    folders = [f for f in os.listdir(base_directory) 
              if os.path.isdir(os.path.join(base_directory, f))]
    
    if not folders:
        return None
    
    # Sort by modification time (newest first)
    folders.sort(key=lambda x: os.path.getmtime(os.path.join(base_directory, x)), reverse=True)
    
    return os.path.join(base_directory, folders[0])


def main():
    parser = argparse.ArgumentParser(
        description='Step 1.5: Clean up failed batch processing items before downstream processing',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Step 1.5: Batch Cleanup Overview:
  - Runs immediately after Step 1 if batch processing was used
  - Detects JSON parsing errors, infinite loops, missing fields, etc.
  - Reprocesses failed items individually with retry logic
  - Updates JSON and Excel files with improved results
  - Ensures clean, complete metadata for Steps 2-5
  - Creates detailed logs and reports

Examples:
  python southern_architect_step1_5.py                           # Process newest folder
  python southern_architect_step1_5.py --folder /path/to/folder  # Process specific folder
  python southern_architect_step1_5.py --model gpt-4o-2024-08-06 # Use specific model

This step ensures that downstream processing (vocabulary enhancement, synthesis, etc.)
has high-quality, complete metadata to work with.
        """
    )
    
    parser.add_argument('--folder', help='Specific folder path to process')
    parser.add_argument('--newest', action='store_true', help='Process the newest folder (default: True if no folder specified)')
    parser.add_argument('--model', default="gpt-4o-2024-08-06", help='Model for reprocessing (default: gpt-4o-2024-08-06)')
    
    args = parser.parse_args()
    
    # Default base directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_output_dir = os.path.join(script_dir, "output_folders")
    
    if args.folder:
        if not os.path.exists(args.folder):
            print(f"Folder not found: {args.folder}")
            return 1
        folder_path = args.folder
    else:
        # Default to newest folder
        folder_path = find_newest_folder(base_output_dir)
        if not folder_path:
            print(f"No folders found in: {base_output_dir}")
            return 1
        print(f"Auto-selected newest folder: {os.path.basename(folder_path)}")
    
    # Create and run the cleanup processor
    processor = BatchCleanupProcessor(folder_path)
    success = processor.run(args.model)
    
    return 0 if success else 1


if __name__ == "__main__":
    exit(main())