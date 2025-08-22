# terms selection - using OpenAI's GPT-4o-mini model

import os
import json
import logging
import time
from datetime import datetime
from typing import List, Dict, Any, Tuple
from openai import OpenAI
import tenacity
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from prompts import SouthernArchitectPrompts
from shared_utilities import APIStats, find_newest_folder

# Import our custom modules
from model_pricing import calculate_cost, get_model_info
from token_logging import create_token_usage_log, log_individual_response
from batch_processor import BatchProcessor

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Suppress verbose HTTP logging
logging.getLogger("openai").setLevel(logging.WARNING)
logging.getLogger("urllib3").setLevel(logging.WARNING)
logging.getLogger("httpx").setLevel(logging.WARNING)

client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))
DEFAULT_MODEL = "gpt-4o-mini-2024-07-18"  # Default model name, change as needed

api_stats = APIStats()

# Add debugging to prepare_batch_requests function in southern_architect_step3.py:

def prepare_batch_requests(entries_with_vocab, vocabulary_selector, model_name):
    """Prepare all vocabulary selection requests for batch processing."""
    batch_requests = []
    custom_id_mapping = {}
        
    for i, (entry_index, entry_data) in enumerate(entries_with_vocab):
        user_prompt = vocabulary_selector.create_user_prompt(entry_data)
        
        if not user_prompt:
            continue  # Skip entries without vocabulary terms
        
        # Create request data
        request_data = {
            "model": model_name,
            "messages": [
                {"role": "system", "content": vocabulary_selector.system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            "max_tokens": 1500,
            "temperature": 0.1
        }
        
        batch_requests.append(request_data)
        custom_id_mapping[f"vocab_selection_{i}"] = {
            "entry_index": entry_index,
            "entry_data": entry_data,
            "row_number": entry_index + 2  # +2 for header row
        }
    
    return batch_requests, custom_id_mapping

class VocabularySelector:
    """Class to select the best vocabulary terms for each page using an LLM."""

    def __init__(self, model_name: str = DEFAULT_MODEL):
        self.model_name = model_name
        self.system_prompt = SouthernArchitectPrompts.get_vocabulary_selection_system_prompt()

    def create_system_prompt(self) -> str:
        """Create the system prompt for vocabulary selection."""
        return SouthernArchitectPrompts.get_vocabulary_selection_system_prompt()

    def create_user_prompt(self, entry_data: Dict[str, Any]) -> str:
        """Create the user prompt for a specific entry with topics organized format."""
        analysis = entry_data.get('analysis', {})
        
        # Build content description
        content_parts = []
        
        # Add main content
        if analysis.get('folder'):
            content_parts.append(f"ISSUE:\n{analysis['folder']}")
        
        # Add visual description for images
        if analysis.get('visual_description'):
            content_parts.append(f"VISUAL DESCRIPTION:\n{analysis['visual_description']}")
        
        # Add TOC entry
        if analysis.get('toc_entry'):
            content_parts.append(f"SUMMARY:\n{analysis['toc_entry']}")
        
        # Add original topics (NOT geographic entities)
        if analysis.get('topics'):
            topics = analysis['topics']
            if isinstance(topics, list):
                content_parts.append(f"TOPICS:\n{', '.join(topics)}")
            else:
                content_parts.append(f"TOPICS:\n{topics}")
        
        content_description = "\n\n".join(content_parts)
        
        # Build topic-organized vocabulary terms
        topic_to_terms = analysis.get('vocabulary_search_results', {})
        
        if not topic_to_terms:
            return None  # No topic-organized vocabulary terms available
        
        topics_section = self._build_topic_organized_terms(topic_to_terms)
        
        # Combine everything
        user_prompt = f"""Analyze this page content and select appropriate vocabulary terms:

{content_description}

AVAILABLE VOCABULARY TERMS BY TOPIC:
{topics_section}

Select the most relevant terms following your instructions. Use exact labels without [source] brackets. Skip topics with no genuinely relevant terms.
"""
        
        return user_prompt

    def _build_topic_organized_terms(self, topic_to_terms: Dict[str, List[Dict]]) -> str:
        """Build the topic-organized terms section from topic_to_terms mapping."""
        sections = []
        
        for topic, terms in topic_to_terms.items():
            if terms:  # Only show topics that have terms
                sections.append(f"  Topic: {topic}")
                
                term_strings = []
                for term in terms:
                    if isinstance(term, dict):
                        label = term.get('label', '').strip()
                        source = term.get('source', 'Unknown')
                        uri = term.get('uri', '')
                        if uri:
                            term_strings.append(f"{label} ({uri}) [{source}]")
                        else:
                            term_strings.append(f"{label} [{source}]")
                
                sections.append(f"  Terms: {'; '.join(term_strings)}")
                sections.append("")  # Empty line between topics
        
        return "\n".join(sections)

    @tenacity.retry(
        wait=tenacity.wait_exponential(multiplier=1, min=4, max=10),
        stop=tenacity.stop_after_attempt(5),
        retry=tenacity.retry_if_exception_type(Exception)
    )
    def select_vocabulary_terms(self, entry_data: Dict[str, Any]) -> Tuple[Dict[str, Any], str, Any, float]:
        """Select vocabulary terms for a single entry."""
        user_prompt = self.create_user_prompt(entry_data)
        
        if not user_prompt:
            # No vocabulary terms available
            return {
                "selected_terms": []
            }, "No vocabulary terms available for selection", None, 0
        
        api_stats.total_requests += 1
        start_time = time.time()
        
        response = client.chat.completions.create(
            model=self.model_name,
            messages=[
                {"role": "system", "content": self.system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            max_tokens=1500,
            temperature=0.1  # Low temperature for consistent, precise selections
        )
        
        processing_time = time.time() - start_time
        api_stats.processing_times.append(processing_time)
        
        api_stats.total_input_tokens += response.usage.prompt_tokens
        api_stats.total_output_tokens += response.usage.completion_tokens
        
        raw_response = response.choices[0].message.content.strip()
        
        # Parse JSON response
        try:
            parsed_response = self.parse_json_response(raw_response)
            return parsed_response, raw_response, response.usage, processing_time
        except Exception as e:
            logging.error(f"Error parsing vocabulary selection response: {e}")
            # Return empty selection on parsing error
            return {
                "selected_terms": []
            }, raw_response, response.usage, processing_time

    def parse_json_response(self, raw_response: str) -> Dict[str, Any]:
        """Parse JSON response from the API."""
        from shared_utilities import parse_json_response_enhanced
        
        parsed_json, error = parse_json_response_enhanced(raw_response)
        
        if parsed_json is None:
            raise ValueError(f"Could not parse JSON response: {error}")
        
        # Validate structure
        if 'selected_terms' not in parsed_json:
            parsed_json['selected_terms'] = []
        
        return parsed_json

class SouthernArchitectVocabularyProcessor:
    """Main class for vocabulary selection and clean output generation."""

    def __init__(self, folder_path: str, model_name: str = DEFAULT_MODEL):
        self.folder_path = folder_path
        self.model_name = model_name
        self.workflow_type = None
        self.json_data = None
        self.excel_path = None
        self.vocabulary_selector = VocabularySelector(model_name)
        self.was_batch_processed = False 
    
    def detect_workflow_type(self) -> bool:
        """Detect workflow type and check for vocabulary enhancement."""
        # Check for expected files in the metadata/collection_metadata subfolder
        metadata_dir = os.path.join(self.folder_path, "metadata", "collection_metadata")
        text_files = ['text_workflow.xlsx', 'text_workflow.json']
        image_files = ['image_workflow.xlsx', 'image_workflow.json']
        
        has_text_files = all(os.path.exists(os.path.join(metadata_dir, f)) for f in text_files)
        has_image_files = all(os.path.exists(os.path.join(metadata_dir, f)) for f in image_files)
        
        if has_text_files and not has_image_files:
            self.workflow_type = 'text'
            self.excel_path = os.path.join(metadata_dir, 'text_workflow.xlsx')
        elif has_image_files and not has_text_files:
            self.workflow_type = 'image'
            self.excel_path = os.path.join(metadata_dir, 'image_workflow.xlsx')
        else:
            logging.error("Could not determine workflow type or multiple workflow files found.")
            return False
        
        # Check if vocabulary enhancement has been run (step 2)
        vocab_report_path = os.path.join(metadata_dir, 'vocabulary_mapping_report.txt')
        if not os.path.exists(vocab_report_path):
            logging.error("Vocabulary enhancement (step 2) must be run before step 3.")
            return False

        return True
    
    def load_json_data(self) -> bool:
        """Load JSON data and verify vocabulary terms exist."""
        json_filename = f"{self.workflow_type}_workflow.json"
        metadata_dir = os.path.join(self.folder_path, "metadata", "collection_metadata")
        json_path = os.path.join(metadata_dir, json_filename)
        
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                self.json_data = json.load(f)
            
            # Check if vocabulary terms exist in the data
            data_items = self.json_data[:-1] if self.json_data and 'api_stats' in self.json_data[-1] else self.json_data
            
            has_vocab_terms = False
            for item in data_items:
                if 'analysis' in item and 'vocabulary_search_results' in item['analysis']:
                    vocab_terms = item['analysis']['vocabulary_search_results']
                    if vocab_terms:  # Check if there are any vocabulary terms
                        has_vocab_terms = True
                        break
            
            if not has_vocab_terms:
                logging.error("No vocabulary terms found. Please run step 2 first.")
                return False
            
            print(f"Loaded JSON data from {json_filename}")
            return True
            
        except Exception as e:
            logging.error(f"Error loading JSON data: {e}")
            return False
    
    def find_entries_with_vocabulary(self) -> List[Tuple[int, Dict[str, Any]]]:
        """Find entries that have vocabulary terms available for selection."""
        entries_with_vocab = []
        
        # Skip the last item if it's API stats
        data_items = self.json_data[:-1] if self.json_data and 'api_stats' in self.json_data[-1] else self.json_data
        
        for i, item in enumerate(data_items):
            if 'analysis' in item and 'vocabulary_search_results' in item['analysis']:
                vocab_terms = item['analysis']['vocabulary_search_results']
                if vocab_terms:  # Only include items with vocabulary terms
                    entries_with_vocab.append((i, item))
        
        return entries_with_vocab
    
    def process_vocabulary_selection(self, entries_with_vocab: List[Tuple[int, Dict[str, Any]]]) -> Dict[int, Dict[str, Any]]:
        """Process vocabulary selection using batch processing when appropriate."""
        selection_results = {}
        
        # Create logs folder
        logs_folder_path = os.path.join(self.folder_path, "logs")
        if not os.path.exists(logs_folder_path):
            os.makedirs(logs_folder_path)
        
        total_entries = len(entries_with_vocab)
        
        # Initialize batch processor and check if we should use batch processing
        processor = BatchProcessor()
        use_batch = processor.should_use_batch(total_entries)
        
        print(f"Processing mode: {'BATCH' if use_batch else 'INDIVIDUAL'}")
        
        if use_batch:
            print(f"Preparing {total_entries} requests for batch processing...")
            
            # Prepare batch requests
            batch_requests, custom_id_mapping = prepare_batch_requests(
                entries_with_vocab, self.vocabulary_selector, self.model_name
            )
            
            if not batch_requests:
                print("No valid requests to process")
                return selection_results
            
            # Estimate costs
            cost_estimate = processor.estimate_batch_cost(batch_requests, self.model_name)
            print(f"Cost estimate: ${cost_estimate['batch_cost']:.4f} (${cost_estimate['savings']:.4f} savings)")
            
            # Convert to batch format
            formatted_requests = processor.create_batch_requests(batch_requests, "vocab_selection")
            
            # Submit batch
            batch_id = processor.submit_batch(
                formatted_requests, 
                f"Southern Architect Vocabulary Selection - {len(batch_requests)} entries - {datetime.now().strftime('%Y-%m-%d')}"
            )
            
            # Wait for completion
            batch_results = processor.wait_for_completion(batch_id, max_wait_hours=24, check_interval_minutes=5)
            
            if batch_results:
                # Process batch results
                processed_results = processor.process_batch_results(batch_results, custom_id_mapping)
                
                print(f"Processing batch results...")
                
                # Track tokens for logging
                api_stats.total_input_tokens = processed_results["summary"]["total_prompt_tokens"]
                api_stats.total_output_tokens = processed_results["summary"]["total_completion_tokens"]
                
                self.was_batch_processed = True
                
                # Process results
                for custom_id, result_data in processed_results["results"].items():
                    if custom_id.startswith("vocab_selection_"):
                        # Extract the index from custom_id
                        parts = custom_id.split("_")
                        if len(parts) >= 3:
                            try:
                                index = int(parts[2])
                                mapping_key = f"vocab_selection_{index}"
                                
                                if mapping_key in custom_id_mapping:
                                    entry_index = custom_id_mapping[mapping_key]["entry_index"]
                                    entry_data = custom_id_mapping[mapping_key]["entry_data"]
                                    row_number = custom_id_mapping[mapping_key]["row_number"]
                                    
                                    if result_data["success"]:
                                        raw_response = result_data["content"]
                                        usage = result_data["usage"]
                                        
                                        # Parse the vocabulary selection response
                                        try:
                                            selection_result = self.vocabulary_selector.parse_json_response(raw_response)
                                        except Exception as e:
                                            logging.error(f"Error parsing vocabulary selection response: {e}")
                                            selection_result = {"selected_terms": []}
                                        
                                        # Log individual response
                                        log_individual_response(
                                            logs_folder_path=logs_folder_path,
                                            script_name="southern_architect_vocabulary_selection",
                                            row_number=row_number,
                                            barcode=f"{entry_data.get('folder', 'unknown')}_page{entry_data.get('page_number', 'unknown')}",
                                            response_text=raw_response,
                                            model_name=self.model_name,
                                            prompt_tokens=usage.get("prompt_tokens", 0),
                                            completion_tokens=usage.get("completion_tokens", 0),
                                            processing_time=0  # Batch processing doesn't track individual timing
                                        )
                                        
                                        selection_results[entry_index] = {
                                            'selection_result': selection_result,
                                            'raw_response': raw_response,
                                            'processing_time': 0
                                        }
                                        
                                        selected_count = len(selection_result.get('selected_terms', []))
                                        print(f"   Entry {entry_index}: Selected {selected_count} vocabulary terms")
                                        
                                    else:
                                        # Handle error case
                                        error_msg = result_data['error']
                                        raw_response = f"Error: {error_msg}"
                                        
                                        # Log error
                                        log_individual_response(
                                            logs_folder_path=logs_folder_path,
                                            script_name="southern_architect_vocabulary_selection",
                                            row_number=row_number,
                                            barcode=f"{entry_data.get('folder', 'unknown')}_page{entry_data.get('page_number', 'unknown')}",
                                            response_text=raw_response,
                                            model_name=self.model_name,
                                            prompt_tokens=0,
                                            completion_tokens=0,
                                            processing_time=0
                                        )
                                        
                                        selection_results[entry_index] = {
                                            'selection_result': {'selected_terms': []},
                                            'raw_response': raw_response,
                                            'processing_time': 0
                                        }
                                        
                                        print(f"   Entry {entry_index}: Processing failed: {error_msg}")
                                        
                            except (ValueError, IndexError) as e:
                                logging.error(f"Error processing custom_id {custom_id}: {e}")
                                continue
                
                processed_entries = len(selection_results)
                print(f"\nBatch processing completed: {processed_entries}/{total_entries} entries processed")
                return selection_results
        
        # Fall back to individual processing (existing code)
        print(f"Using individual processing:")
        self.was_batch_processed = False
        return self.process_vocabulary_selection_individual(entries_with_vocab, logs_folder_path)

    def process_vocabulary_selection_individual(self, entries_with_vocab: List[Tuple[int, Dict[str, Any]]], logs_folder_path: str) -> Dict[int, Dict[str, Any]]:
        """Process vocabulary selection using individual API calls."""
        self.was_batch_processed = False
        selection_results = {}
        total_entries = len(entries_with_vocab)
        processed_entries = 0
        
        for i, (entry_index, entry_data) in enumerate(entries_with_vocab):
            print(f"\nProcessing entry {i+1}/{total_entries}")
            print(f"   Entry index: {entry_index}")
            print(f"   Progress: {((i+1)/total_entries)*100:.1f}%")
            
            try:
                selection_result, raw_response, usage, processing_time = self.vocabulary_selector.select_vocabulary_terms(entry_data)
                
                # Log individual response
                log_individual_response(
                    logs_folder_path=logs_folder_path,
                    script_name="southern_architect_vocabulary_selection",
                    row_number=entry_index + 2,  # +2 for header row
                    barcode=f"{entry_data.get('folder', 'unknown')}_page{entry_data.get('page_number', 'unknown')}",
                    response_text=raw_response,
                    model_name=self.model_name,
                    prompt_tokens=usage.prompt_tokens if usage else 0,
                    completion_tokens=usage.completion_tokens if usage else 0,
                    processing_time=processing_time
                )
                
                selection_results[entry_index] = {
                    'selection_result': selection_result,
                    'raw_response': raw_response,
                    'processing_time': processing_time
                }
                
                selected_count = len(selection_result.get('selected_terms', []))
                print(f"   Selected {selected_count} vocabulary terms")
                processed_entries += 1
                
            except Exception as e:
                logging.error(f"Error processing entry {entry_index}: {e}")
                selection_results[entry_index] = {
                    'selection_result': {'selected_terms': []},
                    'raw_response': f"Error: {str(e)}",
                    'processing_time': 0
                }
                print(f"   Processing failed: {str(e)}")
            
            # Add delay between requests
            time.sleep(0.5)
        
        print(f"\nIndividual processing completed: {processed_entries}/{total_entries} entries processed")
        return selection_results
    
    def normalize_label_for_matching(self, label: str) -> str:
        """Normalize labels for better matching."""
        import re
        
        # Convert to lowercase
        normalized = label.lower().strip()
        
        # Remove extra whitespace
        normalized = re.sub(r'\s+', ' ', normalized)
        
        # Remove common punctuation that might cause mismatches
        normalized = re.sub(r'[.,;:!?]$', '', normalized)
        
        return normalized

    def deduplicate_vocabulary_terms(self, matched_terms: List[Dict]) -> List[Dict]:
        """
        Deduplicate vocabulary terms when they have the same words (same order or different order).
        Priority order: Getty AAT > LCSH > FAST > others
        """
        if not matched_terms:
            return matched_terms
        
        # Define source priority (lower number = higher priority)
        source_priority = {
            'Getty AAT': 1,
            'LCSH': 2, 
            'FAST': 3
        }
        
        def normalize_for_comparison(label: str) -> str:
            """Convert to lowercase letters only, preserving word order."""
            import re
            
            # Convert to lowercase and keep only letters and spaces
            normalized = re.sub(r'[^a-z\s]', '', label.lower())
            
            # Normalize whitespace
            normalized = re.sub(r'\s+', ' ', normalized).strip()
            
            return normalized
        
        def are_semantically_equivalent(term1: Dict, term2: Dict) -> bool:
            """Check if two terms have the same words (same or different order)."""
            label1 = term1.get('label', '')
            label2 = term2.get('label', '')
            
            norm1 = normalize_for_comparison(label1)
            norm2 = normalize_for_comparison(label2)
            
            # Exact match after normalization (same words, same order)
            if norm1 == norm2:
                return True
            
            # Same words, different order
            words1 = set(norm1.split())
            words2 = set(norm2.split())
            
            return words1 == words2 and len(words1) > 0
        
        # Group terms by semantic equivalence
        equivalence_groups = []
        
        for term in matched_terms:
            # Find if this term belongs to an existing group
            added_to_group = False
            
            for group in equivalence_groups:
                if any(are_semantically_equivalent(term, existing_term) for existing_term in group):
                    group.append(term)
                    added_to_group = True
                    break
            
            # If not added to any group, create a new group
            if not added_to_group:
                equivalence_groups.append([term])
        
        # For each group, select the best term based on source priority
        deduplicated_terms = []
        
        for group in equivalence_groups:
            if len(group) == 1:
                # No duplicates, keep the term
                deduplicated_terms.append(group[0])
            else:
                # Multiple equivalent terms, choose based on priority
                best_term = min(group, key=lambda t: source_priority.get(t.get('source', ''), 999))
                deduplicated_terms.append(best_term)
                
                # Log the deduplication for debugging
                removed_terms = [t for t in group if t != best_term]
                removed_labels = [f"{t.get('label')} [{t.get('source')}]" for t in removed_terms]
                best_label = f"{best_term.get('label')} [{best_term.get('source')}]"

        return deduplicated_terms

    def match_selected_labels_to_original_terms(self, selected_labels: List[str], vocab_search_results: Dict[str, List[Dict]]) -> List[Dict]:
        """
        Improved matching that finds all semantically equivalent terms and applies source priority.
        This fixes the issue where LLM selects one source but we end up with a different source in output.
        """
        
        # Create a comprehensive mapping of all available terms
        all_available_terms = []
        for topic, terms in vocab_search_results.items():
            for term in terms:
                if isinstance(term, dict):
                    all_available_terms.append(term)
        
        
        def normalize_for_comparison(label: str) -> str:
            """Normalize labels for semantic comparison - remove punctuation, lowercase, normalize spaces"""
            import re
            normalized = re.sub(r'[^a-z\s]', '', label.lower())
            normalized = re.sub(r'\s+', ' ', normalized).strip()
            return normalized
        
        matched_terms = []
        
        for selected_label in selected_labels:            
            # Find all terms that could semantically match this selected label
            candidate_matches = []
            
            # Normalize the selected label for comparison
            selected_words = set(normalize_for_comparison(selected_label).split())
            
            # Search through all available terms for semantic matches
            for term in all_available_terms:
                term_label = term.get('label', '').strip()
                term_words = set(normalize_for_comparison(term_label).split())
                
                # Check if they have the same words (order doesn't matter)
                if selected_words == term_words and len(selected_words) > 0:
                    candidate_matches.append(term)
                    
            
            # If we found semantic matches, apply source priority
            if candidate_matches:
                # Define source priority (lower number = higher priority)
                source_priority = {
                    'Getty AAT': 1,
                    'LCSH': 2, 
                    'FAST': 3,
                    'Getty TGN': 4
                }
                
                # Choose the best match based on source priority
                best_match = min(candidate_matches, key=lambda t: source_priority.get(t.get('source', ''), 999))
                matched_terms.append(best_match)
                
                if len(candidate_matches) > 1:
                    rejected_terms = [t for t in candidate_matches if t != best_match]
                    rejected_labels = [f"{t.get('label')} [{t.get('source')}]" for t in rejected_terms]

            else:
                # If no semantic matches, try exact string matching as fallback
                selected_normalized = selected_label.lower().strip()
                
                for term in all_available_terms:
                    term_label = term.get('label', '').strip()
                    term_normalized = term_label.lower().strip()
                    
                    if selected_normalized == term_normalized:
                        candidate_matches.append(term)
                
                if candidate_matches:
                    # Apply priority even for exact matches
                    best_match = min(candidate_matches, key=lambda t: source_priority.get(t.get('source', ''), 999))
                    matched_terms.append(best_match)
                
        seen_uris = set()
        deduplicated_terms = []
        for term in matched_terms:
            uri = term.get('uri', '')
            if uri and uri not in seen_uris:
                deduplicated_terms.append(term)
                seen_uris.add(uri)
            elif not uri:  # Keep terms without URIs but they're probably rare
                deduplicated_terms.append(term)
        
        return deduplicated_terms


    def update_json_data(self, selection_results: Dict[int, Dict[str, Any]]) -> bool:
        """Update JSON data with selected vocabulary terms only."""
        try:
            # Skip the last item if it's API stats
            data_items = self.json_data[:-1] if self.json_data and 'api_stats' in self.json_data[-1] else self.json_data
            api_stats = self.json_data[-1] if self.json_data and 'api_stats' in self.json_data[-1] else None
            
            updated_items = []
            
            for i, item in enumerate(data_items):
                if i in selection_results:
                    # Get selected term labels from LLM response
                    selected_term_responses = selection_results[i]['selection_result'].get('selected_terms', [])
                    
                    # Extract clean labels
                    selected_labels = []
                    for term in selected_term_responses:
                        if isinstance(term, dict):
                            label = term.get('label', '').strip()
                            selected_labels.append(label)
                    
                    # Match labels to full term objects from vocabulary_search_results
                    vocab_search_results = item['analysis'].get('vocabulary_search_results', {})
                    matched_terms = self.match_selected_labels_to_original_terms(selected_labels, vocab_search_results)
                    
                    # ONLY store final_selected_terms (remove all other redundant data)
                    item['analysis']['final_selected_terms'] = matched_terms
                else:
                    # Keep empty for entries without vocabulary terms
                    item['analysis']['final_selected_terms'] = []
                
                updated_items.append(item)
            
            # Add API stats back if it existed
            if api_stats:
                updated_items.append(api_stats)
            
            # Save updated JSON
            json_filename = f"{self.workflow_type}_workflow.json"
            metadata_dir = os.path.join(self.folder_path, "metadata", "collection_metadata")
            json_path = os.path.join(metadata_dir, json_filename)
            
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(updated_items, f, indent=2, ensure_ascii=False)
            
            print(f"Updated JSON file with selected vocabulary terms")
            return True
            
        except Exception as e:
            logging.error(f"Error updating JSON data: {e}")
            return False

    def update_excel_file(self, selection_results: Dict[int, Dict[str, Any]]) -> bool:
        """Update Excel file with selected vocabulary terms."""
        try:
            # Load the existing workbook
            wb = load_workbook(self.excel_path)
            analysis_sheet = wb['Analysis']
            
            # Find the Topic Vocabulary Terms column (from step 2)
            topic_vocab_col = None
            for col in range(1, analysis_sheet.max_column + 1):
                header_value = analysis_sheet.cell(row=1, column=col).value
                if header_value and "Topic Vocabulary Terms" in header_value:
                    topic_vocab_col = col
                    break
            
            if topic_vocab_col is None:
                logging.error("Topic Vocabulary Terms column not found. Please run step 2 first.")
                return False
            
            # Update header to reflect that these are selected terms
            header_cell = analysis_sheet.cell(row=1, column=topic_vocab_col)
            header_cell.value = "Selected Topic Vocabulary Terms"
            
            # Get the data for processing
            data_items = self.json_data[:-1] if self.json_data and 'api_stats' in self.json_data[-1] else self.json_data
            
            updated_rows = 0
            for entry_index, result_data in selection_results.items():
                row_num = entry_index + 2  # +2 for header row
                
                # Get final selected terms
                if entry_index < len(data_items):
                    selected_vocab_terms = data_items[entry_index]['analysis'].get('final_selected_terms', [])
                    
                    # Format for Excel display
                    if selected_vocab_terms:
                        formatted_terms = []
                        for term in selected_vocab_terms:
                            if isinstance(term, dict):
                                label = term.get('label', '')
                                uri = term.get('uri', '')
                                source = term.get('source', '')
                                formatted_terms.append(f"{label} ({uri}) [{source}]")
                        
                        cell_value = "; ".join(formatted_terms)
                        updated_rows += 1
                    else:
                        cell_value = ""
                else:
                    cell_value = ""
                
                # Set the cell value
                cell = analysis_sheet.cell(row=row_num, column=topic_vocab_col)
                cell.value = cell_value
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Clear vocabulary terms for entries that weren't processed
            for row_num in range(2, analysis_sheet.max_row + 1):
                entry_index = row_num - 2
                if entry_index not in selection_results:
                    cell = analysis_sheet.cell(row=row_num, column=topic_vocab_col)
                    cell.value = ""
            
            # Save the updated workbook
            wb.save(self.excel_path)
            print(f"Updated Excel file with selected vocabulary terms in {updated_rows} rows")
            return True
            
        except Exception as e:
            logging.error(f"Error updating Excel file: {e}")
            return False

    def create_vocabulary_mapping_report(self, selection_results: Dict[int, Dict[str, Any]]) -> bool:
        """Create vocabulary mapping report showing original terms organized by topic with checkmarks for chosen ones."""
        try:
            # Save vocabulary report in the collection_metadata folder
            metadata_dir = os.path.join(self.folder_path, "metadata", "collection_metadata")
            report_path = os.path.join(metadata_dir, "vocabulary_mapping_report.txt")
            
            # Skip the last item if it's API stats
            data_items = self.json_data[:-1] if self.json_data and 'api_stats' in self.json_data[-1] else self.json_data
            
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write("SOUTHERN ARCHITECT VOCABULARY MAPPING REPORT\n")
                f.write("=" * 50 + "\n\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Workflow Type: {self.workflow_type.upper()}\n\n")
                
                # Process each page
                for i, item in enumerate(data_items):
                    folder = item.get('folder', 'Unknown')
                    page_number = item.get('page_number', 'Unknown')
                    
                    f.write(f"PAGE {page_number} (ISSUE: {folder}):\n")
                    f.write("=" * (len(f"PAGE {page_number} (ISSUE: {folder}):")) + "\n")
                    
                    # Get vocabulary search results, geographic vocabulary results, and final selected terms
                    vocab_search_results = item['analysis'].get('vocabulary_search_results', {})
                    geo_vocab_results = item['analysis'].get('geographic_vocabulary_search_results', {})
                    selected_terms = item['analysis'].get('final_selected_terms', [])
                    
                    # Create set of selected URIs for easy lookup
                    selected_uris = set()
                    if selected_terms:
                        for term in selected_terms:
                            if isinstance(term, dict):
                                uri = term.get('uri', '')
                                if uri:
                                    selected_uris.add(uri)
                    
                    # Show topics and geographic entities on this page
                    topics_list = list(vocab_search_results.keys()) if vocab_search_results else []
                    geo_entities_list = list(geo_vocab_results.keys()) if geo_vocab_results else []
                    
                    if topics_list:
                        f.write(f"Topics on this page ({len(topics_list)}): {', '.join(topics_list)}\n")
                    if geo_entities_list:
                        f.write(f"Geographic entities on this page ({len(geo_entities_list)}): {', '.join(geo_entities_list)}\n")
                    
                    if vocab_search_results:
                        # Count total topics and terms
                        total_topics = len(vocab_search_results)
                        total_terms = sum(len(terms) for terms in vocab_search_results.values())
                        
                        # Show vocabulary terms organized by topic (ONLY TOPICS, NOT GEOGRAPHIC)
                        f.write(f"\nTOPIC VOCABULARY TERMS:\n")
                        selected_count = 0
                        topics_with_selections = 0
                        
                        for topic, terms in vocab_search_results.items():
                            f.write(f"  Topic: {topic}\n")
                            
                            if terms:
                                # Build terms list with checkmarks using URI matching
                                topic_terms = []
                                topic_has_selection = False
                                
                                for term in terms:
                                    if isinstance(term, dict):
                                        label = term.get('label', '')
                                        source = term.get('source', '')
                                        uri = term.get('uri', '')
                                        
                                        # Use URI matching
                                        is_selected = uri in selected_uris
                                        
                                        if is_selected:
                                            selected_count += 1
                                            topic_has_selection = True
                                            topic_terms.append(f"{label} ({uri}) [{source}] ✓")
                                        else:
                                            topic_terms.append(f"{label} ({uri}) [{source}]")

                                f.write(f"    Terms: {'; '.join(topic_terms)}\n")
                                
                                if topic_has_selection:
                                    topics_with_selections += 1
                            else:
                                f.write(f"    Terms: No terms available\n")
                            
                            f.write("\n")
                        
                        # Summary counts with list of chosen terms
                        f.write(f"TOPIC SUMMARY: {selected_count} terms selected from {topics_with_selections}/{total_topics} topics ({total_terms} total terms available)\n")
                        
                        # Add list of chosen terms if any were selected
                        if selected_count > 0:
                            f.write("CHOSEN TOPIC TERMS:\n")
                            for term in selected_terms:
                                if isinstance(term, dict):
                                    label = term.get('label', '')
                                    uri = term.get('uri', '')
                                    source = term.get('source', '')
                                    f.write(f"  - {label} ({uri}) [{source}]\n")
                            f.write("\n")
                    
                    # Show geographic vocabulary terms (NOT selected by LLM, just for display)
                    if geo_vocab_results:
                        f.write(f"GEOGRAPHIC VOCABULARY TERMS (not selected by LLM):\n")
                        for entity, terms in geo_vocab_results.items():
                            f.write(f"  Geographic Entity: {entity}\n")
                            
                            if terms:
                                geo_terms = []
                                for term in terms:
                                    if isinstance(term, dict):
                                        label = term.get('label', '')
                                        source = term.get('source', '')
                                        uri = term.get('uri', '')
                                        geo_terms.append(f"{label} ({uri}) [{source}]")
                                f.write(f"    Terms: {'; '.join(geo_terms)}\n")
                            else:
                                f.write(f"    Terms: No terms available\n")
                            f.write("\n")
                    
                    if not vocab_search_results and not geo_vocab_results:
                        f.write("No vocabulary terms available for this page.\n")
                    
                    f.write("\n" + "=" * 50 + "\n\n")
            
            print(f"Created vocabulary mapping report: {report_path}")
            return True
            
        except Exception as e:
            logging.error(f"Error creating vocabulary mapping report: {e}")
            return False

    def create_page_metadata_files(self) -> bool:
        """Create individual page metadata files with geographic entities included."""
        try:
            # Create page_metadata folder inside metadata
            output_folder = os.path.join(self.folder_path, "metadata", "page_metadata")
            os.makedirs(output_folder, exist_ok=True)
            
            # Skip the last item if it's API stats
            data_items = self.json_data[:-1] if self.json_data and 'api_stats' in self.json_data[-1] else self.json_data
            
            for entry in data_items:
                analysis = entry.get('analysis', {})
                folder_name = entry.get('folder', 'unknown')
                page_number = entry.get('page_number', 0)
                
                # Generate clean filename
                clean_folder = "".join(c for c in folder_name if c.isalnum() or c in ('-', '_')).strip()
                if not clean_folder:
                    clean_folder = "unknown"
                filename = f"{clean_folder}_page{page_number:03d}_metadata.txt"
                file_path = os.path.join(output_folder, filename)
                
                # Generate metadata content
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write("=" * 60 + "\n")
                    f.write("SOUTHERN ARCHITECT - PAGE METADATA\n")
                    f.write("=" * 60 + "\n\n")
                    
                    # Page identification
                    f.write("PAGE IDENTIFICATION:\n")
                    f.write(f"Folder: {folder_name}\n")
                    f.write(f"Page Number: {page_number}\n\n")
                    
                    # Content sections
                    f.write("CONTENT:\n")
                    f.write("-" * 30 + "\n")
                    
                    if self.workflow_type == 'text':
                        text_content = analysis.get('cleaned_text', '').strip()
                        if text_content:
                            f.write("Cleaned OCR Text:\n")
                            f.write(text_content + "\n\n")
                    else:
                        # Image workflow
                        text_transcription = analysis.get('text_transcription', '').strip()
                        if text_transcription:
                            f.write("Text Transcription:\n")
                            f.write(text_transcription + "\n\n")
                        
                        visual_description = analysis.get('visual_description', '').strip()
                        if visual_description:
                            f.write("Visual Description:\n")
                            f.write(visual_description + "\n\n")
                    
                    # TOC entry
                    toc_entry = analysis.get('toc_entry', '').strip()
                    if toc_entry:
                        f.write("Description:\n")
                        f.write(toc_entry + "\n\n")
                    
                    # Metadata
                    f.write("METADATA:\n")
                    f.write("-" * 30 + "\n")
                    
                    # Topics
                    topics = analysis.get('topics', [])
                    if isinstance(topics, str):
                        topics = [s.strip() for s in topics.split(',') if s.strip()]
                    if topics:
                        f.write("Topics:\n")
                        for topic in topics:
                            f.write(f"  - {topic}\n")
                        f.write("\n")

                    # Named Entities
                    named_entities = analysis.get('named_entities', [])
                    if isinstance(named_entities, str):
                        # Handle comma-separated string format
                        named_entities = [s.strip() for s in named_entities.split(',') if s.strip()]
                    if named_entities:
                        f.write("Named Entities:\n")
                        for entity in named_entities:
                            f.write(f"  - {entity}\n")
                        f.write("\n")

                    # ADD: Geographic Entities (from step 1)
                    geographic_entities = analysis.get('geographic_entities', [])
                    if isinstance(geographic_entities, str):
                        # Handle comma-separated string format
                        geographic_entities = [s.strip() for s in geographic_entities.split(',') if s.strip()]
                    if geographic_entities:
                        f.write("Geographic Entities:\n")
                        for entity in geographic_entities:
                            f.write(f"  - {entity}\n")
                        f.write("\n")

                    # Subject headings - use final_selected_terms only (LLM selected topic terms)
                    vocab_terms = analysis.get('final_selected_terms', [])
                    if vocab_terms:
                        f.write("Subject Headings (Selected):\n")
                        for term in vocab_terms:
                            if isinstance(term, dict):
                                label = term.get('label', '')
                                uri = term.get('uri', '')
                                source = term.get('source', '')
                                f.write(f"  - {label} ({uri}) [{source}]\n")
                        f.write("\n")
                    
                    # Geographic Subject Headings (from step 2)
                    geo_vocab_results = analysis.get('geographic_vocabulary_search_results', {})
                    if geo_vocab_results:
                        f.write("Geographic Subject Headings (from vocabulary lookup):\n")
                        for entity, terms in geo_vocab_results.items():
                            if terms:
                                for term in terms:
                                    if isinstance(term, dict):
                                        label = term.get('label', '')
                                        uri = term.get('uri', '')
                                        source = term.get('source', '')
                                        f.write(f"  - {label} ({uri}) [{source}]\n")
                        f.write("\n")
                    
                    # Content Warning if it exists
                    content_warning = analysis.get('content_warning', '').strip()
                    if content_warning and content_warning.lower() != 'none':
                        f.write("Content Warning:\n")
                        f.write(f"  {content_warning}\n\n")
                    
                    f.write("=" * 60 + "\n")
            
            print(f"Created page metadata files in: {output_folder}")
            return True
            
        except Exception as e:
            logging.error(f"Error creating page metadata files: {e}")
            return False

    def create_issue_content_index(self) -> bool:
        """Create separate issue content indexes for each unique folder/issue."""
        try:
            # Skip API stats for processing
            data_items = self.json_data[:-1] if self.json_data and 'api_stats' in self.json_data[-1] else self.json_data
            
            if not data_items:
                print("No data items found for issue content index")
                return False
            
            # Group entries by folder/issue
            issues = {}
            for entry in data_items:
                folder_name = entry.get('folder', 'Unknown')
                if folder_name not in issues:
                    issues[folder_name] = []
                issues[folder_name].append(entry)
            
            print(f"Found {len(issues)} unique issues: {list(issues.keys())}")
            
            # Create issue_metadata folder inside metadata
            issue_metadata_dir = os.path.join(self.folder_path, "metadata", "issue_metadata")
            os.makedirs(issue_metadata_dir, exist_ok=True)

            # Create separate index file for each issue
            for folder_name, entries in issues.items():
                # Create filename and path
                toc_filename = f"{folder_name}_Issue_Content_Index.txt"
                toc_path = os.path.join(issue_metadata_dir, toc_filename)
                
                with open(toc_path, 'w', encoding='utf-8') as f:
                    f.write(f"ISSUE CONTENT INDEX: {folder_name}\n")
                    f.write("=" * (len(f"ISSUE CONTENT INDEX: {folder_name}")) + "\n\n")
                    
                    # Sort entries by page number for logical ordering
                    sorted_entries = sorted(entries, key=lambda x: x.get('page_number', 0))
                    
                    for entry in sorted_entries:
                        analysis = entry.get('analysis', {})
                        page_number = entry.get('page_number', 'Unknown')
                        
                        # TOC entry (summary)
                        toc_entry = analysis.get('toc_entry', '').strip()
                        if not toc_entry or toc_entry.lower() == '[no toc entry]':
                            toc_entry = "[No summary available]"
                        
                        f.write(f"Page {page_number} (Issue: {folder_name}):\n\n{toc_entry}\n\n")
                        
                        # Topics
                        topics = analysis.get('topics', [])
                        if isinstance(topics, str):
                            topics = [s.strip() for s in topics.split(',') if s.strip()]
                        if topics:
                            f.write("Topics:\n")
                            for topic in topics:
                                f.write(f"  - {topic}\n")
                            f.write("\n")
                        
                        # Named Entities
                        named_entities = analysis.get('named_entities', [])
                        if isinstance(named_entities, str):
                            # Handle comma-separated string format
                            named_entities = [s.strip() for s in named_entities.split(',') if s.strip()]
                        if named_entities:
                            f.write("Named Entities:\n")
                            for entity in named_entities:
                                f.write(f"  - {entity}\n")
                            f.write("\n")

                        # Geographic Entities
                        geographic_entities = analysis.get('geographic_entities', [])
                        if isinstance(geographic_entities, str):
                            # Handle comma-separated string format
                            geographic_entities = [s.strip() for s in geographic_entities.split(',') if s.strip()]
                        if geographic_entities:
                            f.write("Geographic Entities:\n")
                            for entity in geographic_entities:
                                f.write(f"  - {entity}\n")
                            f.write("\n")

                        # Subject headings - use final_selected_terms only (LLM selected)
                        vocabulary_terms = analysis.get('final_selected_terms', [])
                        if vocabulary_terms:
                            f.write("Subject Headings (Selected):\n")
                            for term in vocabulary_terms:
                                if isinstance(term, dict):
                                    label = term.get('label', '')
                                    uri = term.get('uri', '')
                                    source = term.get('source', '')
                                    f.write(f"  - {label} ({uri}) [{source}]\n")
                            f.write("\n")
                        
                        # Geographic Subject Headings 
                        geo_vocab_results = analysis.get('geographic_vocabulary_search_results', {})
                        if geo_vocab_results:
                            f.write("Geographic Subject Headings:\n")
                            for entity, terms in geo_vocab_results.items():
                                if terms:
                                    for term in terms:
                                        if isinstance(term, dict):
                                            label = term.get('label', '')
                                            uri = term.get('uri', '')
                                            source = term.get('source', '')
                                            f.write(f"  - {label} ({uri}) [{source}]\n")
                            f.write("\n")
                        
                        # Content Warning if present
                        content_warning = analysis.get('content_warning', '').strip()
                        if content_warning and content_warning.lower() != 'none':
                            f.write(f"Content Warning:\n")
                            f.write(f"  {content_warning}\n")
                        
                        f.write("\n" + "-" * 50 + "\n\n")
                
                print(f"Created issue content index: {toc_path} ({len(entries)} pages)")
            
            return True
            
        except Exception as e:
            logging.error(f"Error creating issue content index: {e}")
            return False
    
    def run(self) -> bool:
        """Main execution method."""
        print(f"\nSOUTHERN ARCHITECT STEP 3 - VOCABULARY SELECTION")
        print(f"Processing folder: {self.folder_path}")
        print(f"Model: {self.model_name}")
        print(f"Note: Geographic entities are included in outputs but NOT sent to LLM for selection")
        print("-" * 50)
        
        # Detect workflow type
        if not self.detect_workflow_type():
            return False
        
        print(f"Detected workflow type: {self.workflow_type.upper()}")
        
        # Load JSON data
        if not self.load_json_data():
            return False
        
        # Find entries with vocabulary terms
        entries_with_vocab = self.find_entries_with_vocabulary()
        if not entries_with_vocab:
            print("No entries with vocabulary terms found")
            return False
        
        print(f"Found {len(entries_with_vocab)} entries with vocabulary terms")
        
        # Show model pricing info
        model_info = get_model_info(self.model_name)
        if model_info:
            print(f"Pricing: ${model_info['input_per_1k']:.5f}/1K input, ${model_info['output_per_1k']:.5f}/1K output")
        
        # Process vocabulary selection
        print(f"\nSelecting best topic vocabulary terms for each page...")
        print(f"Geographic entities will be preserved but not processed by LLM")
        selection_results = self.process_vocabulary_selection(entries_with_vocab)
        
        if not selection_results:
            print("Vocabulary selection failed")
            return False
        
        # Update JSON data with selected terms only
        if not self.update_json_data(selection_results):
            return False
        
        # Update Excel file with selected terms only
        if not self.update_excel_file(selection_results):
            return False
        
        # Create clean vocabulary mapping report
        if not self.create_vocabulary_mapping_report(selection_results):
            return False
        
        # Create page metadata files (with geographic entities)
        if not self.create_page_metadata_files():
            return False

        # Create issue content index (with geographic entities)
        if not self.create_issue_content_index():
            return False
        
        # Calculate and log final metrics
        total_processing_time = sum(result.get('processing_time', 0) for result in selection_results.values())
        
        # Calculate cost
        estimated_cost = calculate_cost(
            model_name=self.model_name,
            prompt_tokens=api_stats.total_input_tokens,
            completion_tokens=api_stats.total_output_tokens,
            is_batch=False
        )
        
        # Create logs folder and token usage log
        logs_folder_path = os.path.join(self.folder_path, "logs")
        if not os.path.exists(logs_folder_path):
            os.makedirs(logs_folder_path)
        
        create_token_usage_log(
        logs_folder_path=logs_folder_path,
        script_name="southern_architect_vocabulary_selection",
        model_name=self.model_name,
        total_items=len(selection_results),
        items_with_issues=0,
        total_time=total_processing_time,
        total_prompt_tokens=api_stats.total_input_tokens,
        total_completion_tokens=api_stats.total_output_tokens,
        additional_metrics={
            "Processing mode": "BATCH" if self.was_batch_processed else "INDIVIDUAL",  
            "Actual cost": f"${estimated_cost:.4f}",
            "Average tokens per entry": f"{(api_stats.total_input_tokens + api_stats.total_output_tokens)/len(selection_results):.0f}" if selection_results else "0",
            "Batch processing used": "Yes" if self.was_batch_processed else "No"  
        }
    )
        
        # Show final summary
        total_selected = sum(len(result['selection_result'].get('selected_terms', [])) for result in selection_results.values())
        entries_with_selections = sum(1 for result in selection_results.values() if result['selection_result'].get('selected_terms'))
        
        print("\n" + "=" * 50)
        print(f"FINAL SUMMARY:")
        print(f"\n ✅ STEP 3 COMPLETE: Selected vocabulary terms in {os.path.basename(self.folder_path)}")
        print(f"Page metadata files, issue indexes, updated Excel/JSON, and vocabulary report created")
        print(f"Entries processed: {len(selection_results)}")
        print(f"Total vocabulary terms selected: {total_selected}")
        print(f"Entries with selections: {entries_with_selections}/{len(selection_results)}")
        print(f"Selection rate: {(entries_with_selections/len(selection_results)*100):.1f}%")
        print(f"Total tokens: {api_stats.total_input_tokens + api_stats.total_output_tokens:,}")
        print(f"Estimated cost: ${estimated_cost:.4f}")

        # Show which issue content index files were created
        data_items = self.json_data[:-1] if self.json_data and 'api_stats' in self.json_data[-1] else self.json_data
        unique_issues = set(entry.get('folder', 'Unknown') for entry in data_items)
        issue_metadata_dir = os.path.join(self.folder_path, "metadata", "issue_metadata")
        for issue in sorted(unique_issues):
            toc_filename = f"{issue}_Issue_Content_Index.txt"
            print(f"  Issue index: {os.path.join(issue_metadata_dir, toc_filename)}")

        return True

def main():
    
    # Default base directory for Southern Architect output folders
    # Get script directory and build path to output folders
    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_output_dir = os.path.join(script_dir, "output_folders")

    model_name = os.getenv('MODEL_NAME', DEFAULT_MODEL)

    # Default folder path (newest folder if not specified)
    folder_path = find_newest_folder(base_output_dir)
    if not folder_path:
        print(f"No folders found in: {base_output_dir}")
        return 1
    print(f"Auto-selected newest folder: {os.path.basename(folder_path)}")


    # Create and run the processor
    processor = SouthernArchitectVocabularyProcessor(folder_path, model_name)
    success = processor.run()
    
    if not success:
        print("Vocabulary selection failed")
        return 1
    
    return 0

if __name__ == "__main__":
    exit(main())